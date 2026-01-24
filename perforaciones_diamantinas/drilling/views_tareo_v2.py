"""
=============================================================================
VISTAS PARA TAREO V2 - Modelo Normalizado con Pivot en Frontend
=============================================================================

Este módulo implementa la visualización tipo matriz (Excel) sobre el modelo
vertical AsistenciaDiaria, utilizando formsets de Django para edición masiva.

Arquitectura:
- Backend: Datos verticales (empleado, fecha, estado)
- Frontend: Transformación a matriz horizontal (estilo Excel)
- Formularios: ModelFormSet para edición masiva eficiente

Autor: Sistema DrillControl
Fecha: Enero 2026
=============================================================================
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.forms import modelformset_factory, ModelForm, Textarea
from datetime import datetime, timedelta, date
from calendar import monthrange
import json
import logging

from .models import Contrato, Trabajador, AsistenciaDiaria, CierreMensualTareo, HistorialCambioAsistencia, Maquina
from .utils.tareo_service import TareoService, CierreMensualService
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)


# =============================================================================
# FORMULARIO PARA ASISTENCIA DIARIA
# =============================================================================
class AsistenciaDiariaForm(ModelForm):
    """
    Formulario individual para cada celda de la matriz de asistencia.
    Incluye validaciones y lógica de negocio específica.
    """
    class Meta:
        model = AsistenciaDiaria
        fields = ['empleado', 'fecha', 'estado', 'observaciones']
        widgets = {
            'observaciones': Textarea(attrs={'class': 'form-control', 'rows': 1}),
        }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Deshabilitar campos que no deben editarse directamente
        self.fields['empleado'].widget.attrs['readonly'] = True
        self.fields['fecha'].widget.attrs['readonly'] = True


# =============================================================================
# VISTA PRINCIPAL: TAREO MENSUAL CON MATRIZ
# =============================================================================
@login_required
def tareo_v2_mensual_view(request):
    """
    Vista principal del Tareo V2 con arquitectura normalizada.
    
    Flujo:
    1. GET: Consulta datos verticales y los transforma a matriz para el template
    2. POST: Recibe formset, valida y actualiza masivamente en BD vertical
    
    Features:
    - Proyección automática mensual
    - Edición masiva con formsets
    - Renderizado eficiente para 70+ empleados
    - Navegación por meses
    """
    user = request.user
    
    # =========================================================================
    # 1. VALIDACIÓN DE PERMISOS
    # =========================================================================
    if not user.can_manage_contract_users():
        messages.error(request, 'No tienes permisos para gestionar el tareo de asistencia')
        return redirect('dashboard')
    
    # =========================================================================
    # 2. DETERMINAR CONTRATO
    # =========================================================================
    if user.has_access_to_all_contracts():
        contrato_id = request.GET.get('contrato')
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO')
        else:
            contrato = Contrato.objects.filter(estado='ACTIVO').first()
        contratos_disponibles = Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato')
    else:
        contrato = user.contrato
        contratos_disponibles = None
    
    if not contrato:
        messages.warning(request, 'No hay contratos activos disponibles')
        return redirect('dashboard')
    
    # =========================================================================
    # 3. CALCULAR RANGO DE FECHAS (MES COMPLETO)
    # =========================================================================
    mes_offset = int(request.GET.get('mes_offset', 0))
    
    # Calcular el mes a mostrar
    hoy = datetime.now().date()
    fecha_base = hoy
    
    # Navegación por meses
    if mes_offset != 0:
        for _ in range(abs(mes_offset)):
            if mes_offset > 0:
                # Mes siguiente
                if fecha_base.month == 12:
                    fecha_base = fecha_base.replace(year=fecha_base.year + 1, month=1, day=1)
                else:
                    fecha_base = fecha_base.replace(month=fecha_base.month + 1, day=1)
            else:
                # Mes anterior
                if fecha_base.month == 1:
                    fecha_base = fecha_base.replace(year=fecha_base.year - 1, month=12, day=1)
                else:
                    fecha_base = fecha_base.replace(month=fecha_base.month - 1, day=1)
    
    # Mes operativo: del 26 del mes anterior al 25 del mes actual
    # Enero 2026 operativo = 26/12/2025 al 25/01/2026
    mes_anterior = fecha_base.month - 1 if fecha_base.month > 1 else 12
    anio_anterior = fecha_base.year if fecha_base.month > 1 else fecha_base.year - 1
    
    fecha_inicio = date(anio_anterior, mes_anterior, 26)
    fecha_fin = date(fecha_base.year, fecha_base.month, 25)
    
    # Nombre del período para mostrar (mes operativo)
    meses_es = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    nombre_periodo = f"{meses_es[fecha_base.month]} {fecha_base.year}"
    
    # =========================================================================
    # 4. GENERAR LISTA DE DÍAS DEL MES
    # =========================================================================
    dias_rango = []
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        nombres_dias = {
            0: 'Lun', 1: 'Mar', 2: 'Mié', 3: 'Jue', 
            4: 'Vie', 5: 'Sáb', 6: 'Dom'
        }
        dias_rango.append({
            'fecha': fecha_actual,
            'dia': fecha_actual.day,
            'nombre_dia': nombres_dias[fecha_actual.weekday()],
            'es_domingo': fecha_actual.weekday() == 6,
            'es_sabado': fecha_actual.weekday() == 5,
        })
        fecha_actual += timedelta(days=1)
    
    # =========================================================================
    # 5. PROCESAMIENTO POST (GUARDAR CAMBIOS)
    # =========================================================================
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Parsear datos del formulario
                asistencias_data = []
                
                for key, value in request.POST.items():
                    if key.startswith('estado_'):
                        # Formato: estado_trabajadorID_YYYY-MM-DD
                        parts = key.split('_')
                        if len(parts) == 4:
                            _, trabajador_id, anio_str, mes_dia_str = parts
                            fecha_str = f"{anio_str}-{mes_dia_str}"
                            
                            try:
                                fecha_asistencia = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                                observaciones_key = f"observaciones_{trabajador_id}_{anio_str}_{mes_dia_str}"
                                observaciones = request.POST.get(observaciones_key, '')
                                
                                asistencias_data.append({
                                    'empleado_id': int(trabajador_id),
                                    'fecha': fecha_asistencia,
                                    'estado': value,
                                    'observaciones': observaciones
                                })
                            except ValueError:
                                continue
                
                # Actualización masiva usando el servicio
                resultado = TareoService.actualizar_masivo_desde_formset(
                    asistencias_data,
                    user
                )
                
                # Mensajes de resultado
                if resultado['errores']:
                    messages.warning(
                        request,
                        f"Guardado parcialmente. Errores: {len(resultado['errores'])}"
                    )
                else:
                    messages.success(
                        request,
                        f"Tareo guardado exitosamente. "
                        f"Actualizados: {resultado['actualizados']}, "
                        f"Creados: {resultado['creados']}"
                    )
                
                # Redirigir para evitar reenvío de formulario
                return redirect(f"{request.path}?contrato={contrato.id}&mes_offset={mes_offset}")
                
        except Exception as e:
            messages.error(request, f"Error al guardar: {str(e)}")
    
    # =========================================================================
    # 6. OBTENER DATOS PARA VISUALIZACIÓN (GET)
    # =========================================================================
    # Usar el servicio para obtener matriz pivoteada
    matriz_tareo = TareoService.obtener_matriz_tareo(contrato, fecha_inicio, fecha_fin)
    
    # Obtener máquinas activas del contrato para los selects
    maquinas_disponibles = Maquina.objects.filter(
        contrato=contrato,
        estado='ACTIVO'
    ).order_by('nombre')
    
    # =========================================================================
    # 7. CONTEXTO PARA EL TEMPLATE
    # =========================================================================
    context = {
        'contrato': contrato,
        'contratos_disponibles': contratos_disponibles,
        'nombre_periodo': nombre_periodo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'dias_rango': dias_rango,
        'matriz_tareo': matriz_tareo,
        'mes_offset': mes_offset,
        'estados_choices': AsistenciaDiaria.ESTADO_CHOICES,
        'maquinas_disponibles': maquinas_disponibles,
        'mes_actual': fecha_inicio.month,
        'anio_actual': fecha_inicio.year,
    }
    
    return render(request, 'drilling/tareo/tareo_v2_mensual.html', context)


# =============================================================================
# API PARA PROYECCIÓN MENSUAL (AJAX)
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_generar_proyeccion(request):
    """
    Endpoint AJAX para generar proyección mensual automática.
    
    POST params:
        - contrato_id: ID del contrato
        - anio: Año de la proyección
        - mes: Mes de la proyección
        - sobrescribir: Boolean, si True elimina proyecciones previas
    
    Returns:
        JSON con estadísticas de la operación
    """
    try:
        # Validar permisos
        if not request.user.can_manage_contract_users():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para esta operación'
            }, status=403)
        
        # Obtener parámetros
        contrato_id = request.POST.get('contrato_id')
        anio = int(request.POST.get('anio'))
        mes = int(request.POST.get('mes'))
        sobrescribir = request.POST.get('sobrescribir', 'false').lower() == 'true'
        
        # Validar contrato
        contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO')
        
        # Ejecutar proyección
        resultado = TareoService.generar_proyeccion_mensual(
            anio=anio,
            mes=mes,
            contrato=contrato,
            sobrescribir=sobrescribir
        )
        
        return JsonResponse({
            'success': True,
            'data': resultado
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# API PARA CORRECCIÓN INDIVIDUAL (AJAX)
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_corregir_asistencia(request):
    """
    Endpoint AJAX para corrección individual de una asistencia.
    
    POST params (JSON):
        - empleado_id: ID del trabajador
        - fecha: Fecha en formato YYYY-MM-DD
        - estado: Nuevo estado
        - observaciones: Observaciones opcionales
    
    Returns:
        JSON con el registro actualizado
    """
    try:
        # Validar permisos
        if not request.user.can_manage_contract_users():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para esta operación'
            }, status=403)
        
        # Parsear datos JSON
        data = json.loads(request.body)
        empleado_id = data.get('empleado_id')
        fecha_str = data.get('fecha')
        estado = data.get('estado')
        observaciones = data.get('observaciones', '')
        
        # Validar fecha
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Ejecutar corrección
        asistencia = TareoService.corregir_asistencia(
            empleado_id=empleado_id,
            fecha=fecha,
            nuevo_estado=estado,
            usuario=request.user,
            observaciones=observaciones
        )
        
        return JsonResponse({
            'success': True,
            'data': {
                'id': asistencia.id,
                'estado': asistencia.estado,
                'estado_display': asistencia.get_estado_display(),
                'es_proyeccion': asistencia.es_proyeccion
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# API PARA GUARDAR ASISTENCIA POR DÍA (AJAX)
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_guardar_dia_tareo(request):
    """
    Endpoint AJAX para guardar la asistencia completa de un día específico.
    
    POST params (JSON):
        - fecha: Fecha en formato YYYY-MM-DD
        - contrato_id: ID del contrato
        - asistencias: Lista de objetos con:
            - empleado_id: ID del trabajador
            - estado: Estado de asistencia
            - maquina_id: ID de máquina asignada (opcional)
            - observaciones: Observaciones opcionales
    
    Returns:
        JSON con estadísticas de la operación
    """
    try:
        # Validar permisos
        if not request.user.can_manage_contract_users():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para esta operación'
            }, status=403)
        
        # Parsear datos JSON
        data = json.loads(request.body)
        fecha_str = data.get('fecha')
        contrato_id = data.get('contrato_id')
        asistencias_data = data.get('asistencias', [])
        
        # Validar fecha
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Validar contrato
        contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO')
        
        stats = {
            'actualizados': 0,
            'creados': 0,
            'errores': []
        }
        
        # Procesar cada asistencia
        with transaction.atomic():
            for asist_data in asistencias_data:
                try:
                    empleado_id = asist_data.get('empleado_id')
                    estado = asist_data.get('estado')
                    maquina_id = asist_data.get('maquina_id')
                    observaciones = asist_data.get('observaciones', '')
                    
                    if not empleado_id or not estado:
                        continue
                    
                    # Obtener trabajador
                    trabajador = Trabajador.objects.get(id=empleado_id, contrato=contrato)
                    
                    # Obtener máquina si se especificó
                    maquina = None
                    if maquina_id and maquina_id != '':
                        try:
                            maquina = Maquina.objects.get(id=maquina_id)
                        except Maquina.DoesNotExist:
                            pass
                    
                    # Actualizar o crear asistencia
                    asistencia, created = AsistenciaDiaria.objects.update_or_create(
                        empleado=trabajador,
                        fecha=fecha,
                        defaults={
                            'estado': estado,
                            'maquina_snapshot': maquina,
                            'observaciones': observaciones,
                            'es_proyeccion': False,
                            'registrado_por': request.user,
                            'guardia_snapshot': trabajador.guardia_asignada
                        }
                    )
                    
                    if created:
                        stats['creados'] += 1
                    else:
                        stats['actualizados'] += 1
                    
                except Exception as e:
                    stats['errores'].append(f"Error procesando empleado {empleado_id}: {str(e)}")
                    logger.error(f"Error guardando asistencia: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        logger.error(f"Error en api_guardar_dia_tareo: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# API PARA OBTENER MÁQUINAS DEL CONTRATO
# =============================================================================
@login_required
@require_http_methods(["GET"])
def api_obtener_maquinas(request):
    """
    Endpoint AJAX para obtener las máquinas activas de un contrato.
    
    GET params:
        - contrato_id: ID del contrato (opcional)
    
    Returns:
        JSON con lista de máquinas
    """
    try:
        contrato_id = request.GET.get('contrato_id')
        
        # Filtrar máquinas
        maquinas_query = Maquina.objects.filter(estado='ACTIVO')
        
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id)
            maquinas_query = maquinas_query.filter(contrato=contrato)
        
        # Serializar máquinas
        maquinas_data = [
            {
                'id': maq.id,
                'nombre': maq.nombre,
                'codigo': maq.codigo,
                'tipo': maq.tipo_maquina
            }
            for maq in maquinas_query.order_by('nombre')
        ]
        
        return JsonResponse({
            'success': True,
            'data': maquinas_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# VISTA DE ESTADÍSTICAS (DASHBOARD)
# =============================================================================
@login_required
def tareo_v2_estadisticas(request):
    """
    Vista de estadísticas y resumen del tareo mensual.
    
    Muestra:
    - Total de días trabajados vs proyectados
    - Ausentismos por tipo
    - Trabajadores por guardia
    - Gráficos de tendencia
    """
    user = request.user
    
    # Determinar contrato
    if user.has_access_to_all_contracts():
        contrato_id = request.GET.get('contrato')
        contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO') if contrato_id else None
        contratos_disponibles = Contrato.objects.filter(estado='ACTIVO')
    else:
        contrato = user.contrato
        contratos_disponibles = None
    
    if not contrato:
        messages.warning(request, 'Seleccione un contrato')
        return redirect('dashboard')
    
    # Calcular estadísticas del mes actual
    hoy = date.today()
    primer_dia_mes = hoy.replace(day=1)
    
    # Query de asistencias del mes
    asistencias_mes = AsistenciaDiaria.objects.filter(
        empleado__contrato=contrato,
        fecha__gte=primer_dia_mes,
        fecha__lte=hoy
    )
    
    # Estadísticas básicas
    total_registros = asistencias_mes.count()
    proyecciones = asistencias_mes.filter(es_proyeccion=True).count()
    correcciones = asistencias_mes.filter(es_proyeccion=False).count()
    
    # Por estado
    stats_por_estado = {}
    for estado_code, estado_label in AsistenciaDiaria.ESTADO_CHOICES:
        count = asistencias_mes.filter(estado=estado_code).count()
        if count > 0:
            stats_por_estado[estado_label] = count
    
    context = {
        'contrato': contrato,
        'contratos_disponibles': contratos_disponibles,
        'total_registros': total_registros,
        'proyecciones': proyecciones,
        'correcciones': correcciones,
        'stats_por_estado': stats_por_estado,
        'mes_actual': primer_dia_mes,
    }
    
    return render(request, 'drilling/tareo/tareo_v2_estadisticas.html', context)


# =============================================================================
# VISTAS DE CIERRE MENSUAL Y AUDITORÍA
# =============================================================================

@login_required
def tareo_cierre_mensual(request):
    """
    Vista para revisar y cerrar contablemente el mes.
    Muestra resumen completo antes del cierre.
    """
    # Determinar contrato
    if request.user.is_staff or request.user.is_superuser:
        contratos = Contrato.objects.all()
        contrato_id = request.GET.get('contrato')
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id)
        else:
            contrato = contratos.first()
    else:
        contrato = request.user.contrato
        contratos = Contrato.objects.filter(id=contrato.id)
    
    # Mes y año
    mes = int(request.GET.get('mes', date.today().month))
    anio = int(request.GET.get('anio', date.today().year))
    
    # Obtener resumen del mes
    resumen = CierreMensualService.obtener_resumen_mes(contrato, anio, mes)
    
    # Obtener estado del cierre
    try:
        cierre = CierreMensualTareo.objects.get(contrato=contrato, anio=anio, mes=mes)
    except CierreMensualTareo.DoesNotExist:
        cierre = None
    
    context = {
        'contrato': contrato,
        'contratos': contratos,
        'mes': mes,
        'anio': anio,
        'resumen': resumen,
        'cierre': cierre,
        'meses': range(1, 13),
        'anios': range(date.today().year - 1, date.today().year + 2),
    }
    
    return render(request, 'drilling/tareo/cierre_mensual.html', context)


@login_required
@require_http_methods(["POST"])
def api_cerrar_mes(request):
    """
    API para cerrar contablemente un mes.
    """
    from .utils.tareo_service import CierreMensualService
    
    try:
        contrato_id = request.POST.get('contrato_id')
        mes = int(request.POST.get('mes'))
        anio = int(request.POST.get('anio'))
        observaciones = request.POST.get('observaciones', '')
        
        contrato = get_object_or_404(Contrato, id=contrato_id)
        
        # Verificar permisos
        if not (request.user.is_staff or request.user.role in ['manager', 'admin']):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        # Cerrar mes
        resultado = CierreMensualService.cerrar_mes(
            contrato=contrato,
            anio=anio,
            mes=mes,
            usuario=request.user,
            observaciones=observaciones
        )
        
        if resultado['success']:
            return JsonResponse({
                'success': True,
                'mensaje': resultado['mensaje'],
                'cierre': {
                    'id': resultado['cierre'].id,
                    'estado': resultado['cierre'].estado,
                    'fecha_cierre': resultado['cierre'].fecha_cierre.strftime('%Y-%m-%d %H:%M'),
                    'total_trabajadores': resultado['cierre'].total_trabajadores,
                    'total_dias_trabajo': resultado['cierre'].total_dias_trabajo,
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'error': resultado['error']
            }, status=400)
            
    except Exception as e:
        logger.error(f"Error cerrando mes: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_reabrir_mes(request):
    """
    API para reabrir un mes cerrado (caso excepcional).
    """
    from .utils.tareo_service import CierreMensualService
    
    try:
        contrato_id = request.POST.get('contrato_id')
        mes = int(request.POST.get('mes'))
        anio = int(request.POST.get('anio'))
        motivo = request.POST.get('motivo', '')
        
        contrato = get_object_or_404(Contrato, id=contrato_id)
        
        # Solo admin puede reabrir
        if not request.user.is_staff:
            return JsonResponse({'error': 'Solo administradores pueden reabrir meses'}, status=403)
        
        resultado = CierreMensualService.reabrir_mes(
            contrato=contrato,
            anio=anio,
            mes=mes,
            usuario=request.user,
            motivo=motivo
        )
        
        if resultado['success']:
            return JsonResponse({
                'success': True,
                'mensaje': resultado['mensaje']
            })
        else:
            return JsonResponse({
                'success': False,
                'error': resultado['error']
            }, status=400)
            
    except Exception as e:
        logger.error(f"Error reabriendo mes: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def tareo_historial_trabajador(request, trabajador_id):
    """
    Vista de historial completo de cambios de un trabajador.
    Útil para auditorías y resolución de disputas.
    """
    from .utils.tareo_service import AuditoriaAsistenciaService
    from .models import HistorialCambioAsistencia
    
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    
    # Verificar permisos
    if not (request.user.is_staff or request.user.contrato == trabajador.contrato):
        messages.error(request, 'No tiene permisos para ver este historial')
        return redirect('dashboard')
    
    # Filtros opcionales
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # Obtener historial
    historial = AuditoriaAsistenciaService.obtener_historial_trabajador(
        trabajador=trabajador,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(historial, 50)  # 50 cambios por página
    page_number = request.GET.get('page')
    historial_paginado = paginator.get_page(page_number)
    
    context = {
        'trabajador': trabajador,
        'historial': historial_paginado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_cambios': historial.count(),
    }
    
    return render(request, 'drilling/tareo/historial_trabajador.html', context)


@login_required
def tareo_reporte_nomina(request):
    """
    Vista para generar reporte de nómina del mes cerrado.
    Solo muestra meses cerrados para garantizar integridad.
    """
    from .models import CierreMensualTareo, AsistenciaDiaria
    from calendar import monthrange
    
    # Determinar contrato
    if request.user.is_staff or request.user.is_superuser:
        contratos = Contrato.objects.all()
        contrato_id = request.GET.get('contrato')
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id)
        else:
            contrato = contratos.first()
    else:
        contrato = request.user.contrato
        contratos = Contrato.objects.filter(id=contrato.id)
    
    # Obtener meses cerrados
    cierres_disponibles = CierreMensualTareo.objects.filter(
        contrato=contrato,
        estado='CERRADO'
    ).order_by('-anio', '-mes')
    
    # Seleccionar cierre
    cierre_id = request.GET.get('cierre_id')
    if cierre_id:
        cierre = get_object_or_404(CierreMensualTareo, id=cierre_id)
    elif cierres_disponibles.exists():
        cierre = cierres_disponibles.first()
    else:
        cierre = None
    
    # Generar reporte si hay cierre seleccionado
    reporte_trabajadores = []
    if cierre:
        # Usar fechas del mes operativo
        primer_dia = cierre.get_fecha_inicio_periodo()
        ultimo_dia = cierre.get_fecha_fin_periodo()
        
        trabajadores = Trabajador.objects.filter(
            contrato=contrato,
            estado='ACTIVO'
        ).order_by('apellidos', 'nombres')
        
        for trabajador in trabajadores:
            asistencias = AsistenciaDiaria.objects.filter(
                empleado=trabajador,
                fecha__gte=primer_dia,
                fecha__lte=ultimo_dia,
                es_proyeccion=False  # Solo registros reales
            )
            
            dias_trabajo = asistencias.filter(estado='TRABAJO').count()
            dias_descanso = asistencias.filter(estado='DESCANSO').count()
            faltas = asistencias.filter(estado='FALTA').count()
            vacaciones = asistencias.filter(estado='VACACIONES').count()
            permisos = asistencias.filter(estado='PERMISO').count()
            dm = asistencias.filter(estado='DM').count()
            
            reporte_trabajadores.append({
                'trabajador': trabajador,
                'dias_trabajo': dias_trabajo,
                'dias_descanso': dias_descanso,
                'faltas': faltas,
                'vacaciones': vacaciones,
                'permisos': permisos,
                'descanso_medico': dm,
                'total_dias': asistencias.count(),
            })
    
    context = {
        'contrato': contrato,
        'contratos': contratos,
        'cierres_disponibles': cierres_disponibles,
        'cierre': cierre,
        'reporte_trabajadores': reporte_trabajadores,
    }
    
    return render(request, 'drilling/tareo/reporte_nomina.html', context)


@login_required
def api_exportar_nomina_excel(request, cierre_id):
    """
    Exporta el reporte de nómina a Excel.
    """
    from .models import CierreMensualTareo, AsistenciaDiaria
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from calendar import monthrange
    import io
    
    cierre = get_object_or_404(CierreMensualTareo, id=cierre_id)
    
    # Verificar permisos
    if not (request.user.is_staff or request.user.contrato == cierre.contrato):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Nómina {cierre.mes}-{cierre.anio}"
    
    # Encabezados
    headers = [
        'DNI', 'Apellidos', 'Nombres', 'Cargo', 'Régimen',
        'Días Trabajo', 'Días Descanso', 'Faltas', 'Vacaciones',
        'Permisos', 'DM', 'Total Días'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0470AC", end_color="0470AC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos - usar mes operativo
    primer_dia = cierre.get_fecha_inicio_periodo()
    ultimo_dia = cierre.get_fecha_fin_periodo()
    
    trabajadores = Trabajador.objects.filter(
        contrato=cierre.contrato,
        estado='ACTIVO'
    ).order_by('apellidos', 'nombres')
    
    row_num = 2
    for trabajador in trabajadores:
        asistencias = AsistenciaDiaria.objects.filter(
            empleado=trabajador,
            fecha__gte=primer_dia,
            fecha__lte=ultimo_dia,
            es_proyeccion=False
        )
        
        ws.cell(row=row_num, column=1, value=trabajador.dni)
        ws.cell(row=row_num, column=2, value=trabajador.apellidos)
        ws.cell(row=row_num, column=3, value=trabajador.nombres)
        ws.cell(row=row_num, column=4, value=trabajador.cargo.nombre if trabajador.cargo else '')
        ws.cell(row=row_num, column=5, value=trabajador.regimen_laboral or '')
        ws.cell(row=row_num, column=6, value=asistencias.filter(estado='TRABAJO').count())
        ws.cell(row=row_num, column=7, value=asistencias.filter(estado='DESCANSO').count())
        ws.cell(row=row_num, column=8, value=asistencias.filter(estado='FALTA').count())
        ws.cell(row=row_num, column=9, value=asistencias.filter(estado='VACACIONES').count())
        ws.cell(row=row_num, column=10, value=asistencias.filter(estado='PERMISO').count())
        ws.cell(row=row_num, column=11, value=asistencias.filter(estado='DM').count())
        ws.cell(row=row_num, column=12, value=asistencias.count())
        
        row_num += 1
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="nomina_{cierre.contrato.nombre_contrato}_{cierre.mes}_{cierre.anio}.xlsx"'
    
    return response
