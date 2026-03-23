"""
Vistas para el tareo de asistencia de trabajadores (V1 + V2 unificado)
Permite a los managers de contrato registrar la asistencia diaria.

Incluye:
- Tareo mensual estilo matriz (V1/legacy y V2 normalizado)
- Proyección automática por régimen laboral
- Cierre mensual auditable
- Export Excel con 3 hojas (Tareo, Leyenda, Informe)
- Historial de cambios y reporte de nómina
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.middleware.csrf import CsrfViewMiddleware
from django.db.models import Count
from django.forms import ModelForm, Textarea
from datetime import datetime, timedelta, date
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import locale
import logging
import io

from .models import (
    Contrato, Trabajador,
    AsistenciaTrabajador,          # Modelo V1 (legacy)
    Maquina,
    FechaCerrada,
)
from .tareo_compat import AsistenciaDiaria, CierreMensualTareo, HistorialCambioAsistencia, NEW_TAREO
from .utils.tareo_service import TareoService, CierreMensualService
from .utils.tareo_service import TareoEngine

logger = logging.getLogger(__name__)


# Compat helper: nombre de la FK en AsistenciaDiaria puede ser `empleado` (legacy) o `trabajador` (nuevo)
try:
    from .tareo_compat import NEW_TAREO
except Exception:
    NEW_TAREO = False


def _emp_field_name():
    # Detectar dinámicamente el nombre del FK en AsistenciaDiaria
    try:
        # Preferir `trabajador` si existe en el modelo
        AsistenciaDiaria._meta.get_field('trabajador')
        return 'trabajador'
    except Exception:
        try:
            AsistenciaDiaria._meta.get_field('empleado')
            return 'empleado'
        except Exception:
            # Fallback a la configuración previa si no se puede inspeccionar
            return 'trabajador' if NEW_TAREO else 'empleado'


def _asistencias_v2_qs(contrato, fecha_inicio, fecha_fin):
    """Intentar consultar AsistenciaDiaria usando el FK `trabajador` o `empleado`.
    Devuelve una tupla (QuerySet, usado_emp_field) donde `usado_emp_field`
    es 'trabajador' o 'empleado' si alguno funcionó, o (empty_qs, None) si falla.
    """
    from django.core.exceptions import FieldError
    from .tareo_compat import AsistenciaDiaria

    for field in ('trabajador', 'empleado'):
        try:
            qs = AsistenciaDiaria.objects.filter(**{f"{field}__contrato": contrato, 'fecha__gte': fecha_inicio, 'fecha__lte': fecha_fin}).select_related(field, 'maquina_snapshot')
            return qs, field
        except FieldError:
            logger.debug("AsistenciaDiaria: field %s not valid for model, trying next", field)
        except Exception:
            logger.exception("Error querying AsistenciaDiaria with field %s", field)

    return AsistenciaDiaria.objects.none(), None


def _trabajadores_tareo_qs(contrato, fecha_inicio=None, fecha_fin=None):
    from django.db.models import Q
    qs = Trabajador.objects.filter(contrato=contrato)
    if fecha_inicio or fecha_fin:
        # Trabajadores activos (estado ACTIVO) o inactivos que trabajaron en el rango
        activos = Q(estado="ACTIVO")
        inactivos_en_rango = (
            ~Q(estado="ACTIVO") &
            (
                (Q(fecha_inicio_labores__isnull=True) | Q(fecha_inicio_labores__lte=fecha_fin)) &
                (Q(fecha_cese__isnull=True) | Q(fecha_cese__gte=fecha_inicio))
            )
        )
        qs = qs.filter(activos | inactivos_en_rango)
    return qs


def _trabajador_field_name():
    return 'trabajador'


def _emp_field_name():
    return _trabajador_field_name()


def _asistencias_v2_qs(contrato, fecha_inicio, fecha_fin):
    qs = AsistenciaDiaria.objects.filter(
        trabajador__contrato=contrato,
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin,
    ).select_related('trabajador', 'maquina_snapshot')
    return qs, 'trabajador'


# Configurar locale para español
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
    except:
        pass  # Usar locale por defecto si no se puede configurar español

# Configurar locale para español
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
    except:
        pass  # Usar locale por defecto si no se puede configurar español


# Helper: calcular inicio de semana operativa según contrato
from datetime import date as _date


def _inicio_semana_operativa(contrato, referencia=None):
    """Calcula la fecha de inicio de la semana operativa según
    `contrato.dia_cambio_guardia`.

    Si `referencia` es None se usa la fecha de hoy. Retorna la fecha
    correspondiente al último día igual a `dia_cambio_guardia` (<= referencia).
    """
    if referencia is None:
        referencia = _date.today()

    dia_objetivo = contrato.dia_cambio_guardia if getattr(contrato, 'dia_cambio_guardia', None) is not None else 0
    # weekday(): Monday=0 .. Sunday=6
    delta = (referencia.weekday() - int(dia_objetivo)) % 7
    inicio = referencia - timedelta(days=delta)
    return inicio


# Estados que cuentan como días activos en mina (V1 y V2)
ESTADOS_ACTIVOS_MINA = {
    'TRABAJADO', 'DIA_APOYO', 'STAND_BY',
    'INDUCCION', 'INDUCCION_VIRTUAL', 'RECORRIDO',
    'TRABAJO',  # alias V2
}


def _calcular_dias_previos_al_rango(trabajador_id, fecha_inicio_rango, limite_dias=90):
    """
    Cuenta cuántos días activos consecutivos tuvo el trabajador
    INMEDIATAMENTE antes de fecha_inicio_rango (sin límite mensual).
    Se consulta hasta `limite_dias` hacia atrás para cubrir casos extremos.
    """
    fecha_hasta = fecha_inicio_rango - timedelta(days=1)
    fecha_minima = fecha_hasta - timedelta(days=limite_dias)

    # Preferir datos normalizados (AsistenciaDiaria) si existen
    asistencias_v2 = AsistenciaDiaria.objects.filter(
        empleado_id=trabajador_id,
        fecha__gte=fecha_minima,
        fecha__lte=fecha_hasta,
    ).values('fecha', 'estado')

    asist_map = {a['fecha']: a['estado'] for a in asistencias_v2}

    # Completar con legacy V1 solo para fechas no cubiertas por V2
    faltantes = AsistenciaTrabajador.objects.filter(
        trabajador_id=trabajador_id,
        fecha__gte=fecha_minima,
        fecha__lte=fecha_hasta,
    ).exclude(fecha__in=list(asist_map.keys())).order_by('-fecha').values('fecha', 'estado')

    for a in faltantes:
        asist_map[a['fecha']] = a['estado']

    # Recorremos hacia atrás día a día verificando continuidad
    count = 0
    fecha_esperada = fecha_hasta
    while fecha_esperada >= fecha_minima:
        estado = asist_map.get(fecha_esperada)
        if estado and estado in ESTADOS_ACTIVOS_MINA:
            count += 1
            fecha_esperada -= timedelta(days=1)
        else:
            break

    return count


@login_required
def tareo_mensual_view(request):
    """Vista principal del tareo por semanas/rango personalizado"""
    user = request.user
    
    # Verificar permisos (solo managers de contrato y superiores)
    if not user.can_manage_contract_users():
        messages.error(request, 'No tienes permisos para gestionar el tareo de asistencia')
        return redirect('dashboard')
    
    # Determinar contrato
    if user.has_access_to_all_contracts():
        contrato_id = request.GET.get('contrato')
        if contrato_id:
            try:
                contrato = Contrato.objects.get(id=contrato_id, estado='ACTIVO')
            except Contrato.DoesNotExist:
                messages.error(request, 'Contrato no encontrado')
                contrato = None
        else:
            contrato = Contrato.objects.filter(estado='ACTIVO').first()
        contratos_disponibles = Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato')
    else:
        contrato = user.contrato
        contratos_disponibles = None
    
    if not contrato:
        messages.warning(request, 'No hay contratos activos disponibles')
        return redirect('dashboard')
    
    # NUEVA LÓGICA: Siempre mostrar mes completo, navegación por meses
    mes_offset = int(request.GET.get('mes_offset', 0))
    
    # Calcular el mes a mostrar
    hoy = datetime.now().date()
    if mes_offset == 0:
        # Mes actual
        fecha_base = hoy
    else:
        # Agregar/restar meses
        fecha_base = hoy
        for _ in range(abs(mes_offset)):
            if mes_offset > 0:
                # Mes siguiente
                if fecha_base.month == 12:
                    fecha_base = fecha_base.replace(year=fecha_base.year + 1, month=1, day=1)
                else:
                    fecha_base = fecha_base.replace(month=fecha_base.month + 1, day=1)
            else:
                # Mes anterior
                if fecha_base.month == 1:
                    fecha_base = fecha_base.replace(year=fecha_base.year - 1, month=12, day=1)
                else:
                    fecha_base = fecha_base.replace(month=fecha_base.month - 1, day=1)
    
    # Calcular primer y último día del mes operativo (26 al 25)
    mes_anterior = fecha_base.month - 1 if fecha_base.month > 1 else 12
    anio_anterior = fecha_base.year if fecha_base.month > 1 else fecha_base.year - 1
    
    fecha_inicio = date(anio_anterior, mes_anterior, 26)
    fecha_fin = date(fecha_base.year, fecha_base.month, 25)
    dias_a_mostrar = (fecha_fin - fecha_inicio).days + 1
    
    # Nombre del período (usar el mes operativo objetivo)
    meses_es = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    nombre_periodo = f"{meses_es[fecha_base.month]} {fecha_base.year}"
    
    # Generar lista de días del rango
    dia_cambio_guardia = contrato.dia_cambio_guardia if contrato.dia_cambio_guardia is not None else 6
    dia_previo_cambio  = (dia_cambio_guardia - 1) % 7

    dias_rango = []
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        # Nombres de días en español
        nombres_dias = {
            0: 'Lun', 1: 'Mar', 2: 'Mié', 3: 'Jue', 
            4: 'Vie', 5: 'Sáb', 6: 'Dom'
        }
        wd = fecha_actual.weekday()
        dias_rango.append({
            'fecha': fecha_actual,
            'dia': fecha_actual.day,
            'nombre_dia': nombres_dias[wd],
            'es_cambio_guardia': wd == dia_cambio_guardia,
            'es_previo_cambio':  wd == dia_previo_cambio,
        })
        fecha_actual += timedelta(days=1)
    
    # Obtener trabajadores activos del contrato
    from django.db.models import Case, When, Value, IntegerField as IntF
    trabajadores = _trabajadores_tareo_qs(
        contrato,
        fecha_inicio,
        fecha_fin,
    ).select_related('maquina_asignada').annotate(
        grupo_ord=Case(
            When(grupo='LINEA_MANDO',          then=Value(1)),
            When(grupo='OPERADORES',            then=Value(2)),
            When(grupo='SERVICIOS_GEOLOGICOS',  then=Value(3)),
            When(grupo='CONDUCTORES',     then=Value(4)),
            default=Value(5),
            output_field=IntF()
        ),
        cargo_ord=Case(
            When(cargo__icontains='RESIDENTE',  then=Value(1)),
            When(cargo__icontains='JEFE',       then=Value(2)),
            When(cargo__icontains='SUPERVISOR', then=Value(3)),
            When(cargo__icontains='INGENIERO',  then=Value(4)),
            When(cargo__icontains='PERFORISTA', then=Value(1)),
            When(cargo__icontains='AYUDANTE',   then=Value(2)),
            default=Value(9),
            output_field=IntF()
        )
    ).order_by('maquina_asignada__nombre', 'guardia_asignada', 'grupo_ord', 'cargo_ord', 'apepat', 'apemat', 'nombres')
    
    # Obtener asistencias del rango (legacy V1)
    asistencias_v1 = AsistenciaTrabajador.objects.filter(
        trabajador__contrato=contrato,
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin
    ).select_related('trabajador')

    # Crear diccionario de asistencias a partir de V1
    asistencias_dict = {}
    for asist in asistencias_v1:
        # Ignorar registros anteriores al inicio histórico del tareo
        if hasattr(TareoService, 'HISTORICO_START') and asist.fecha < TareoService.HISTORICO_START:
            continue
        if asist.trabajador.id not in asistencias_dict:
            asistencias_dict[asist.trabajador.id] = {}
        asistencias_dict[asist.trabajador.id][asist.fecha] = {
            'estado': asist.estado,
            'estado_display': asist.get_estado_display(),
            'tipo': asist.tipo,
            'tipo_display': asist.get_tipo_display(),
            'observaciones': asist.observaciones,
            'id': getattr(asist, 'id', None),
            'guardia_snapshot': getattr(asist, 'guardia_snapshot', None),
            'maquina_id': None,
            'maquina_nombre': None,
        }

    # Overlay: prefer AsistenciaDiaria (V2) when exists — proyecciones/actuales V2
    # Intentar consultar AsistenciaDiaria probando los posibles nombres de FK
    try:
        asistencias_v2, used_emp_field = _asistencias_v2_qs(contrato, fecha_inicio, fecha_fin)
        emp_field = used_emp_field or _emp_field_name()
        logger.debug("tareo_semanal: emp_field=%s (used=%s)", emp_field, used_emp_field)
        logger.debug("tareo_semanal: AsistenciaDiaria fields=%s", [f.name for f in AsistenciaDiaria._meta.get_fields()])
    except Exception:
        logger.exception("Error al intentar obtener AsistenciaDiaria desde compat")
        asistencias_v2 = AsistenciaDiaria.objects.none()

    for a in asistencias_v2:
        # admitir tanto `empleado` (legacy) como `trabajador` (nuevo esquema)
        emp_id = getattr(a, 'empleado_id', None) or getattr(a, 'trabajador_id', None)
        if emp_id not in asistencias_dict:
            asistencias_dict[emp_id] = {}

        # determinar si es proyección según esquema
        if hasattr(a, 'es_proyeccion'):
            es_proy = getattr(a, 'es_proyeccion')
        else:
            es_proy = (getattr(a, 'tipo', None) == 'PROY')

        asistencias_dict[emp_id][a.fecha] = {
            'estado': a.estado,
            'estado_display': a.get_estado_display() if hasattr(a, 'get_estado_display') else a.estado,
            'tipo': 'PAGABLE' if a.estado in ('TRABAJO', 'TD', 'TN') else 'NO_PAGABLE',
            'tipo_display': 'Pagable' if a.estado in ('TRABAJO', 'TD', 'TN') else 'No Pagable',
            'observaciones': getattr(a, 'observaciones', '') or '',
            'id': getattr(a, 'id', None),
            'guardia_snapshot': getattr(a, 'guardia_snapshot', None) or '',
            'maquina_id': getattr(a, 'maquina_snapshot_id', None),
            'maquina_nombre': a.maquina_snapshot.nombre if getattr(a, 'maquina_snapshot', None) else None,
            'es_proyeccion': es_proy,
        }
    
    # Combinar trabajadores con sus asistencias y agrupar por campo `grupo`
    GRUPO_META = {
        'LINEA_MANDO':         {'nombre': 'Línea de Mando',       'order': 1},
        'CONDUCTORES':         {'nombre': 'Conductores',           'order': 2},
        'OPERADORES':          {'nombre': 'Operadores',            'order': 3},
        '__STAND_BY__':        {'nombre': 'Personal Stand By',     'order': 4},
        'SERVICIOS_GEOLOGICOS':{'nombre': 'Servicios Geológicos',  'order': 5},
        '__SIN_GRUPO__':       {'nombre': 'Sin Grupo Asignado',    'order': 6},
    }

    trabajadores_por_grupo = {}

    for trabajador in trabajadores:
        tiene_maquina = bool(getattr(trabajador, 'maquina_asignada', None))
        if trabajador.es_standby and not tiene_maquina:
            primary_key = '__STAND_BY__'
        else:
            primary_key = trabajador.grupo if trabajador.grupo else '__SIN_GRUPO__'
        guardia_key = trabajador.guardia_asignada if trabajador.guardia_asignada else 'SIN_GUARDIA'

        if primary_key not in trabajadores_por_grupo:
            meta = GRUPO_META.get(primary_key, {'nombre': primary_key, 'order': 99})
            trabajadores_por_grupo[primary_key] = {
                'nombre': meta['nombre'],
                'order': meta['order'],
                'guardias': {}
            }


    @login_required
    def tareo_engine_preview(request):
        """
        Vista simplificada que genera una proyección determinista de 30/31 días
        usando el `TareoEngine` puro.

        Parámetros GET:
        - trabajador_id (int) required
        - start (YYYY-MM-DD) optional, default hoy
        - days (int) optional, default 30
        """
        user = request.user
        if not user.can_manage_contract_users():
            return JsonResponse({'error': 'No tienes permisos'}, status=403)

        trabajador_id = request.GET.get('trabajador_id')
        if not trabajador_id:
            return JsonResponse({'error': 'trabajador_id es requerido'}, status=400)

        try:
            trabajador = Trabajador.objects.select_related('contrato').get(id=int(trabajador_id))
        except Exception:
            return JsonResponse({'error': 'Trabajador no encontrado'}, status=404)

        start_str = request.GET.get('start')
        days = int(request.GET.get('days', 30))
        try:
            if start_str:
                fecha_inicio = datetime.strptime(start_str, '%Y-%m-%d').date()
            else:
                fecha_inicio = datetime.now().date()
        except Exception:
            return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}, status=400)

        fecha_fin = fecha_inicio + timedelta(days=days-1)

        proyeccion = TareoEngine.proyectar_rango(trabajador, fecha_inicio, fecha_fin)
        # Serializar fechas a ISO
        proyeccion_serial = [{'fecha': p['fecha'].isoformat(), 'estado': p['estado']} for p in proyeccion]

        return JsonResponse({'trabajador_id': trabajador.id, 'proyeccion': proyeccion_serial})

        if guardia_key not in trabajadores_por_grupo[primary_key]['guardias']:
            trabajadores_por_grupo[primary_key]['guardias'][guardia_key] = {
                'nombre': f"Guardia {guardia_key}" if guardia_key != 'SIN_GUARDIA' else 'Sin Guardia',
                'trabajadores': []
            }

        # Preparar asistencias del trabajador
        asistencias_trabajador = []
        for dia_info in dias_rango:
            fecha = dia_info['fecha']
            asist_dia = asistencias_dict.get(trabajador.id, {}).get(fecha)

            # Celda bloqueada si la fecha es anterior al inicio de labores del trabajador
            bloqueada = bool(
                (trabajador.fecha_inicio_labores and fecha < trabajador.fecha_inicio_labores)
                or (trabajador.fecha_cese and fecha > trabajador.fecha_cese)
            )

            # Normalizar estados legacy 'TRABAJADO'/'TRABAJO' a 'TD'/'TN' para operadores
            if asist_dia and asist_dia.get('estado') in ('TRABAJADO', 'TRABAJO'):
                # Solo reasignar cuando el trabajador tiene máquina asignada (OPERADORES)
                if getattr(trabajador, 'maquina_asignada', None):
                    try:
                        nuevo = TareoService.calcular_estado_dia(trabajador, fecha, forzar_alineacion=True)
                        if nuevo:
                            asist_dia['estado'] = nuevo
                            if nuevo == 'TD':
                                asist_dia['estado_display'] = 'Turno Día'
                            elif nuevo == 'TN':
                                asist_dia['estado_display'] = 'Turno Noche'
                    except Exception:
                        # Mantener estado legacy si falla la conversión
                        pass

            # Calcular estado sugerido si no hay asistencia (solo celdas no bloqueadas)
            # No sugerir ni generar estados para fechas anteriores al inicio histórico
            estado_sugerido = None
            if not asist_dia and not bloqueada and fecha >= TareoService.HISTORICO_START:
                try:
                    estado_sugerido = TareoService.calcular_estado_dia(trabajador, fecha, forzar_alineacion=True)
                except Exception:
                    estado_sugerido = trabajador.calcular_estado_regimen(fecha)

            asistencias_trabajador.append({
                'fecha': fecha,
                'bloqueada': bloqueada,
                'estado': asist_dia['estado'] if asist_dia else None,
                'estado_sugerido': estado_sugerido,
                'estado_display': asist_dia['estado_display'] if asist_dia else '-',
                'tipo': asist_dia['tipo'] if asist_dia else 'PAGABLE',
                'tipo_display': asist_dia['tipo_display'] if asist_dia else 'Pagable',
                'observaciones': asist_dia['observaciones'] if asist_dia else '',
                'es_cambio_guardia': dia_info['es_cambio_guardia'],
                'es_previo_cambio':  dia_info['es_previo_cambio']
            })
        
        dias_previos = _calcular_dias_previos_al_rango(trabajador.id, fecha_inicio)

        trabajadores_por_grupo[primary_key]['guardias'][guardia_key]['trabajadores'].append({
            'trabajador': trabajador,
            'asistencias': asistencias_trabajador,
            'dias_previos': dias_previos,
        })
    
    # 3. Construir lista ordenada final
    grupos_ordenados = []

    # Preferencia explícita en el orden de presentación de grupos
    preferred_order = ['LINEA_MANDO', 'OPERADORES', '__STAND_BY__', 'CONDUCTORES']
    ordered_keys = [k for k in preferred_order if k in trabajadores_por_grupo]
    # Añadir el resto respetando la orden definida en 'order'
    ordered_keys += [k for k in sorted(trabajadores_por_grupo.keys(), key=lambda k: trabajadores_por_grupo[k]['order']) if k not in preferred_order]

    for grupo_key in ordered_keys:
        grupo_data = trabajadores_por_grupo[grupo_key]

        # Ordenar guardias (A, B, C, luego sin guardia)
        guardias_ordenadas = []
        for guardia_key in ['A', 'B', 'C', 'SIN_GUARDIA']:
            if guardia_key in grupo_data['guardias']:
                guardia_info = grupo_data['guardias'][guardia_key]
                guardias_ordenadas.append({
                    'key': guardia_key,
                    'nombre': guardia_info['nombre'],
                    'trabajadores': guardia_info['trabajadores']
                })

        # Para OPERADORES: construir sub-agrupación por máquina
        maquinas_list = []
        if grupo_key == 'OPERADORES':
            import re as _re
            maquinas_dict = {}
            for gd in guardias_ordenadas:
                for item in gd['trabajadores']:
                    maq = item['trabajador'].maquina_asignada
                    maq_key = maq.nombre if maq else '__SIN_MAQUINA__'
                    maq_nombre = maq.nombre if maq else 'Sin Máquina Asignada'
                    maq_css = 'maq-' + _re.sub(r'[^a-zA-Z0-9]', '-', maq_key)
                    if maq_key not in maquinas_dict:
                        maquinas_dict[maq_key] = {'nombre': maq_nombre, 'css_key': maq_css, 'guardias': {}}
                    if gd['key'] not in maquinas_dict[maq_key]['guardias']:
                        maquinas_dict[maq_key]['guardias'][gd['key']] = {'nombre': gd['nombre'], 'trabajadores': []}
                    maquinas_dict[maq_key]['guardias'][gd['key']]['trabajadores'].append(item)

            for maq_key in sorted(maquinas_dict.keys(), key=lambda k: (k == '__SIN_MAQUINA__', k)):
                md = maquinas_dict[maq_key]
                guardias_maq = []
                for gk in ['A', 'B', 'C', 'SIN_GUARDIA']:
                    if gk in md['guardias']:
                        guardias_maq.append({'key': gk, 'nombre': md['guardias'][gk]['nombre'],
                                             'trabajadores': md['guardias'][gk]['trabajadores']})
                maquinas_list.append({
                    'nombre': md['nombre'], 'css_key': md['css_key'],
                    'guardias': guardias_maq,
                    'total_personas': sum(len(g['trabajadores']) for g in guardias_maq),
                })

        grupos_ordenados.append({
            'key': grupo_key,
            'nombre': grupo_data['nombre'],
            'guardias': guardias_ordenadas,
            'maquinas': maquinas_list,
            'total_trabajadores': sum(len(g['trabajadores']) for g in guardias_ordenadas)
        })

    
    # Contexto para el template
    context = {
        'contrato': contrato,
        'contratos_disponibles': contratos_disponibles,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'nombre_periodo': nombre_periodo,
        'dias_rango': dias_rango,
        'grupos_ordenados': grupos_ordenados,
        'total_trabajadores': trabajadores.count(),
        'total_dias': dias_a_mostrar,
        'estados_asistencia': AsistenciaTrabajador.ESTADO_ASISTENCIA_CHOICES,
        'mes_offset': mes_offset,
    }
    
    return render(request, 'drilling/tareo/mensual.html', context)


@require_http_methods(["POST"])
def guardar_asistencia(request):
    """API para guardar asistencia individual

    Esta vista maneja autenticación y verificación CSRF manualmente para
    devolver siempre JSON (evita respuestas HTML que rompen parsers AJAX).
    """
    # Autenticación: devolver JSON 401 en lugar de redirect HTML
    user = request.user
    if not user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'No autenticado'}, status=401)

    # CSRF: validar explícitamente y devolver JSON si falla
    csrf_resp = CsrfViewMiddleware().process_view(request, None, (), {})
    if csrf_resp is not None:
        return JsonResponse({'success': False, 'message': 'CSRF verification failed'}, status=403)

    if not user.can_manage_contract_users():
        return JsonResponse({'success': False, 'message': 'Sin permisos'}, status=403)

    try:
        data = json.loads(request.body)
        trabajador_id = data.get('trabajador_id')
        fecha_str = data.get('fecha')
        estado = data.get('estado')
        tipo = data.get('tipo', 'PAGABLE')  # Por defecto PAGABLE
        observaciones = data.get('observaciones', '')
        
        # Validaciones
        if not all([trabajador_id, fecha_str, estado]):
            return JsonResponse({'success': False, 'message': 'Datos incompletos'}, status=400)
        
        trabajador = Trabajador.objects.get(id=trabajador_id)
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Verificar que el usuario tenga acceso al contrato del trabajador
        if not user.has_contract_permission(trabajador.contrato):
            return JsonResponse({'success': False, 'message': 'Sin acceso a este contrato'}, status=403)
        
        # Crear o actualizar asistencia
        asistencia, created = AsistenciaTrabajador.objects.update_or_create(
            trabajador=trabajador,
            fecha=fecha,
            defaults={
                'estado': estado,
                'tipo': tipo,
                'observaciones': observaciones,
                'registrado_por': user,
                'cargo_snapshot': trabajador.cargo or None,
                'guardia_snapshot': trabajador.guardia_asignada
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Asistencia guardada correctamente',
            'estado_display': asistencia.get_estado_display(),
            'tipo_display': asistencia.get_tipo_display()
        })
        
    except Trabajador.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Trabajador no encontrado'}, status=404)
    except ValueError as e:
        return JsonResponse({'success': False, 'message': f'Error en formato de fecha: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


@require_http_methods(["POST"])
def guardar_asistencias_masivas(request):
    """API para guardar múltiples asistencias en una sola operación

    Igual que `guardar_asistencia`, esta vista valida autenticación y CSRF
    para devolver JSON legible por el frontend en caso de fallo.
    """
    user = request.user
    if not user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'No autenticado'}, status=401)

    csrf_resp = CsrfViewMiddleware().process_view(request, None, (), {})
    if csrf_resp is not None:
        return JsonResponse({'success': False, 'message': 'CSRF verification failed'}, status=403)

    if not user.can_manage_contract_users():
        return JsonResponse({'success': False, 'message': 'Sin permisos'}, status=403)

    try:
        data = json.loads(request.body)
        asistencias_data = data.get('asistencias', [])
        
        if not asistencias_data:
            return JsonResponse({'success': False, 'message': 'No hay datos para guardar'}, status=400)
        
        guardadas = 0
        turnos_actualizados = 0
        errores = []

        with transaction.atomic():
            for item in asistencias_data:
                trabajador_id = item.get('trabajador_id')
                fecha_str = item.get('fecha')
                estado = item.get('estado')
                turno = item.get('turno')
                observaciones = item.get('observaciones', '')
                tipo = item.get('tipo', None)

                # Validar existencia
                try:
                    trabajador = Trabajador.objects.get(id=trabajador_id)
                except Trabajador.DoesNotExist:
                    errores.append(f"Trabajador {trabajador_id} no existe")
                    continue

                # Verificar permisos al contrato
                if not user.has_contract_permission(trabajador.contrato):
                    errores.append(f"Sin permiso en trabajador {trabajador_id}")
                    continue

                try:
                    # Si se envía turno, actualizar la guardia asignada del trabajador
                    if turno:
                        trabajador.guardia_asignada = turno
                        trabajador.save(update_fields=['guardia_asignada'])
                        turnos_actualizados += 1

                    # Si se envía estado, crear/actualizar asistencia
                    if estado:
                        if not fecha_str:
                            errores.append(f"Falta fecha para trabajador {trabajador_id}")
                            continue
                        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                        defaults = {
                            'estado': estado,
                            'observaciones': observaciones,
                            'registrado_por': user,
                            'cargo_snapshot': trabajador.cargo or None,
                            'guardia_snapshot': trabajador.guardia_asignada
                        }
                        if tipo:
                            defaults['tipo'] = tipo

                        AsistenciaTrabajador.objects.update_or_create(
                            trabajador=trabajador,
                            fecha=fecha,
                            defaults=defaults
                        )
                        guardadas += 1

                except Exception as e:
                    errores.append(f"Error en trabajador {trabajador_id}: {str(e)}")

        mensaje = f'{guardadas} asistencias guardadas, {turnos_actualizados} turnos actualizados'
        if errores:
            return JsonResponse({
                'success': True,
                'message': mensaje + f' con {len(errores)} errores',
                'errores': errores
            })

        return JsonResponse({
            'success': True,
            'message': mensaje
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


# MAPEO DE ESTADOS DEL SISTEMA A CÓDIGOS
MAPEO_CODIGOS = {
    'TRABAJADO': 'T',
    'TRABAJO': 'T',
    'T': 'T',
    'TRABAJO_DIA': 'TD',
    'TRABAJO_NOCHE': 'TN',
    'TD': 'TD',
    'TN': 'TN',
    'DIA_LIBRE': 'DL',
    'DESCANSO': 'DL',
    'DL': 'DL',
    'DIA_APOYO': 'DA',
    'DA': 'DA',
    'PERMISO_PATERNIDAD': 'P',
    'PT': 'P',
    'DESCANSO_MEDICO': 'DM',
    'DM': 'DM',
    'STAND_BY': 'SB',
    'SB': 'SB',
    'SUBSIDIO': 'SUB',
    'SUB': 'SUB',
    'INDUCCION': 'IND',
    'INDUCCION_VIRTUAL': 'IND',
    'IND': 'IND',
    'I': 'IND',
    'IV': 'IND',
    'RECORRIDO': 'REC',
    'REC': 'REC',
    'R': 'REC',
    'FALTA': 'F',
    'F': 'F',
    'PERMISO': 'P',
    'P': 'P',
    'SUSPENSION': 'S',
    'S': 'S',
    'VACACIONES': 'V',
    'V': 'V',
    'LICENCIA': 'L',
    'LICENCIA_SIN_GOCE': 'L',
    'LSG': 'L',
    'CESADO': 'C',
    'TRABAJO_CALIENTE': 'TC',
    'LICENCIA_FALLECIMIENTO': 'L',
    'LICENCIA_CON_GOCE': 'L',
    'LF': 'L',
    'LCG': 'L',
    'L': 'L',
}

# LEYENDA DE CÓDIGOS
LEYENDA = {
    'T': 'TRABAJADO',
    'TD': 'TRABAJADO DÍA',
    'TN': 'TRABAJADO NOCHE',
    'T1': 'TRABAJADO + 1 H.E.',
    'T2': 'TRABAJADO + 2 H.E.',
    'DL': 'DIA LIBRE',
    'DA': 'DIA APOYO',
    'DM': 'DESCANSO MEDICO',
    'SB': 'STAND BY',
    'SUB': 'SUBSIDIO',
    'IND': 'INDUCCION',
    'REC': 'RECORRIDO',
    'F': 'FALTA',
    'P': 'PERMISO',
    'S': 'SUSPENSION',
    'V': 'VACACIONES',
    'L': 'LICENCIA',
    'C': 'CESADO',
    'TC': 'TRABAJO EN CALIENTE',
}

# Colores vivos por código para el Excel (formato ARGB sin #)
COLORES_EXCEL = {
    'T':   'FF00C853',  # Verde brillante — Trabajado
    'TD':  'FFFFB300',  # Ámbar — Trabajo Día
    'TN':  'FF0D47A1',  # Azul oscuro — Trabajo Noche
    'DL':  'FF2979FF',  # Azul eléctrico — Día Libre
    'DA':  'FF1565C0',  # Azul marino — Día Apoyo
    'PT':  'FFF57F17',  # Ámbar — Permiso Paternidad
    'DM':  'FFD50000',  # Rojo vivo — Descanso Médico
    'SB':  'FF00BFA5',  # Teal — Stand By
    'SUB': 'FF6D4C41',  # Marrón — Subsidio
    'I':   'FFFF6D00',  # Naranja brillante — Inducción
    'IV':  'FFFF6D00',  # Naranja brillante — Inducción Virtual
    'R':   'FF558B2F',  # Verde oscuro — Recorrido
    'F':   'FFB71C1C',  # Rojo sangre — Falta
    'P':   'FFE65100',  # Naranja oscuro — Permiso
    'S':   'FF880E4F',  # Fucsia — Suspensión
    'V':   'FF6A1B9A',  # Violeta — Vacaciones
    'LSG': 'FF37474F',  # Gris oscuro — Licencia Sin Goce
    'LCG': 'FF455A64',  # Gris azulado — Licencia Con Goce
    'C':   'FF212121',  # Negro — Cesado
    'TC':  'FFFF1744',  # Rojo neón — Trabajo en Caliente
    'LF':  'FF4A148C',  # Morado oscuro — Licencia por Fallecimiento
}


@login_required
def exportar_asistencias_excel(request):
    """
    Exportar asistencias a Excel en formato completo con 3 hojas:
    - Tareo: Tabla principal con trabajadores y marcaciones diarias
    - Leyenda: Códigos de asistencia
    - Informe: Estadísticas y resúmenes
    """
    from django.http import HttpResponse
    from django.db.models import Count
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    user = request.user
    
    if not user.can_manage_contract_users():
        return HttpResponse('Sin permisos', status=403)
    
    # Obtener parámetros
    contrato_id = request.GET.get('contrato')
    modo = request.GET.get('modo', 'semana')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    
    # Validar contrato
    if user.has_access_to_all_contracts():
        if not contrato_id:
            contrato = Contrato.objects.filter(estado='ACTIVO').first()
        else:
            contrato = Contrato.objects.filter(id=contrato_id, estado='ACTIVO').first()
    else:
        contrato = user.contrato
    
    if not contrato:
        return HttpResponse('Contrato no encontrado', status=404)
    
    try:
        # Calcular rango de fechas
        if fecha_inicio_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            except:
                fecha_inicio = date.today()
        else:
            fecha_inicio = date.today()
        
        # Calcular fecha_fin según modo
        if modo == 'semana':
            dias_a_mostrar = 7
            fecha_inicio = _inicio_semana_operativa(contrato, fecha_inicio)
            fecha_fin = fecha_inicio + timedelta(days=dias_a_mostrar - 1)
        elif modo == 'quincena':
            dias_a_mostrar = 15
            fecha_fin = fecha_inicio + timedelta(days=dias_a_mostrar - 1)
        elif modo == 'mes':
            # Si vienen ambas fechas (mes operativo 26→25), usarlas directamente.
            # Solo recalcular al mes calendario cuando NO viene fecha_fin.
            if fecha_fin_str:
                try:
                    fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                except Exception:
                    fecha_fin = fecha_inicio.replace(day=25)
            else:
                # Fallback: mes calendario (día 1 → último día del mes)
                primer_dia = fecha_inicio.replace(day=1)
                if primer_dia.month == 12:
                    fecha_fin = primer_dia.replace(year=primer_dia.year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    fecha_fin = primer_dia.replace(month=primer_dia.month + 1, day=1) - timedelta(days=1)
                fecha_inicio = primer_dia
            dias_a_mostrar = (fecha_fin - fecha_inicio).days + 1
        else:  # personalizado
            if fecha_fin_str:
                try:
                    fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                except:
                    fecha_fin = fecha_inicio + timedelta(days=7)
            else:
                fecha_fin = fecha_inicio + timedelta(days=7)
            dias_a_mostrar = (fecha_fin - fecha_inicio).days + 1
        
        num_dias = dias_a_mostrar
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto

        include_mes_anterior = request.GET.get('include_mes_anterior') == '1'

        if include_mes_anterior:
            # Calcular mes operativo anterior (restar 1 mes al inicio y fin)
            if fecha_inicio.month == 1:
                fi_ant = fecha_inicio.replace(year=fecha_inicio.year - 1, month=12)
            else:
                fi_ant = fecha_inicio.replace(month=fecha_inicio.month - 1)
            if fecha_fin.month == 1:
                ff_ant = fecha_fin.replace(year=fecha_fin.year - 1, month=12)
            else:
                ff_ant = fecha_fin.replace(month=fecha_fin.month - 1)
            nd_ant = (ff_ant - fi_ant).days + 1

            # Hoja mes anterior (1ª)
            ws_ant = wb.create_sheet(f"Tareo {fi_ant.strftime('%b %Y').upper()}", 0)
            _crear_hoja_tareo(ws_ant, contrato, fi_ant, ff_ant, nd_ant)

        # 1. CREAR HOJA TAREO (mes actual — se agrega al final)
        ws_tareo = wb.create_sheet(f"Tareo {fecha_inicio.strftime('%b %Y').upper()}")
        _crear_hoja_tareo(ws_tareo, contrato, fecha_inicio, fecha_fin, num_dias)

        # 2. CREAR HOJA LEYENDA
        ws_leyenda = wb.create_sheet("LEYENDA")
        _crear_hoja_leyenda(ws_leyenda)

        # 3. CREAR HOJA INFORME
        ws_informe = wb.create_sheet("Informe")
        _crear_hoja_informe(ws_informe, contrato, fecha_inicio, fecha_fin)
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        mes_nombre = fecha_inicio.strftime('%B').capitalize()
        sufijo = '_con_MesAnterior' if include_mes_anterior else ''
        filename = f"Tareo_{contrato.nombre_contrato.replace(' ', '_')}_{mes_nombre}_{fecha_inicio.year}{sufijo}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error al generar Excel: {str(e)}', status=500)


def _crear_hoja_tareo(ws, contrato, fecha_inicio, fecha_fin, num_dias):
    """Crea la hoja principal de tareo"""
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    subheader_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # FILA 1: Encabezado principal
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = f"TAREO MES DE : {fecha_inicio.strftime('%B %Y').upper()}"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('E1:H1')
    cell = ws['E1']
    cell.value = f"CONTRATO : {contrato.nombre_contrato.upper()}"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # FILA 2: Etiquetas de semanas
    col_actual = 9  # Columna I (después de las 8 columnas fijas)
    fecha_actual = fecha_inicio
    
    while fecha_actual <= fecha_fin:
        # Calcular semana
        semana = fecha_actual.isocalendar()[1]
        col_letter = get_column_letter(col_actual)
        
        # Buscar inicio de semana
        if fecha_actual.weekday() == 0 or fecha_actual == fecha_inicio:  # Lunes o primer día
            inicio_semana = col_actual
            # Contar días de esta semana en el rango
            dias_semana = 0
            temp_fecha = fecha_actual
            while temp_fecha <= fecha_fin and temp_fecha.weekday() < 7:
                dias_semana += 1
                temp_fecha += timedelta(days=1)
                if temp_fecha.weekday() == 0:
                    break
            
            # Merge cells para la semana
            if dias_semana > 1:
                fin_semana = inicio_semana + dias_semana - 1
                ws.merge_cells(start_row=2, start_column=inicio_semana, end_row=2, end_column=fin_semana)
            
            cell = ws.cell(row=2, column=inicio_semana)
            cell.value = f"Semana {semana}"
            cell.font = Font(bold=True, size=10)
            cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        col_actual += 1
        fecha_actual += timedelta(days=1)
    
    # FILA 3: Headers de columnas
    headers = [
        'ITEM', 'CODIGO', 'APELLIDOS Y NOMBRES', 'Cargo', 'MAQUINA', 'CATEGORIA',
        'GRUPO', 'GUARDIA'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    
    # Headers de días
    fecha_actual = fecha_inicio
    col_num = 9  # Después de "GUARDIA" (ahora 8 columnas fijas)
    while fecha_actual <= fecha_fin:
        cell = ws.cell(row=3, column=col_num)
        cell.value = fecha_actual
        cell.number_format = 'DD/MM'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        col_num += 1
        fecha_actual += timedelta(days=1)
    
    # Headers de resumen (SIMPLIFICADO)
    headers_resumen = [
        'TRABAJADO (T)', 'DIAS LIBRES (DL)', 'FALTAS (F)', 
        'VACACIONES (V)', 'D. MEDICO (DM)', 'TOTAL DIAS'
    ]
    
    for header in headers_resumen:
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
        col_num += 1
    
    # DATOS DE TRABAJADORES
    trabajadores = _trabajadores_tareo_qs(
        contrato,
        fecha_inicio,
        fecha_fin,
    ).order_by('grupo', 'apepat', 'apemat', 'nombres')
    
    # Obtener asistencias desde V2 (AsistenciaDiaria). Usar V1 solo como
    # fallback si no hay registros en V2 para el contrato/periodo.
    from .tareo_compat import AsistenciaDiaria
    filter_kwargs = {f"{emp_field}__contrato": contrato, 'fecha__gte': fecha_inicio, 'fecha__lte': fecha_fin}
    asistencias_v2 = AsistenciaDiaria.objects.filter(
        **filter_kwargs
    ).select_related(emp_field, 'maquina_snapshot')

    # Diccionario de asistencias (preferir V2)
    asist_dict = {}
    if asistencias_v2.exists():
        for asist in asistencias_v2:
            emp_id = getattr(asist, f"{emp_field}_id")
            if emp_id not in asist_dict:
                asist_dict[emp_id] = {}
            asist_dict[emp_id][asist.fecha] = asist
    else:
        # Fallback a tabla legacy AsistenciaTrabajador
        asistencias = AsistenciaTrabajador.objects.filter(
            trabajador__contrato=contrato,
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin
        ).select_related('trabajador')
        for asist in asistencias:
            if asist.trabajador.id not in asist_dict:
                asist_dict[asist.trabajador.id] = {}
            asist_dict[asist.trabajador.id][asist.fecha] = asist
    
    # Agrupar por CATEGORIA -> MAQUINA (solo para OPERADORES) -> GUARDIA -> trabajadores
    import re as _re
    categorias = {}
    order_map = {
        'LINEA_MANDO': 1,
        'CONDUCTORES': 2,
        'OPERADORES': 3,
        '__STAND_BY__': 4,
        'SERVICIOS_GEOLOGICOS': 5,
        '__SIN_GRUPO__': 6
    }

    for trabajador in trabajadores:
        tiene_maquina = bool(getattr(trabajador, 'maquina_asignada', None))
        if getattr(trabajador, 'es_standby', False) and not tiene_maquina:
            cat_key = '__STAND_BY__'
        else:
            cat_key = trabajador.grupo if trabajador.grupo else '__SIN_GRUPO__'

        guardia_key = trabajador.guardia_asignada if trabajador.guardia_asignada else 'SIN_GUARDIA'

        if cat_key not in categorias:
            display = ' '.join(cat_key.split('_')).title() if cat_key not in ['__STAND_BY__', '__SIN_GRUPO__'] else (
                'Personal Stand By' if cat_key == '__STAND_BY__' else 'Sin Grupo Asignado'
            )
            categorias[cat_key] = {
                'nombre': display,
                'order': order_map.get(cat_key, 99),
                'maquinas': {},  # solo usada para OPERADORES
                'guardias': {}
            }

        # Si es OPERADORES, agrupar por máquina
        if cat_key == 'OPERADORES':
            maq = getattr(trabajador, 'maquina_asignada', None)
            maq_key = maq.nombre if maq else '__SIN_MAQUINA__'
            maq_nombre = maq.nombre if maq else 'Sin Máquina Asignada'
            maq_css = 'maq-' + _re.sub(r'[^a-zA-Z0-9]', '-', maq_key)
            if maq_key not in categorias[cat_key]['maquinas']:
                categorias[cat_key]['maquinas'][maq_key] = {'nombre': maq_nombre, 'css_key': maq_css, 'guardias': {}}
            categorias[cat_key]['maquinas'][maq_key]['guardias'].setdefault(guardia_key, []).append(trabajador)
        else:
            # Para otros grupos, agrupar directamente por guardia
            categorias[cat_key]['guardias'].setdefault(guardia_key, []).append(trabajador)

    row_num = 4
    item_idx = 1

    # Ordenar categorías por orden definido
    sorted_cats = sorted(categorias.items(), key=lambda kv: (categorias[kv[0]]['order'], kv[0]))
    for cat_key, cat_data in sorted_cats:
        # Escribir encabezado de categoría
        total_cols = len(headers) + num_dias + len(headers_resumen)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=total_cols)
        cell = ws.cell(row=row_num, column=1)
        cell.value = f"CATEGORÍA: {cat_data['nombre'].upper()}"
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        row_num += 1

        if cat_key == 'OPERADORES':
            # Dentro de operadores, ordenar máquinas
            sorted_maquinas = sorted(cat_data['maquinas'].items(), key=lambda kv: (kv[0] == '__SIN_MAQUINA__', kv[0]))
            for maq_key, maq_data in sorted_maquinas:
                # Encabezado de máquina
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=total_cols)
                cell = ws.cell(row=row_num, column=1)
                cell.value = f"MÁQUINA: {maq_data['nombre']}"
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')
                row_num += 1

                # Por cada guardia en la máquina
                guardia_order = lambda g: (g != 'A', g != 'B', g != 'C', g)
                for guardia_key in sorted(maq_data['guardias'].keys(), key=guardia_order):
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=total_cols)
                    cell = ws.cell(row=row_num, column=1)
                    cell.value = f"Guardia: {guardia_key if guardia_key != 'SIN_GUARDIA' else 'Sin Guardia'}"
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                    row_num += 1

                    for trabajador in maq_data['guardias'][guardia_key]:
                        ws.cell(row=row_num, column=1).value = item_idx
                        ws.cell(row=row_num, column=2).value = trabajador.dni
                        ws.cell(row=row_num, column=3).value = f"{trabajador.apellidos}, {trabajador.nombres}"
                        ws.cell(row=row_num, column=4).value = trabajador.cargo or ""
                        ws.cell(row=row_num, column=5).value = (trabajador.maquina_asignada.nombre if getattr(trabajador, 'maquina_asignada', None) else "")
                        # Categoria legible
                        grupo_val = getattr(trabajador, 'grupo', '') or ''
                        categoria_display = ''
                        if grupo_val == 'LINEA_MANDO':
                            categoria_display = 'LINEA DE MANDO'
                        elif grupo_val == 'OPERADORES':
                            categoria_display = 'OPERADORES'
                        elif grupo_val == 'CONDUCTORES':
                            categoria_display = 'CONDUCTORES'
                        else:
                            categoria_display = grupo_val or ''
                        ws.cell(row=row_num, column=6).value = categoria_display
                        ws.cell(row=row_num, column=7).value = trabajador.grupo or ''
                        ws.cell(row=row_num, column=8).value = trabajador.guardia_asignada if trabajador.guardia_asignada else ""

                        # Marcaciones diarias
                        fecha_actual = fecha_inicio
                        col_num = 9
                        contadores = {'T': 0, 'TD': 0, 'TN': 0, 'DL': 0, 'F': 0, 'V': 0, 'DM': 0}
                        while fecha_actual <= fecha_fin:
                            asist = asist_dict.get(trabajador.id, {}).get(fecha_actual)
                            cell = ws.cell(row=row_num, column=col_num)
                            if asist:
                                codigo = getattr(asist, 'estado', None) or getattr(asist, 'estado', '')
                                codigo = MAPEO_CODIGOS.get(codigo, codigo)
                                guardia = getattr(asist, 'guardia_snapshot', None) or ''
                                cell.value = f"{codigo}{(' - ' + guardia) if guardia else ''}"
                                hex_color = COLORES_EXCEL.get(codigo)
                                if hex_color:
                                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFF', size=10)
                                if codigo in contadores:
                                    contadores[codigo] += 1
                                else:
                                    # Fallback: si nos llega el estado largo, intentar mapear
                                    corto = MAPEO_CODIGOS.get(codigo)
                                    if corto and corto in contadores:
                                        contadores[corto] += 1
                            else:
                                cell.value = ""

                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = border_thin
                            col_num += 1
                            fecha_actual += timedelta(days=1)

                        # Totales Simplificados
                        ws.cell(row=row_num, column=col_num).value = contadores['T']
                        ws.cell(row=row_num, column=col_num+1).value = contadores['DL']
                        ws.cell(row=row_num, column=col_num+2).value = contadores['F']
                        ws.cell(row=row_num, column=col_num+3).value = contadores['V']
                        ws.cell(row=row_num, column=col_num+4).value = contadores['DM']
                        total_registrados = sum(contadores.values())
                        ws.cell(row=row_num, column=col_num+5).value = total_registrados

                        row_num += 1
                        item_idx += 1
        else:
            # Para categorías distintas de OPERADORES, agrupar por guardia directamente
            guardia_order = lambda g: (g != 'A', g != 'B', g != 'C', g)
            for guardia_key in sorted(cat_data['guardias'].keys(), key=guardia_order):
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=total_cols)
                cell = ws.cell(row=row_num, column=1)
                cell.value = f"Guardia: {guardia_key if guardia_key != 'SIN_GUARDIA' else 'Sin Guardia'}"
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                row_num += 1

                for trabajador in cat_data['guardias'][guardia_key]:
                    ws.cell(row=row_num, column=1).value = item_idx
                    ws.cell(row=row_num, column=2).value = trabajador.dni
                    ws.cell(row=row_num, column=3).value = f"{trabajador.apellidos}, {trabajador.nombres}"
                    ws.cell(row=row_num, column=4).value = trabajador.cargo or ""
                    ws.cell(row=row_num, column=5).value = (trabajador.maquina_asignada.nombre if getattr(trabajador, 'maquina_asignada', None) else "")
                    grupo_val = getattr(trabajador, 'grupo', '') or ''
                    categoria_display = ''
                    if grupo_val == 'LINEA_MANDO':
                        categoria_display = 'LINEA DE MANDO'
                    elif grupo_val == 'OPERADORES':
                        categoria_display = 'OPERADORES'
                    elif grupo_val == 'CONDUCTORES':
                        categoria_display = 'CONDUCTORES'
                    else:
                        categoria_display = grupo_val or ''
                    ws.cell(row=row_num, column=6).value = categoria_display
                    ws.cell(row=row_num, column=7).value = trabajador.grupo or ''
                    ws.cell(row=row_num, column=8).value = trabajador.guardia_asignada if trabajador.guardia_asignada else ""

                    fecha_actual = fecha_inicio
                    col_num = 9
                    contadores = {'T': 0, 'TD': 0, 'TN': 0, 'DL': 0, 'F': 0, 'V': 0, 'DM': 0}
                    while fecha_actual <= fecha_fin:
                        asist = asist_dict.get(trabajador.id, {}).get(fecha_actual)
                        cell = ws.cell(row=row_num, column=col_num)
                        if asist:
                            codigo = getattr(asist, 'estado', None) or getattr(asist, 'estado', '')
                            codigo = MAPEO_CODIGOS.get(codigo, codigo)
                            guardia = getattr(asist, 'guardia_snapshot', None) or ''
                            cell.value = f"{codigo}{(' - ' + guardia) if guardia else ''}"
                            hex_color = COLORES_EXCEL.get(codigo)
                            if hex_color:
                                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                                cell.font = Font(bold=True, color='FFFFFF', size=10)
                            if codigo in contadores:
                                contadores[codigo] += 1
                            else:
                                corto = MAPEO_CODIGOS.get(codigo)
                                if corto and corto in contadores:
                                    contadores[corto] += 1
                        else:
                            cell.value = ""

                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border_thin
                        col_num += 1
                        fecha_actual += timedelta(days=1)

                    ws.cell(row=row_num, column=col_num).value = contadores['T']
                    ws.cell(row=row_num, column=col_num+1).value = contadores['DL']
                    ws.cell(row=row_num, column=col_num+2).value = contadores['F']
                    ws.cell(row=row_num, column=col_num+3).value = contadores['V']
                    ws.cell(row=row_num, column=col_num+4).value = contadores['DM']
                    total_registrados = sum(contadores.values())
                    ws.cell(row=row_num, column=col_num+5).value = total_registrados

                    row_num += 1
                    item_idx += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10

    # Días (columnas de marcación) — empiezan en columna 9 (I)
    for col in range(9, 9 + num_dias):
        ws.column_dimensions[get_column_letter(col)].width = 4


def _crear_hoja_leyenda(ws):
    """Crea la hoja de leyenda con códigos"""
    ws.merge_cells('A1:B1')
    cell = ws['A1']
    cell.value = "LEYENDA: CODIFICACION"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 3
    for codigo, descripcion in LEYENDA.items():
        ws.cell(row=row, column=1).value = codigo
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = descripcion
        row += 1
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40


def _crear_hoja_informe(ws, contrato, fecha_inicio, fecha_fin):
    """Crea la hoja de informe con estadísticas"""
    ws['A1'] = f"INFORME DE TAREO - {contrato.nombre_contrato.upper()}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    
    # Estadísticas
    trabajadores = _trabajadores_tareo_qs(contrato, fecha_inicio, fecha_fin)
    # Preferir AsistenciaDiaria (V2) para estadísticas del informe
    from .tareo_compat import AsistenciaDiaria
    emp_field = _emp_field_name()
    asistencias = AsistenciaDiaria.objects.filter(
        **{f"{emp_field}__contrato": contrato, 'fecha__gte': fecha_inicio, 'fecha__lte': fecha_fin}
    )
    
    row = 4
    ws.cell(row=row, column=1).value = "Total Trabajadores"
    ws.cell(row=row, column=2).value = trabajadores.count()
    ws.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=1).value = "Total Registros de Asistencia"
    ws.cell(row=row, column=2).value = asistencias.count()
    ws.cell(row=row, column=1).font = Font(bold=True)
    
    row += 2
    ws.cell(row=row, column=1).value = "Distribución por Estado:"
    ws.cell(row=row, column=1).font = Font(bold=True, underline="single")
    
    row += 1
    # Contar por estado
    estados = asistencias.values('estado').annotate(total=Count('estado')).order_by('-total')
    for estado in estados:
        row += 1
        codigo = MAPEO_CODIGOS.get(estado['estado'], estado['estado'])
        descripcion = LEYENDA.get(codigo, estado['estado'])
        ws.cell(row=row, column=1).value = f"{codigo} - {descripcion}"
        ws.cell(row=row, column=2).value = estado['total']


@login_required
def exportar_tareo_semanal(request):
    """
    Exportador semanal con dos hojas:
    - Matriz por trabajador (fila por trabajador, columna por día)
    - Distribución por máquina (bloques por máquina y guardia)
    Usa los parámetros `contrato` y `fecha_inicio` (YYYY-MM-DD). Si no vienen,
    usa el contrato del usuario y la semana actual (lunes como inicio).
    """
    from django.http import HttpResponse
    try:
        user = request.user
        if not user.can_manage_contract_users():
            return HttpResponse('Sin permisos', status=403)

        contrato_id = request.GET.get('contrato')
        if user.has_access_to_all_contracts():
            contrato = Contrato.objects.filter(id=contrato_id, estado='ACTIVO').first() if contrato_id else Contrato.objects.filter(estado='ACTIVO').first()
        else:
            contrato = user.contrato

        if not contrato:
            return HttpResponse('Contrato no encontrado', status=404)

        fecha_inicio_str = request.GET.get('fecha_inicio')
        if fecha_inicio_str:
            fecha_inicio_ref = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_inicio = _inicio_semana_operativa(contrato, fecha_inicio_ref)
        else:
            # semana actual: tomar inicio según dia_cambio_guardia del contrato
            fecha_inicio = _inicio_semana_operativa(contrato)

        fecha_fin = fecha_inicio + timedelta(days=6)
        num_dias = 7

        wb = Workbook()
        wb.remove(wb.active)

        # Hoja 1: Matriz (reusar _crear_hoja_tareo pero limitada a la semana)
        ws_mat = wb.create_sheet('Matriz semanal')
        _crear_hoja_tareo(ws_mat, contrato, fecha_inicio, fecha_fin, num_dias)

        # Hoja 2: Distribución por máquina
        ws_dist = wb.create_sheet('Distribución por Máquina')
        _crear_hoja_distribucion(ws_dist, contrato, fecha_inicio, fecha_fin)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Tareo_Semanal_{contrato.nombre_contrato.replace(' ', '_')}_{fecha_inicio.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error al generar Excel semanal: {str(e)}', status=500)


@login_required
def mostrar_tareo_semanal(request):
    """
    Muestra en la aplicación la representación semanal (matriz + distribución)
    y permite descargar la vista como imagen desde el cliente.
    Parámetros GET: `contrato` y `fecha_inicio` (YYYY-MM-DD) opcionales.
    """
    try:
        user = request.user
        if not user.can_manage_contract_users():
            return HttpResponse('Sin permisos', status=403)

        contrato_id = request.GET.get('contrato')
        if user.has_access_to_all_contracts():
            contrato = Contrato.objects.filter(id=contrato_id, estado='ACTIVO').first() if contrato_id else Contrato.objects.filter(estado='ACTIVO').first()
        else:
            contrato = user.contrato

        if not contrato:
            return HttpResponse('Contrato no encontrado', status=404)

        fecha_inicio_str = request.GET.get('fecha_inicio')
        if fecha_inicio_str:
            fecha_inicio_ref = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_inicio = _inicio_semana_operativa(contrato, fecha_inicio_ref)
        else:
            fecha_inicio = _inicio_semana_operativa(contrato)

        fecha_fin = fecha_inicio + timedelta(days=6)

        # Obtener asistencias V2 (semana) — intentar ambos nombres de FK
        from .tareo_compat import AsistenciaDiaria
        try:
            asistencias_v2, used_emp_field = _asistencias_v2_qs(contrato, fecha_inicio, fecha_fin)
            emp_field = used_emp_field or _emp_field_name()
        except Exception:
            logger.exception("Error al intentar obtener AsistenciaDiaria desde compat (mostrar_tareo_semanal)")
            asistencias_v2 = AsistenciaDiaria.objects.none()

        asist_dict = {}
        asistencias_v1 = AsistenciaTrabajador.objects.filter(
            trabajador__contrato=contrato,
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin,
        ).select_related('trabajador')
        for asist in asistencias_v1:
            asist_dict.setdefault(asist.trabajador_id, {})[asist.fecha] = {
                'estado': asist.estado,
                'guardia_snapshot': getattr(asist, 'guardia_snapshot', None) or '',
                'maquina_snapshot': getattr(asist.trabajador, 'maquina_asignada', None),
                'es_proyeccion': False,
            }
        if asistencias_v2.exists():
            for asist in asistencias_v2:
                emp_id = getattr(asist, f"{emp_field}_id")
                asist_dict.setdefault(emp_id, {})[asist.fecha] = {
                    'estado': getattr(asist, 'estado', '') or '',
                    'guardia_snapshot': getattr(asist, 'guardia_snapshot', None) or '',
                    'maquina_snapshot': getattr(asist, 'maquina_snapshot', None),
                    'es_proyeccion': getattr(asist, 'es_proyeccion', False) if hasattr(asist, 'es_proyeccion') else (getattr(asist, 'tipo', None) == 'PROY'),
                }

        # Estadísticas de verificación: cuántas filas son proyección vs reales
        try:
            if NEW_TAREO:
                model = getattr(asistencias_v2, 'model', None)
                has_tipo = False
                has_es_proy = False
                if model is not None:
                    fields = [f.name for f in model._meta.get_fields()]
                    has_tipo = 'tipo' in fields
                    has_es_proy = 'es_proyeccion' in fields

                if has_tipo:
                    proyecciones_count = asistencias_v2.filter(tipo='PROY').count()
                    reales_count = asistencias_v2.filter(tipo='REAL').count()
                elif has_es_proy:
                    proyecciones_count = asistencias_v2.filter(es_proyeccion=True).count()
                    reales_count = asistencias_v2.filter(es_proyeccion=False).count()
                else:
                    # Ultimate fallback: iterate instances and inspect attributes
                    proyecciones_count = 0
                    reales_count = 0
                    for a in asistencias_v2:
                        if getattr(a, 'tipo', None) == 'PROY' or getattr(a, 'es_proyeccion', False):
                            proyecciones_count += 1
                        else:
                            reales_count += 1
            else:
                proyecciones_count = asistencias_v2.filter(es_proyeccion=True).count()
                reales_count = asistencias_v2.filter(es_proyeccion=False).count()
        except Exception:
            logger.exception('Error counting proyecciones/reales in mostrar_tareo_semanal')
            proyecciones_count = reales_count = 0

        # Trabajadores activos
        trabajadores = _trabajadores_tareo_qs(contrato, fecha_inicio, fecha_fin).select_related('maquina_asignada')

        # Construir estructura por categorias -> maquinas -> guardias -> trabajadores
        import re as _re
        CAT_NOMBRES = {
            'LINEA_MANDO':          'Línea de Mando',
            'CONDUCTORES':          'Conductores',
            'OPERADORES':           'Operadores',
            '__STAND_BY__':         'Personal Stand By',
            'SERVICIOS_GEOLOGICOS': 'Servicios Geológicos',
            '__SIN_GRUPO__':        'Sin Grupo Asignado',
        }
        categorias = {}
        for trabajador in trabajadores.order_by('grupo', 'maquina_asignada__nombre', 'apepat'):
            tiene_maquina = bool(getattr(trabajador, 'maquina_asignada', None))
            if getattr(trabajador, 'es_standby', False) and not tiene_maquina:
                cat_key = '__STAND_BY__'
            else:
                cat_key = trabajador.grupo if trabajador.grupo else '__SIN_GRUPO__'

            if cat_key not in categorias:
                categorias[cat_key] = {'nombre': CAT_NOMBRES.get(cat_key, cat_key), 'maquinas': {}}

            maq = trabajador.maquina_asignada
            maq_key = maq.nombre if maq else '__SIN_MAQUINA__'
            maq_nombre = maq.nombre if maq else 'Sin Máquina'
            maq_css = 'maq-' + _re.sub(r'[^a-zA-Z0-9]', '-', maq_key)
            maquinas = categorias[cat_key]['maquinas']
            if maq_key not in maquinas:
                maquinas[maq_key] = {'nombre': maq_nombre, 'css': maq_css, 'guardias': {}}

            # guardia grouping will be determined from per-cell guardia_snapshot below

            # Construir asistencias por día para la semana
            dias = []
            d = fecha_inicio
            while d <= fecha_fin:
                asist = asist_dict.get(trabajador.id, {}).get(d)
                if asist:
                    codigo = MAPEO_CODIGOS.get(asist.get('estado', '') or '', '')
                    guardia_snap = asist.get('guardia_snapshot', '') or ''
                    maq_snap = asist.get('maquina_snapshot', None)
                    maq_snap_name = maq_snap.nombre if maq_snap else ''
                    es_proy = asist.get('es_proyeccion', False)
                else:
                    # si no existe registro, dejar vacío (puede usarse proyección)
                    codigo = ''
                    guardia_snap = ''
                    maq_snap_name = ''
                    es_proy = False
                color = COLORES_EXCEL.get(codigo, '') if codigo else ''
                # COLORES_EXCEL guarda valores ARGB (8 hex), pero CSS expects
                # #RRGGBB (or #RRGGBBAA). If the value is ARGB with full alpha
                # 'FF' prefix, strip it so template receives 'RRGGBB'. Also
                # tolerate 6-char values.
                if color:
                    c = color.strip()
                    if len(c) == 8 and c.upper().startswith('FF'):
                        c = c[2:]
                    elif len(c) == 8:
                        # If it's AARRGGBB, convert to RRGGBB by dropping first two
                        c = c[2:]
                    # ensure we send 6-char hex (fallback to empty if malformed)
                    color = c if len(c) == 6 else ''
                dias.append({'fecha': d, 'codigo': codigo, 'guardia_snapshot': guardia_snap, 'color': color, 'maquina_snapshot': maq_snap_name, 'es_proyeccion': es_proy})
                d += timedelta(days=1)

            # Determinar si hay una máquina asignada en las celdas (maquina_snapshot)
            maq_from_cells = [x['maquina_snapshot'] for x in dias if x.get('maquina_snapshot')]
            if maq_from_cells:
                # elegir la máquina más frecuente entre las celdas
                from collections import Counter
                most_common = Counter(maq_from_cells).most_common(1)
                if most_common:
                    maq_name_from_cells = most_common[0][0]
                else:
                    maq_name_from_cells = ''
            else:
                maq_name_from_cells = ''

            # Guardia es fija por trabajador: siempre usar guardia_asignada para agrupar.
            # La máquina puede variar por día según maquina_snapshot.
            assigned_maq_name = trabajador.maquina_asignada.nombre if getattr(trabajador, 'maquina_asignada', None) else '__SIN_MAQUINA__'
            assigned_guardia = trabajador.guardia_asignada if trabajador.guardia_asignada else 'SIN_GUARDIA'

            comb_map = {}
            # initialize empty week placeholders when creating a new combo
            def empty_week():
                return [ {'fecha': fecha_inicio + timedelta(days=i), 'codigo':'', 'guardia_snapshot':'', 'color':'', 'maquina_snapshot': ''} for i in range(7) ]

            for idx, day in enumerate(dias):
                day_maq = day.get('maquina_snapshot') or assigned_maq_name
                # La guardia NO cambia por día: siempre se agrupa por guardia_asignada del trabajador.
                # El guardia_snapshot es solo metadata histórica por celda, no define el grupo.
                day_guardia = assigned_guardia
                combo = (day_maq, day_guardia)
                if combo not in comb_map:
                    comb_map[combo] = empty_week()
                # place the day's data into the correct slot
                comb_map[combo][idx] = day

            maquinas = categorias[cat_key]['maquinas']
            for (mname, gname), weekdias in comb_map.items():
                maq_key = mname if mname else '__SIN_MAQUINA__'
                maq_nombre = mname if mname else 'Sin Máquina'
                maq_css = 'maq-' + _re.sub(r'[^a-zA-Z0-9]', '-', maq_key)
                if maq_key not in maquinas:
                    maquinas[maq_key] = {'nombre': maq_nombre, 'css': maq_css, 'guardias': {}}

                gmap = maquinas[maq_key]['guardias']
                guardia_key = gname if gname else (trabajador.guardia_asignada or 'SIN_GUARDIA')
                if guardia_key not in gmap:
                    gmap[guardia_key] = []

                gmap[guardia_key].append({
                    'trabajador': trabajador,
                    'dias': weekdias,
                })

        # Pasar a lista ordenada para el template
        preferred_order = ['LINEA_MANDO', 'CONDUCTORES', 'OPERADORES', '__STAND_BY__', 'SERVICIOS_GEOLOGICOS', '__SIN_GRUPO__']
        categorias_list = []
        # Añadir en el orden preferido si existen
        for key in preferred_order:
            if key in categorias:
                v = categorias[key]
                maquinas_list = []
                for mk, md in v['maquinas'].items():
                    guardias_list = []
                    for gk, tlist in md['guardias'].items():
                        guardias_list.append({'key': gk, 'trabajadores': tlist})
                    total_workers = sum(len(tlist) for tlist in md['guardias'].values())
                    maquinas_list.append({'key': mk, 'nombre': md['nombre'], 'css': md['css'], 'guardias': guardias_list, 'total_workers': total_workers})
                categorias_list.append({'key': key, 'nombre': v['nombre'], 'maquinas': maquinas_list})

        # Añadir el resto de categorías que no están en preferred_order, ordenadas por key
        for k in sorted([kk for kk in categorias.keys() if kk not in preferred_order]):
            v = categorias[k]
            maquinas_list = []
            for mk, md in v['maquinas'].items():
                guardias_list = []
                for gk, tlist in md['guardias'].items():
                    guardias_list.append({'key': gk, 'trabajadores': tlist})
                maquinas_list.append({'key': mk, 'nombre': md['nombre'], 'css': md['css'], 'guardias': guardias_list})
            categorias_list.append({'key': k, 'nombre': v['nombre'], 'maquinas': maquinas_list})

        context = {
            'contrato': contrato,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'categorias': categorias_list,
            'dias': [fecha_inicio + timedelta(days=i) for i in range(7)],
            'proyecciones_count': proyecciones_count,
            'reales_count': reales_count,
        }

        return render(request, 'drilling/tareo/tareo_semanal_vista.html', context)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error generando vista semanal: {str(e)}', status=500)


def _crear_hoja_distribucion(ws, contrato, fecha_inicio, fecha_fin):
    """Crea una hoja distribuida por máquina y guardia similar al cartel."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'].value = f"DISTRIBUCIÓN SEMANAL: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Cabecera de columnas simples
    ws.cell(row=2, column=1).value = 'MÁQUINA / GUARDIA'
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # Cols por día (dentro de cada guardia mostraremos nombres)
    dias = []
    fecha = fecha_inicio
    col_base = 2
    while fecha <= fecha_fin:
        cell = ws.cell(row=2, column=col_base)
        cell.value = fecha.strftime('%a %d')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        col_base += 1
        dias.append(fecha)
        fecha += timedelta(days=1)

    # Obtener asistencias V1 y V2 para el rango, priorizando V2 cuando exista.
    from .tareo_compat import AsistenciaDiaria
    emp_field = _emp_field_name()
    distrib = {}

    asist_v1 = AsistenciaTrabajador.objects.filter(
        trabajador__contrato=contrato,
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin,
    ).select_related('trabajador', 'trabajador__maquina_asignada')
    for a in asist_v1:
        maq = getattr(a.trabajador, 'maquina_asignada', None)
        maq_key = maq.nombre if maq else '__SIN_MAQUINA__'
        guardia = getattr(a, 'guardia_snapshot', None) or (getattr(a.trabajador, 'guardia_asignada', None) or 'SIN_GUARDIA')
        distrib.setdefault(maq_key, {}).setdefault(guardia, {}).setdefault(a.fecha, []).append(a.trabajador)

    asist_v2 = AsistenciaDiaria.objects.filter(
        **{f"{emp_field}__contrato": contrato, 'fecha__gte': fecha_inicio, 'fecha__lte': fecha_fin}
    ).select_related(emp_field, 'maquina_snapshot')
    for a in asist_v2:
        maq = getattr(a, 'maquina_snapshot', None) or getattr(getattr(a, emp_field), 'maquina_asignada', None)
        maq_key = maq.nombre if maq else '__SIN_MAQUINA__'
        trabajador_obj = getattr(a, emp_field)
        guardia = getattr(a, 'guardia_snapshot', None) or (getattr(trabajador_obj, 'guardia_asignada', None) or 'SIN_GUARDIA')
        bucket = distrib.setdefault(maq_key, {}).setdefault(guardia, {}).setdefault(a.fecha, [])
        bucket = [t for t in bucket if getattr(t, 'id', None) != getattr(trabajador_obj, 'id', None)]
        bucket.append(trabajador_obj)
        distrib[maq_key][guardia][a.fecha] = bucket

    # Ordenar máquinas
    sorted_maqs = sorted(distrib.items(), key=lambda kv: (kv[0] == '__SIN_MAQUINA__', kv[0]))
    row = 3
    for maq_key, maq_data in sorted_maqs:
        # Encabezado máquina
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(dias))
        ws.cell(row=row, column=1).value = (maq_key if maq_key != '__SIN_MAQUINA__' else 'Sin Máquina Asignada')
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Por guardia
        for guardia_key in sorted(maq_data.keys()):
            ws.cell(row=row, column=1).value = f"Guardia: {guardia_key}"
            ws.cell(row=row, column=1).font = Font(bold=True)
            col = 2
            for d in dias:
                trabajadores = maq_data.get(guardia_key, {}).get(d, [])
                if trabajadores:
                    names = '\n'.join([f"{t.apellidos}, {t.nombres}" for t in trabajadores])
                    ws.cell(row=row, column=col).value = names
                else:
                    ws.cell(row=row, column=col).value = ''
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
                col += 1
            row += 1

        # Separador
        row += 1

    # Ajustes de ancho
    ws.column_dimensions['A'].width = 30
    for i in range(2, 2 + len(dias)):
        ws.column_dimensions[get_column_letter(i)].width = 20


@login_required
@require_http_methods(["POST"])
@login_required
@require_http_methods(["POST"])
def limpiar_asistencias_mes(request):
    """
    Limpia todas las asistencias del contrato para el rango de fechas especificado.
    Elimina registros de AsistenciaTrabajador excepto los estados protegidos si se solicita.
    """
    user = request.user
    if not user.can_manage_contract_users():
        return JsonResponse({'success': False, 'message': 'Sin permisos'}, status=403)
        
    try:
        data = json.loads(request.body)
        contrato_id = data.get('contrato_id')
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')
        mantener_protegidos = data.get('mantener_protegidos', True)
        
        print(f"DEBUG Limpiar mes - Contrato: {contrato_id}, Fechas: {fecha_inicio_str} a {fecha_fin_str}")
        
        if not all([contrato_id, fecha_inicio_str, fecha_fin_str]):
            return JsonResponse({'success': False, 'message': 'Faltan datos requeridos'}, status=400)
            
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        
        # Validar periodo (máximo 40 días para evitar limpiezas masivas accidentales)
        dias = (fecha_fin - fecha_inicio).days + 1
        print(f"DEBUG - Días en el rango: {dias}")
        
        if dias > 40:
             return JsonResponse({'success': False, 'message': f'Rango de fechas demasiado amplio ({dias} días, máx 40)'}, status=400)
             
        query = AsistenciaTrabajador.objects.filter(
            trabajador__contrato_id=contrato_id,
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin
        )
        
        total_antes = query.count()
        print(f"DEBUG - Total registros antes de filtrar: {total_antes}")
        
        if mantener_protegidos:
            ESTADOS_PROTEGIDOS = ['VACACIONES', 'DESCANSO_MEDICO', 'LICENCIA', 'PERMISO', 'SUBSIDIO']
            query = query.exclude(estado__in=ESTADOS_PROTEGIDOS)
            
        total_a_eliminar = query.count()
        print(f"DEBUG - Total a eliminar (sin protegidos): {total_a_eliminar}")
        
        count, _ = query.delete()
        
        print(f"DEBUG - Registros eliminados: {count}")
        
        return JsonResponse({
            'success': True, 
            'message': f'Se eliminaron {count} de {total_antes} registros de asistencia.'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def auto_rellenar_turno_view(request):
    """
    Retorna trabajadores sugeridos para un turno en fecha y guardia específica.
    Se basa en la asistencia ya registrada (tareo).
    Si el trabajador tiene asistencia 'TRABAJADO' y coincide la guardia, se sugiere.
    """
    try:
        data = json.loads(request.body)
        contrato_id = data.get('contrato_id')
        fecha_str = data.get('fecha')
        guardia = data.get('guardia')
        
        if not all([contrato_id, fecha_str, guardia]):
            return JsonResponse({'success': False, 'message': 'Faltan datos'}, status=400)
            
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # 1. Obtener trabajadores que tengan asistencia TRABAJADO en esa fecha
        asistencias = AsistenciaTrabajador.objects.filter(
            trabajador__contrato_id=contrato_id,
            fecha=fecha,
            estado='TRABAJADO' # Solo los que han ido a trabajar
        ).select_related('trabajador')
        
        # 2. Filtrar por guardia.
        # La guardia se puede tomar de la asistencia (parametro 'turno') o del perfil del trabajador.
        # Prioridad: turno en asistencia > perfil trabajador.
        
        trabajadores_list = []
        for a in asistencias:
            t = a.trabajador
            
            # Verificar guardia
            guardia_trabajador = a.turno if a.turno else t.guardia_asignada
            
            # Si coincide la guardia, agregar
            if guardia_trabajador == guardia:
                trabajadores_list.append({
                    'id': t.id,
                    'nombres': t.nombres,
                    'apellidos': t.apellidos,
                    'cargo': t.cargo or '',
                    'observaciones': ''
                })
        
        # Si no hay asistencias registradas (quizás no se ha hecho el tareo aún), 
        # fallback a sugerir según asignación estática del perfil
        if not trabajadores_list:
            trabajadores_perfil = Trabajador.objects.filter(
                contrato_id=contrato_id,
                estado='ACTIVO',
                subestado='EN_OPERACION',
                guardia_asignada=guardia
            )
            
            for t in trabajadores_perfil:
                # Verificar si hoy le toca trabajar según régimen (opcional, para no sugerir en días libres)
                # estado_calc = t.calcular_estado_regimen(fecha)
                # if estado_calc != 'TRABAJADO': continue 
                
                trabajadores_list.append({
                    'id': t.id,
                    'nombres': t.nombres,
                    'apellidos': t.apellidos,
                    'cargo': t.cargo or '',
                    'observaciones': 'Sugerido por perfil'
                })

        return JsonResponse({
            'success': True, 
            'trabajadores': trabajadores_list
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def auto_rellenar_asistencia(request):
    """
    Auto-rellena la asistencia basada en el régimen laboral.
    Sobrescribe estados 'TRABAJADO', 'DIA_LIBRE', 'FALTA' o vacíos.
    Respeta licencias, vacaciones, descansos médicos.
    """
    user = request.user
    if not user.can_manage_contract_users():
        return JsonResponse({'success': False, 'message': 'Sin permisos'}, status=403)

    try:
        data = json.loads(request.body)
        contrato_id = data.get('contrato_id')
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')

        if not all([contrato_id, fecha_inicio_str, fecha_fin_str]):
            return JsonResponse({'success': False, 'message': 'Faltan datos'}, status=400)

        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

        # Obtener trabajadores
        trabajadores = Trabajador.objects.filter(
            contrato_id=contrato_id,
            estado='ACTIVO'
        )

        count_updated = 0
        
        # Estados que NO se deben sobrescribir (protegidos)
        ESTADOS_PROTEGIDOS = ['VACACIONES', 'DESCANSO_MEDICO', 'LICENCIA', 'PERMISO', 'SUBSIDIO']

        with transaction.atomic():
            # 1. ASIGNACIÓN AUTOMÁTICA DE GUARDIAS (BALANCEO)
            # Si hay trabajadores sin guardia, asignarlos para balancear A, B, C
            # EXCEPCIÓN: No asignar guardia a LINEA_MANDO
            trabajadores_sin_guardia = [
                t for t in trabajadores 
                if not t.guardia_asignada and not any(c in (t.cargo or '').upper() for c in ('RESIDENTE','SUPERVISOR','JEFE','ADMINISTRADOR','INGENIERO','GERENTE'))
            ]
            
            if trabajadores_sin_guardia:
                # Contar actuales (excluyendo linea de mando)
                conteos = {'A': 0, 'B': 0, 'C': 0}
                for t in trabajadores:
                    if t.guardia_asignada in conteos and not any(c in (t.cargo or '').upper() for c in ('RESIDENTE','SUPERVISOR','JEFE','ADMINISTRADOR','INGENIERO','GERENTE')):
                        conteos[t.guardia_asignada] += 1
                
                # Asignar round-robin a la guardia con menos gente
                for t in trabajadores_sin_guardia:
                    # Encontrar guardia con menor conteo
                    min_guardia = min(conteos, key=conteos.get)
                    t.guardia_asignada = min_guardia
                    t.save(update_fields=['guardia_asignada'])
                    conteos[min_guardia] += 1

            # 2. GENERACIÓN DE ASISTENCIA CON ROTACIÓN
            for trabajador in trabajadores:
                # Iterar días
                delta = (fecha_fin - fecha_inicio).days + 1
                for i in range(delta):
                    fecha = fecha_inicio + timedelta(days=i)
                    
                    # Calcular estado según régimen y guardia (ROTACIÓN 14x7)
                    estado_regimen = None
                    
                    # CASO ESPECIAL: LINEA DE MANDO (Siempre TRABAJADO, excepto domingos si aplica, o según régimen simple)
                    if any(c in (trabajador.cargo or '').upper() for c in ('RESIDENTE','SUPERVISOR','JEFE','ADMINISTRADOR','INGENIERO','GERENTE')):
                        # Asumimos TRABAJADO por defecto para línea de mando, o lógica simple
                        # Si tienen régimen, intentar respetarlo, pero sin offsets de guardia
                        if trabajador.regimen_laboral:
                             try:
                                dias_trabajo, dias_descanso = map(int, trabajador.regimen_laboral.lower().split('x'))
                                # Para linea de mando, a veces es continuo. 
                                # Si es 14x7 pero sin guardia, ¿cómo sabemos cuándo empieza?
                                # Usamos fecha_inicio_ciclo si existe, sino asumimos TRABAJADO siempre
                                if trabajador.fecha_inicio_ciclo:
                                    ciclo_total = dias_trabajo + dias_descanso
                                    delta_ciclo = (fecha - trabajador.fecha_inicio_ciclo).days
                                    if delta_ciclo >= 0:
                                        dia_en_ciclo = delta_ciclo % ciclo_total
                                        if dia_en_ciclo < dias_trabajo:
                                            estado_regimen = 'TRABAJADO'
                                        else:
                                            estado_regimen = 'DIA_LIBRE'
                                else:
                                    # Si no hay fecha inicio ciclo, asumir TRABAJADO siempre (ej. régimen administrativo)
                                    estado_regimen = 'TRABAJADO'
                             except:
                                 estado_regimen = 'TRABAJADO'
                        else:
                            estado_regimen = 'TRABAJADO'

                    # Lógica específica para 14x7 con rotación de 3 guardias (SOLO SI TIENE GUARDIA)
                    elif trabajador.regimen_laboral == '14x7' and trabajador.guardia_asignada:
                        # Use deterministic TareoEngine to compute TD/TN/DL with guard offsets
                        from .utils.tareo_service import TareoEngine
                        try:
                            estado_tareo = TareoEngine.estado_para_fecha(trabajador, fecha)
                            # Map engine states to asistencia states
                            if estado_tareo in ('TD', 'TN'):
                                estado_regimen = 'TRABAJADO'
                            else:
                                estado_regimen = 'DIA_LIBRE'
                        except Exception:
                            # Fallback: si falla el motor, marcar TRABAJADO
                            estado_regimen = 'TRABAJADO'
                            
                    # Fallback para otros regímenes (lógica simple inicio mes)
                    elif trabajador.regimen_laboral:
                        try:
                            dias_trabajo, dias_descanso = map(int, trabajador.regimen_laboral.lower().split('x'))
                            ciclo_total = dias_trabajo + dias_descanso
                            inicio_mes = fecha.replace(day=1)
                            delta_dias = (fecha - inicio_mes).days
                            dia_en_ciclo = delta_dias % ciclo_total
                            
                            if dia_en_ciclo < dias_trabajo:
                                estado_regimen = 'TRABAJADO'
                            else:
                                estado_regimen = 'DIA_LIBRE'
                        except ValueError:
                            pass

                    if not estado_regimen:
                        continue 

                    # Verificar estado actual
                    asistencia, created = AsistenciaTrabajador.objects.get_or_create(
                        trabajador=trabajador,
                        fecha=fecha,
                        defaults={
                            'estado': estado_regimen,
                            'registrado_por': user
                        }
                    )

                    if created:
                        count_updated += 1
                    else:
                        # Si ya existe, verificar si es sobrescribible
                        if asistencia.estado not in ESTADOS_PROTEGIDOS:
                            if asistencia.estado != estado_regimen:
                                asistencia.estado = estado_regimen
                                asistencia.save()
                                count_updated += 1
        
        # --- VALIDACIÓN DE COBERTURA ---
        # "SIEMPRE debemos de tener 1 perforista y por lo menos 1 ayudante" por turno (Guardia)
        
        alertas = []
        
        # Consultar asistencias del rango para verificar
        asistencias = AsistenciaTrabajador.objects.filter(
            trabajador__contrato_id=contrato_id,
            fecha__range=[fecha_inicio, fecha_fin],
            estado='TRABAJADO'
        ).select_related('trabajador')

        # Agrupar por Fecha -> Guardia
        cobertura = {}
        for asist in asistencias:
            fecha_str = asist.fecha.strftime('%Y-%m-%d')
            guardia = asist.trabajador.guardia_asignada
            if not guardia:
                continue # Ignorar si no tiene guardia asignada
            
            key = f"{fecha_str}|{guardia}"
            if key not in cobertura:
                cobertura[key] = {'perforistas': 0, 'ayudantes': 0}
            
            cargo_nombre = (asist.trabajador.cargo or '').upper()
            if 'PERFORISTA' in cargo_nombre and 'AYUDANTE' not in cargo_nombre:
                cobertura[key]['perforistas'] += 1
            elif 'AYUDANTE' in cargo_nombre:
                cobertura[key]['ayudantes'] += 1

        # Verificar mínimos
        for key, counts in cobertura.items():
            fecha_s, guardia = key.split('|')
            perf = counts['perforistas']
            ayu = counts['ayudantes']
            
            if perf < 1 or ayu < 1:
                # Formatear fecha para mensaje
                f_obj = datetime.strptime(fecha_s, '%Y-%m-%d')
                f_fmt = f_obj.strftime('%d/%m')
                alertas.append(f"Día {f_fmt} Guardia {guardia}: {perf} Perf. / {ayu} Ayud. (Mínimo 1 Perf + 1 Ayud)")

        msg = f'Se actualizaron {count_updated} registros.'
        if alertas:
            msg += " ADVERTENCIA DE COBERTURA: " + "; ".join(alertas[:5])
            if len(alertas) > 5:
                msg += f" ... y {len(alertas)-5} más."

        return JsonResponse({
            'success': True, 
            'message': msg,
            'alertas': alertas
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def actualizar_grupos_trabajadores(request):
    """
    Solo lee el estado actual de los trabajadores y devuelve un resumen.
    NO modifica ningún dato — la tabla del tareo se reconstruye con un reload
    usando los grupos y el flag es_standby ya guardados en cada trabajador.
    """
    try:
        contrato_id = request.POST.get('contrato_id')

        qs = Trabajador.objects.filter(estado='ACTIVO')
        if contrato_id:
            qs = qs.filter(contrato_id=contrato_id)

        stats = {}
        standby_count = 0
        for t in qs.only('grupo', 'es_standby', 'maquina_asignada'):
            if t.es_standby and not t.maquina_asignada_id:
                standby_count += 1
                key = 'Stand By'
            else:
                key = t.grupo or 'Sin Grupo'
            stats[key] = stats.get(key, 0) + 1

        total = sum(stats.values())
        detalles = ", ".join([f"{k}: {v}" for k, v in sorted(stats.items())])
        message = f'Tareo actualizado — {total} trabajadores activos. Distribución: {detalles}.'

        return JsonResponse({'success': True, 'message': message})
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'message': str(e) + '\n' + traceback.format_exc()}, status=500)

@login_required
def debug_trabajadores(request):
    """Vista de depuración para listar trabajadores y sus grupos"""
    if not request.user.is_staff:
        return HttpResponseForbidden("Solo staff")
        
    contrato_id = request.GET.get('contrato')
    if contrato_id:
        trabajadores = Trabajador.objects.filter(contrato_id=contrato_id).select_related('contrato').order_by('grupo', 'apepat', 'apemat')
    else:
        trabajadores = Trabajador.objects.all().select_related('contrato').order_by('contrato', 'grupo', 'apepat', 'apemat')
        
    html = """
    <html>
    <head>
        <title>Debug Trabajadores</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="p-4">
        <h1>Debug Trabajadores</h1>
        <table class="table table-striped table-bordered table-sm">
            <thead class="table-dark">
                <tr>
                    <th>ID</th>
                    <th>Contrato</th>
                    <th>Nombres</th>
                    <th>Apellidos</th>
                    <th>Cargo</th>
                    <th>Grupo (DB)</th>
                    <th>Grupo (Calculado)</th>
                    <th>Estado</th>
                    <th>Guardia</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for t in trabajadores:
        grupo_calc = t.asignar_grupo_automatico()
        match_style = "text-success" if t.grupo == grupo_calc else "text-danger fw-bold"
        
        html += f"""
        <tr>
            <td>{t.id}</td>
            <td>{t.contrato.nombre_contrato if t.contrato else '-'}</td>
            <td>{t.nombres}</td>
            <td>{t.apellidos}</td>
            <td>{t.cargo or '-'}</td>
            <td>{t.grupo or '-'}</td>
            <td class="{match_style}">{grupo_calc}</td>
            <td>{t.estado}</td>
            <td>{t.guardia_asignada}</td>
        </tr>
        """
        
    html += """
            </tbody>
        </table>
    </body>
    </html>
    """
    return HttpResponse(html)


@login_required
@require_http_methods(["POST"])
@login_required
@require_http_methods(["POST"])
def generar_guardias_automaticas(request):
    """
    Genera y asigna guardias A, B, C automáticamente con lógica flexible:
    - EXCLUYE personal marcado como STANDBY (personal de reserva)
    - Distribuye perforistas y ayudantes proporcionalmente según disponibilidad
    - Si no hay suficiente para 3 guardias completas, forma las que sean posibles
    - Línea de mando no recibe guardias (trabajan independiente)
    
    Composición ideal por guardia: 1 perforista + 2 ayudantes
    Pero se adapta al personal disponible
    """
    user = request.user
    
    if not user.can_manage_contract_users():
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
    
    # Determinar contrato
    if user.has_access_to_all_contracts():
        contrato_id = request.POST.get('contrato_id')
        if not contrato_id:
            return JsonResponse({'success': False, 'error': 'Debe especificar contrato'}, status=400)
        try:
            contrato = Contrato.objects.get(id=contrato_id)
        except Contrato.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Contrato no encontrado'}, status=404)
    else:
        contrato = user.contrato
        if not contrato:
            return JsonResponse({'success': False, 'error': 'Usuario sin contrato asignado'}, status=400)
    
    try:
        # Separar trabajadores por tipo de cargo
        perforistas = []
        ayudantes = []
        otros_operadores = []
        personal_standby = []
        
        trabajadores = Trabajador.objects.filter(
            contrato=contrato,
            estado='ACTIVO'
        ).exclude(grupo='LINEA_MANDO').order_by('cargo', 'apepat', 'nombres')
        
        if not trabajadores.exists():
            return JsonResponse({
                'success': False,
                'error': 'No hay trabajadores para asignar guardias'
            }, status=400)
        
        # Clasificar trabajadores según cargo y si son STANDBY
        for trabajador in trabajadores:
            # Separar personal STANDBY
            if trabajador.es_standby:
                personal_standby.append(trabajador)
                continue
            
            cargo_upper = (trabajador.cargo or '').upper()
            
            # Identificar PERFORISTAS
            if 'PERFORISTA' in cargo_upper and 'AYUDANTE' not in cargo_upper:
                perforistas.append(trabajador)
            # Identificar AYUDANTES
            elif 'AYUDANTE' in cargo_upper:
                ayudantes.append(trabajador)
            # Otros (auxiliares, conductores, etc.)
            else:
                otros_operadores.append(trabajador)
        
        num_perforistas = len(perforistas)
        num_ayudantes = len(ayudantes)
        num_otros = len(otros_operadores)
        num_standby = len(personal_standby)
        
        # Validar que haya al menos algo para asignar
        if num_perforistas == 0 and num_ayudantes == 0:
            return JsonResponse({
                'success': False,
                'error': 'No hay perforistas ni ayudantes activos para asignar guardias (excluidos STANDBY).',
                'personal_standby': num_standby
            }, status=400)
        
        # Calcular cuántas guardias podemos formar
        # Idealmente: 1 perforista + 2 ayudantes por guardia (maqueta completa)
        # Pero flexible: Aceptamos formar con menos si falta gente (priorizar A, B, C)
        
        # Siempre intentamos formar 3 guardias si hay suficiente personal mínimo
        num_guardias = 3
        
        # Validar si realmente tenemos CERO capacidad para 3
        if num_perforistas < 1 and num_ayudantes < 1:
            num_guardias = 0
        elif num_perforistas < 3 and num_ayudantes < 3:
             # Si hay muy poco personal total, reducir guardias (solo si es extremo)
             # Ej: 1 Perforista total -> 1 Guardia
             # Ej: 2 Perforistas total -> 2 Guardias
             # Pero si hay 3 perforistas, o 2P + 3A, intentamos estirar a 3
             if num_perforistas > 0:
                 num_guardias = min(3, num_perforistas) 
             else:
                 # Solo ayudantes
                 num_guardias = min(3, max(1, num_ayudantes // 2))
        
        # Forzar 3 guardias si el usuario lo pide implícitamente (lógica de negocio habitual)
        # Salvo que sea absurdo (ej: 1 sola persona)
        if (num_perforistas + num_ayudantes) >= 3:
             num_guardias = 3

        guardias = ['A', 'B', 'C'][:num_guardias]  # Guardias activas
        
        asignados = 0
        distribucion = {g: 0 for g in guardias}
        detalles = {
            'perforistas': {g: 0 for g in guardias}, 
            'ayudantes': {g: 0 for g in guardias}, 
            'otros': {g: 0 for g in guardias}
        }
        
        with transaction.atomic():
            # 1. Asignar PERFORISTAS de forma cíclica (A, B, C, A, B...)
            for i, perforista in enumerate(perforistas):
                guardia = guardias[i % len(guardias)]
                perforista.guardia_asignada = guardia
                perforista.save(update_fields=['guardia_asignada'])
                asignados += 1
                distribucion[guardia] += 1
                detalles['perforistas'][guardia] += 1
            
            # 2. Asignar AYUDANTES de forma cíclica (repartir equitativamente)
            # Antes: 2xA, 2xB... (llenado agresivo)
            # Ahora: A, B, C, A, B, C... (reparto equilibrado para cubrir mínimos)
            for i, ayudante in enumerate(ayudantes):
                guardia = guardias[i % len(guardias)]
                ayudante.guardia_asignada = guardia
                ayudante.save(update_fields=['guardia_asignada'])
                asignados += 1
                distribucion[guardia] += 1
                detalles['ayudantes'][guardia] += 1
            
            # 3. Asignar OTROS de forma equitativa
            for i, trabajador in enumerate(otros_operadores):
                guardia = guardias[i % len(guardias)]
                trabajador.guardia_asignada = guardia
                trabajador.save(update_fields=['guardia_asignada'])
                asignados += 1
                distribucion[guardia] += 1
                detalles['otros'][guardia] += 1
            
            # 4. Limpiar guardias del personal STANDBY (no tienen guardia fija)
            for trabajador in personal_standby:
                if trabajador.guardia_asignada:
                    trabajador.guardia_asignada = None
                    trabajador.save(update_fields=['guardia_asignada'])
        
        # Preparar mensaje informativo
        mensaje = f'✅ Guardias asignadas exitosamente'
        advertencias = []
        
        if num_guardias < 3:
            advertencias.append(f'⚠️ Solo se formaron {num_guardias} guardia(s) por personal limitado')
        
        if num_standby > 0:
            advertencias.append(f'ℹ️ {num_standby} trabajador(es) STANDBY excluidos (sin guardia fija)')
        
        # Verificar composición
        for g in guardias:
            perf = detalles['perforistas'][g]
            ayud = detalles['ayudantes'][g]
            if perf == 0:
                advertencias.append(f'⚠️ Guardia {g} sin perforistas')
            if ayud < 2:
                advertencias.append(f'⚠️ Guardia {g} con solo {ayud} ayudante(s)')
        
        if advertencias:
            mensaje += '\n\n' + '\n'.join(advertencias)
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'asignados': asignados,
            'guardias_formadas': num_guardias,
            'distribucion_total': distribucion,
            'detalles': detalles,
            'resumen': {
                'perforistas_totales': num_perforistas,
                'ayudantes_totales': num_ayudantes,
                'otros_totales': num_otros,
                'personal_standby': num_standby
            },
            'advertencias': advertencias
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al generar guardias: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def autocompletar_tareo_por_regimen(request):
    """
    Autocompleta el tareo del mes según el régimen laboral de cada trabajador.
    
    Lógica:
    - Calcula días de trabajo/descanso según régimen (14x7, 20x10, etc.)
    - Asegura que siempre hay 2 guardias activas cada día
    - Rota las guardias para cumplir con sus regímenes
    - Línea de mando trabaja todos los días hábiles
    """
    user = request.user
    
    if not user.can_manage_contract_users():
        return JsonResponse({'success': False, 'message': 'Sin permisos'}, status=403)
    
    try:
        # Obtener datos de POST (FormData)
        contrato_id = request.POST.get('contrato_id')
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_fin_str = request.POST.get('fecha_fin')
        
        if not fecha_inicio_str or not fecha_fin_str:
            return JsonResponse({'success': False, 'message': 'Fechas no especificadas'}, status=400)
        
        # Parsear fechas
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        
        # Determinar contrato
        if user.has_access_to_all_contracts():
            if not contrato_id:
                return JsonResponse({'success': False, 'message': 'Debe especificar contrato'}, status=400)
            contrato = Contrato.objects.get(id=contrato_id)
        else:
            contrato = user.contrato
        
        # Obtener trabajadores activos
        trabajadores = Trabajador.objects.filter(
            contrato=contrato,
            estado='ACTIVO'
        )
        
        registros_creados = 0
        registros_actualizados = 0
        trabajadores_sin_ciclo = 0
        
        with transaction.atomic():
            # Recalcular fecha_inicio_ciclo para los trabajadores NO línea de mando.
            # El desfase entre guardias = ciclo_total / 3 (proporcional al régimen).
            # Ejemplo: 14x7 → ciclo 21 → desfase 7 días | 20x10 → ciclo 30 → desfase 10 días
            # Esto garantiza que siempre haya 2 guardias trabajando y 1 descansando.
            REGIMENES_CICLO = {
                '14x7':  21,
                '20x10': 30,
                '28x14': 42,
                '5x2':   7,
                '6x1':   7,
            }

            # Anclar el inicio del ciclo al día de cambio de guardia del contrato
            # más cercano anterior o igual a fecha_inicio.
            # Así el ciclo siempre arranca en un día de cambio real (ej: viernes para Cuculí).
            dia_cambio = contrato.dia_cambio_guardia if contrato.dia_cambio_guardia is not None else 6
            dias_hasta_cambio = (fecha_inicio.weekday() - dia_cambio) % 7
            fecha_ancla_ciclo = fecha_inicio - timedelta(days=dias_hasta_cambio)

            for trabajador in trabajadores:
                if trabajador.grupo != 'LINEA_MANDO':
                    fecha_anterior = trabajador.fecha_inicio_ciclo

                    # Calcular desfase proporcional al ciclo del trabajador
                    ciclo_total = REGIMENES_CICLO.get(trabajador.regimen_laboral or '', 21)
                    desfase = ciclo_total // 3  # 1/3 del ciclo por cada guardia

                    # Recalcular inicio de ciclo según guardia, anclado al día de cambio
                    if trabajador.guardia_asignada == 'A':
                        nueva_fecha_ciclo = fecha_ancla_ciclo
                    elif trabajador.guardia_asignada == 'B':
                        nueva_fecha_ciclo = fecha_ancla_ciclo - timedelta(days=desfase)
                    elif trabajador.guardia_asignada == 'C':
                        nueva_fecha_ciclo = fecha_ancla_ciclo - timedelta(days=desfase * 2)
                    else:
                        nueva_fecha_ciclo = fecha_ancla_ciclo
                    
                    if fecha_anterior != nueva_fecha_ciclo:
                        # Usar QuerySet.update() para evitar disparar el save() hook
                        # (que llama a asignar_grupo_automatico() y sobrescribe el grupo manual)
                        Trabajador.objects.filter(pk=trabajador.pk).update(fecha_inicio_ciclo=nueva_fecha_ciclo)
                        trabajador.fecha_inicio_ciclo = nueva_fecha_ciclo  # actualizar instancia en memoria
                        trabajadores_sin_ciclo += 1
            
            # Recorrer cada día del mes
            fecha_actual = fecha_inicio
            while fecha_actual <= fecha_fin:
                
                # Para cada trabajador, determinar su estado ese día
                for trabajador in trabajadores:
                    # Respetar ventana vigente del trabajador en el tareo
                    if not trabajador.vigente_en_fecha(fecha_actual):
                        continue

                    # Intenta calcular según régimen configurado (para TODOS, incluida Línea de Mando)
                    estado = trabajador.calcular_estado_regimen(fecha_actual)
                    
                    # Si no tiene régimen configurado (return None), aplicar defaults por grupo
                    if not estado:
                        if trabajador.grupo == 'LINEA_MANDO':
                            # Default para línea de mando sin régimen: Lunes a Viernes
                            estado = 'TRABAJADO' if fecha_actual.weekday() < 5 else 'DIA_LIBRE'
                        else:
                            # Default para operativos sin régimen: Siempre Trabajo
                            estado = 'TRABAJADO'
                    
                    # Verificar si ya existe registro
                    asistencia, created = AsistenciaTrabajador.objects.get_or_create(
                        trabajador=trabajador,
                        fecha=fecha_actual,
                        defaults={
                            'estado': estado,
                            'guardia_snapshot': trabajador.guardia_asignada,
                            'cargo_snapshot': trabajador.cargo or '',
                            'registrado_por': user
                        }
                    )
                    
                    if created:
                        registros_creados += 1
                    else:
                        # Actualizar si cambió el estado
                        if asistencia.estado != estado:
                            asistencia.estado = estado
                            asistencia.guardia_snapshot = trabajador.guardia_asignada
                            asistencia.cargo_snapshot = trabajador.cargo or ''
                            asistencia.save()
                            registros_actualizados += 1
                
                fecha_actual += timedelta(days=1)
        
        mensaje = f'✅ Tareo autocompletado: {registros_creados} registros creados, {registros_actualizados} actualizados'
        if trabajadores_sin_ciclo > 0:
            mensaje += f' ({trabajadores_sin_ciclo} trabajadores inicializados con fecha de ciclo)'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'creados': registros_creados,
            'actualizados': registros_actualizados,
            'inicializados': trabajadores_sin_ciclo,
            'periodo': f'{fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Error al autocompletar tareo: {str(e)}'
        }, status=500)


# =============================================================================
# SISTEMA TAREO V2 - Modelo Normalizado (integrado desde views_tareo_v2.py)
# =============================================================================

# =============================================================================
class AsistenciaDiariaForm(ModelForm):
    """
    Formulario individual para cada celda de la matriz de asistencia.
    Incluye validaciones y lógica de negocio específica.
    """
    class Meta:
        model = AsistenciaDiaria
        fields = [_emp_field_name(), 'fecha', 'estado', 'observaciones']
        widgets = {
            'observaciones': Textarea(attrs={'class': 'form-control', 'rows': 1}),
        }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Deshabilitar campos que no deben editarse directamente (nombre de FK puede ser `empleado` o `trabajador`)
        emp_field = _emp_field_name()
        if emp_field in self.fields:
            try:
                self.fields[emp_field].widget.attrs['readonly'] = True
            except Exception:
                pass
        if 'fecha' in self.fields:
            try:
                self.fields['fecha'].widget.attrs['readonly'] = True
            except Exception:
                pass


# =============================================================================
# VISTA PRINCIPAL: TAREO MENSUAL CON MATRIZ
# =============================================================================
@login_required
def tareo_v2_mensual_view(request):
    """
    Vista principal del Tareo V2 con arquitectura normalizada.
    
    Flujo:
    1. GET: Consulta datos verticales y los transforma a matriz para el template
    2. POST: Recibe formset, valida y actualiza masivamente en BD vertical
    
    Features:
    - Proyección automática mensual
    - Edición masiva con formsets
    - Renderizado eficiente para 70+ empleados
    - Navegación por meses
    """
    user = request.user
    
    # =========================================================================
    # 1. VALIDACIÓN DE PERMISOS
    # =========================================================================
    if not user.can_manage_contract_users():
        messages.error(request, 'No tienes permisos para gestionar el tareo de asistencia')
        return redirect('dashboard')
    
    # =========================================================================
    # 2. DETERMINAR CONTRATO
    # =========================================================================
    if user.has_access_to_all_contracts():
        contrato_id = request.GET.get('contrato')
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO')
        else:
            contrato = Contrato.objects.filter(estado='ACTIVO').first()
        contratos_disponibles = Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato')
    else:
        contrato = user.contrato
        contratos_disponibles = None
    
    if not contrato:
        messages.warning(request, 'No hay contratos activos disponibles')
        return redirect('dashboard')
    
    # =========================================================================
    # 3. CALCULAR RANGO DE FECHAS
    # =========================================================================
    vista = request.GET.get('vista', 'mes')               # 'mes' | 'semana'
    mes_offset = int(request.GET.get('mes_offset', 0))
    semana_offset = int(request.GET.get('semana_offset', 0))

    # Día de cambio de guardia (necesario para ambos modos de vista)
    dia_cambio_guardia = contrato.dia_cambio_guardia if contrato.dia_cambio_guardia is not None else 0
    dia_previo_cambio  = (dia_cambio_guardia - 1) % 7

    meses_es = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    dias_abrev = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    hoy = datetime.now().date()

    if vista == 'semana':
        # Semana de guardia: ventana de 7 días que arranca en el dia_cambio_guardia
        # más próximo anterior (o hoy mismo si coincide) desplazada por semana_offset
        dias_desde_cambio = (hoy.weekday() - dia_cambio_guardia) % 7
        inicio_semana_actual = hoy - timedelta(days=dias_desde_cambio)
        fecha_inicio = inicio_semana_actual + timedelta(weeks=semana_offset)
        fecha_fin    = fecha_inicio + timedelta(days=6)
        nombre_periodo = (
            f"Semana de Guardia: "
            f"{dias_abrev[fecha_inicio.weekday()]} {fecha_inicio.strftime('%d/%m')} "
            f"– {dias_abrev[fecha_fin.weekday()]} {fecha_fin.strftime('%d/%m/%Y')}"
        )
        fecha_base     = fecha_fin
        mes_operativo  = fecha_base.month
        anio_operativo = fecha_base.year
    else:
        # Mes operativo: del 26 del mes anterior al 25 del mes actual
        # Por ejemplo: Enero 2026 operativo = 26/12/2025 al 25/01/2026
        fecha_base = hoy
        if mes_offset != 0:
            for _ in range(abs(mes_offset)):
                if mes_offset > 0:
                    if fecha_base.month == 12:
                        fecha_base = fecha_base.replace(year=fecha_base.year + 1, month=1, day=1)
                    else:
                        fecha_base = fecha_base.replace(month=fecha_base.month + 1, day=1)
                else:
                    if fecha_base.month == 1:
                        fecha_base = fecha_base.replace(year=fecha_base.year - 1, month=12, day=1)
                    else:
                        fecha_base = fecha_base.replace(month=fecha_base.month - 1, day=1)

        mes_anterior   = fecha_base.month - 1 if fecha_base.month > 1 else 12
        anio_anterior  = fecha_base.year      if fecha_base.month > 1 else fecha_base.year - 1
        fecha_inicio   = date(anio_anterior, mes_anterior, 26)
        fecha_fin      = date(fecha_base.year, fecha_base.month, 25)
        nombre_periodo = f"{meses_es[fecha_base.month]} {fecha_base.year}"
        mes_operativo  = fecha_base.month
        anio_operativo = fecha_base.year

    # ―― LÍMITE MÍNIMO: nunca mostrar periodos anteriores al 26 Feb 2026 ―――――――
    FECHA_MIN_OPERATIVA = date(2026, 2, 26)
    # El período operativo mínimo es Marzo 2026 (inicia 26 Feb 2026).
    # Calculamos cuántos meses hacia atrás desde hoy llega hasta ese período.
    _min_op_year  = FECHA_MIN_OPERATIVA.year  if FECHA_MIN_OPERATIVA.month < 12 else FECHA_MIN_OPERATIVA.year + 1
    _min_op_month = FECHA_MIN_OPERATIVA.month + 1 if FECHA_MIN_OPERATIVA.month < 12 else 1
    min_mes_offset = (_min_op_year * 12 + _min_op_month) - (hoy.year * 12 + hoy.month)

    if vista == 'mes' and mes_offset < min_mes_offset:
        # Si alguien fuerza un offset menor (URL manual), redirigir al mínimo
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(
            f"{request.path}?contrato={contrato.id}&vista=mes&mes_offset={min_mes_offset}"
        )

    es_mes_minimo = (vista == 'mes' and mes_offset <= min_mes_offset)

    # =========================================================================
    # 4. GENERAR LISTA DE DÍAS DEL MES
    # =========================================================================

    dias_rango = []
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        nombres_dias = {
            0: 'Lun', 1: 'Mar', 2: 'Mié', 3: 'Jue', 
            4: 'Vie', 5: 'Sáb', 6: 'Dom'
        }
        wd = fecha_actual.weekday()
        dias_rango.append({
            'fecha': fecha_actual,
            'dia': fecha_actual.day,
            'nombre_dia': nombres_dias[wd],
            'es_cambio_guardia': wd == dia_cambio_guardia,   # rojo  → día cambio de guardia
            'es_previo_cambio':  wd == dia_previo_cambio,    # amarillo → día previo al cambio
        })
        fecha_actual += timedelta(days=1)
    
    # =========================================================================
    # 5. PROCESAMIENTO POST (GUARDAR CAMBIOS)
    # =========================================================================
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Parsear datos del formulario
                asistencias_data = []
                
                for key, value in request.POST.items():
                    if key.startswith('estado_'):
                        # Formato: estado_trabajadorID_YYYY-MM-DD
                        parts = key.split('_')
                        if len(parts) == 4:
                            _, trabajador_id, anio_str, mes_dia_str = parts
                            fecha_str = f"{anio_str}-{mes_dia_str}"
                            
                            try:
                                fecha_asistencia = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                                observaciones_key = f"observaciones_{trabajador_id}_{anio_str}_{mes_dia_str}"
                                observaciones = request.POST.get(observaciones_key, '')
                                maquina_raw = request.POST.get(f"maquina_{trabajador_id}_{anio_str}_{mes_dia_str}", '')
                                guardia_raw = request.POST.get(f"guardia_{trabajador_id}_{anio_str}_{mes_dia_str}", '')

                                asistencias_data.append({
                                    'trabajador_id': int(trabajador_id),
                                    'fecha': fecha_asistencia,
                                    'estado': value,
                                    'observaciones': observaciones,
                                    'maquina_id': int(maquina_raw) if maquina_raw else None,
                                    'guardia_snapshot': guardia_raw or None,
                                })
                            except ValueError:
                                continue
                
                # Actualización masiva usando el servicio
                resultado = TareoService.actualizar_masivo_desde_formset(
                    asistencias_data,
                    user
                )
                
                # Mensajes de resultado
                if resultado['errores']:
                    messages.warning(
                        request,
                        f"Guardado parcialmente. Errores: {len(resultado['errores'])}"
                    )
                else:
                    messages.success(
                        request,
                        f"Tareo guardado exitosamente. "
                        f"Actualizados: {resultado['actualizados']}, "
                        f"Creados: {resultado['creados']}"
                    )
                
                # Redirigir para evitar reenvío de formulario
                return redirect(f"{request.path}?contrato={contrato.id}&mes_offset={mes_offset}&vista={vista}&semana_offset={semana_offset}")
                
        except Exception as e:
            messages.error(request, f"Error al guardar: {str(e)}")
    
    # =========================================================================
    # 6. PRE-CARGAR DATOS DE V1 → V2 (GET AUTOMÁTICO)
    # =========================================================================
    # Importar registros de AsistenciaTrabajador (V1) que aún no existen en
    # AsistenciaDiaria (V2). Solo crea/actualiza proyecciones; nunca toca
    # correcciones manuales (es_proyeccion=False).
    try:
        resultado_v1 = TareoService.importar_desde_v1(
            contrato=contrato,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            usuario=request.user,
        )
        if resultado_v1['importados'] > 0 or resultado_v1['actualizados'] > 0:
            logger.info(
                f"Pre-carga V1→V2: contrato={contrato.id} "
                f"creados={resultado_v1['importados']} actualizados={resultado_v1['actualizados']}"
            )
    except Exception as e:
        logger.warning(f"Error en pre-carga V1→V2: {str(e)}")

    # =========================================================================
    # 7. OBTENER DATOS PARA VISUALIZACIÓN (GET)
    # =========================================================================
    # Usar el servicio para obtener matriz pivoteada
    matriz_tareo = TareoService.obtener_matriz_tareo(contrato, fecha_inicio, fecha_fin)
    
    # Obtener máquinas activas del contrato para los selects
    maquinas_disponibles = Maquina.objects.filter(
        contrato=contrato,
        estado='OPERATIVO'  # Estado correcto según modelo Maquina
    ).order_by('nombre')
    
    # =========================================================================
    # 8. CONTEXTO PARA EL TEMPLATE
    # =========================================================================
    context = {
        'contrato': contrato,
        'contratos_disponibles': contratos_disponibles,
        'nombre_periodo': nombre_periodo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'dias_rango': dias_rango,
        'matriz_tareo': matriz_tareo,
        'mes_offset': mes_offset,
        'vista': vista,
        'semana_offset': semana_offset,
        'estados_choices': AsistenciaDiaria.ESTADO_CHOICES,
        'maquinas_disponibles': maquinas_disponibles,
        'mes_operativo': mes_operativo,
        'anio_operativo': anio_operativo,
        'dia_cambio_guardia': dia_cambio_guardia,
        'dia_previo_cambio': dia_previo_cambio,
        'nombre_dia_cambio': ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo'][dia_cambio_guardia],
        'es_mes_minimo': es_mes_minimo,
    }
    
    return render(request, 'drilling/tareo/tareo_v2_mensual.html', context)


# =============================================================================
# API PARA PROYECCIÓN MENSUAL (AJAX)
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_generar_proyeccion(request):
    """
    Endpoint AJAX para generar proyección mensual automática.
    
    POST params:
        - contrato_id: ID del contrato
        - anio: Año de la proyección
        - mes: Mes de la proyección
        - sobrescribir: Boolean, si True elimina proyecciones previas
    
    Returns:
        JSON con estadísticas de la operación
    """
    try:
        # Validar permisos
        if not request.user.can_manage_contract_users():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para esta operación'
            }, status=403)
        
        # Obtener parámetros
        contrato_id = request.POST.get('contrato_id')
        anio = int(request.POST.get('anio'))
        mes = int(request.POST.get('mes'))
        sobrescribir = request.POST.get('sobrescribir', 'false').lower() == 'true'
        
        # Validar contrato
        contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO')
        
        # Ejecutar proyección
        resultado = TareoService.generar_proyeccion_mensual(
            anio=anio,
            mes=mes,
            contrato=contrato,
            sobrescribir=sobrescribir
        )
        
        return JsonResponse({
            'success': True,
            'data': resultado
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_generar_proyeccion_todos(request):
    """
    Endpoint AJAX para generar/actualizar proyección mensual para TODOS los
    contratos activos a la vez.

    Calcula automáticamente el mes operativo actual (26 del mes anterior al
    25 del mes en curso) basándose en la fecha del servidor.  No requiere
    pasar contrato_id — usa contrato=None en TareoService para procesar
    todos los trabajadores activos.

    POST params (todos opcionales):
        - sobrescribir: 'true' | 'false' (default: 'false')
          Si es true, elimina todas las proyecciones existentes del período
          antes de regenerarlas.

    Returns:
        JSON con estadísticas agregadas de la operación.
    """
    try:
        if not request.user.can_manage_contract_users():
            return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

        sobrescribir = request.POST.get('sobrescribir', 'false').lower() == 'true'

        # Determinar mes operativo actual
        hoy = date.today()
        # El mes operativo "X" abarca del 26 del mes X-1 al 25 del mes X.
        # Si hoy está entre el 1 y el 25 → operativo = mes actual.
        # Si hoy está entre el 26 y el 31 → operativo = mes siguiente.
        if hoy.day >= 26:
            mes_op = hoy.month + 1 if hoy.month < 12 else 1
            anio_op = hoy.year if hoy.month < 12 else hoy.year + 1
        else:
            mes_op = hoy.month
            anio_op = hoy.year

        resultado = TareoService.generar_proyeccion_mensual(
            anio=anio_op,
            mes=mes_op,
            contrato=None,        # ← todos los contratos
            sobrescribir=sobrescribir,
        )

        return JsonResponse({'success': True, 'data': resultado})

    except Exception as e:
        logger.error(f"Error en proyección todos los contratos: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# API PARA CORRECCIÓN INDIVIDUAL (AJAX)
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_corregir_asistencia(request):
    """
    Endpoint AJAX para corrección individual de una asistencia.
    
    POST params (JSON):
        - empleado_id: ID del trabajador
        - fecha: Fecha en formato YYYY-MM-DD
        - estado: Nuevo estado
        - observaciones: Observaciones opcionales
    
    Returns:
        JSON con el registro actualizado
    """
    try:
        # Validar permisos
        if not request.user.can_manage_contract_users():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para esta operación'
            }, status=403)
        
        # Parsear datos JSON
        data = json.loads(request.body)
        trabajador_id = data.get('trabajador_id') or data.get('empleado_id')
        fecha_str = data.get('fecha')
        estado = data.get('estado')
        observaciones = data.get('observaciones', '')

        trabajador = get_object_or_404(Trabajador, id=trabajador_id)
        if not request.user.has_access_to_all_contracts():
            if not request.user.contrato_id or trabajador.contrato_id != request.user.contrato_id:
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes acceso al contrato del trabajador'
                }, status=403)
        
        # Validar fecha
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Ejecutar corrección
        asistencia = TareoService.corregir_asistencia(
            trabajador_id=trabajador_id,
            fecha=fecha,
            nuevo_estado=estado,
            usuario=request.user,
            observaciones=observaciones
        )
        
        return JsonResponse({
            'success': True,
            'data': {
                'id': asistencia.id,
                'estado': asistencia.estado,
                'estado_display': asistencia.get_estado_display(),
                'es_proyeccion': (getattr(asistencia, 'es_proyeccion', None) if hasattr(asistencia, 'es_proyeccion') else (getattr(asistencia, 'tipo', None) == 'PROY'))
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# API PARA GUARDAR ASISTENCIA POR DÍA (AJAX)
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_guardar_dia_tareo(request):
    """
    Endpoint AJAX para guardar la asistencia completa de un día específico.
    
    POST params (JSON):
        - fecha: Fecha en formato YYYY-MM-DD
        - contrato_id: ID del contrato
        - asistencias: Lista de objetos con:
            - empleado_id: ID del trabajador
            - estado: Estado de asistencia
            - maquina_id: ID de máquina asignada (opcional)
            - observaciones: Observaciones opcionales
    
    Returns:
        JSON con estadísticas de la operación
    """
    try:
        # Validar permisos
        if not request.user.can_manage_contract_users():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para esta operación'
            }, status=403)
        
        # Parsear datos JSON
        data = json.loads(request.body)
        fecha_str = data.get('fecha')
        contrato_id = data.get('contrato_id')
        asistencias_data = data.get('asistencias', [])
        
        # Validar fecha
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Validar contrato
        contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO')
        if not request.user.has_access_to_all_contracts():
            if not request.user.contrato_id or contrato.id != request.user.contrato_id:
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes acceso a este contrato'
                }, status=403)
        
        stats = {
            'actualizados': 0,
            'creados': 0,
            'errores': []
        }
        
        # Procesar cada asistencia
        with transaction.atomic():
            for asist_data in asistencias_data:
                try:
                    trabajador_id = asist_data.get('trabajador_id') or asist_data.get('empleado_id')
                    estado = asist_data.get('estado')
                    maquina_id = asist_data.get('maquina_id')
                    guardia_snapshot = asist_data.get('guardia_snapshot') if 'guardia_snapshot' in asist_data else None
                    observaciones = asist_data.get('observaciones', '')
                    
                    if not trabajador_id or not estado:
                        continue
                    
                    # Obtener trabajador
                    trabajador = Trabajador.objects.get(id=trabajador_id, contrato=contrato)
                    
                    # Obtener máquina si se especificó
                    maquina = None
                    if maquina_id and maquina_id != '':
                        try:
                            maquina = Maquina.objects.get(id=maquina_id)
                        except Maquina.DoesNotExist:
                            pass
                    
                    # Actualizar o crear asistencia (compat: empleado vs trabajador, es_proyeccion vs tipo)
                    emp_field = _emp_field_name()
                    lookup = {f"{emp_field}": trabajador, 'fecha': fecha}
                    defaults = {
                        'estado': estado,
                        'maquina_snapshot': maquina,
                        'observaciones': observaciones,
                        'registrado_por': request.user,
                        'guardia_snapshot': (guardia_snapshot if guardia_snapshot is not None else trabajador.guardia_asignada)
                    }
                    if NEW_TAREO:
                        defaults['tipo'] = 'REAL'
                    else:
                        defaults['es_proyeccion'] = False

                    asistencia, created = AsistenciaDiaria.objects.update_or_create(
                        **lookup,
                        defaults=defaults
                    )
                    
                    if created:
                        stats['creados'] += 1
                    else:
                        stats['actualizados'] += 1
                    
                except Exception as e:
                    stats['errores'].append(f"Error procesando trabajador {trabajador_id}: {str(e)}")
                    logger.error(f"Error guardando asistencia: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        logger.error(f"Error en api_guardar_dia_tareo: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# API PARA GUARDAR SELECCIÓN ARBITRARIA (FILA / GRUPO) (AJAX)
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_guardar_seleccion(request):
    """
    Guarda un conjunto arbitrario de registros de asistencia en una sola
    operación bulk. Usado por los botones "Guardar Fila" y "Guardar Grupo".

    POST body (JSON):
        {
            "contrato_id": int,
            "registros": [
                {"trabajador_id": int, "fecha": "YYYY-MM-DD",
                 "estado": str, "maquina_id": int|null,
                 "observaciones": str}
            ]
        }
    """
    try:
        if not request.user.can_manage_contract_users():
            return JsonResponse(
                {'success': False, 'error': 'Sin permisos'}, status=403
            )

        payload   = json.loads(request.body)
        registros = payload.get('registros', [])

        asistencias_data = []
        for r in registros:
            try:
                asistencias_data.append({
                    'trabajador_id':  int(r['trabajador_id']),
                    'fecha':        datetime.strptime(r['fecha'], '%Y-%m-%d').date(),
                    'estado':       r['estado'],
                    'observaciones': r.get('observaciones', ''),
                    'maquina_id':   r.get('maquina_id') or None,
                    'guardia_snapshot': r.get('guardia_snapshot') if 'guardia_snapshot' in r else None,
                })
            except (KeyError, ValueError):
                continue

        resultado = TareoService.actualizar_masivo_desde_formset(
            asistencias_data, request.user
        )

        return JsonResponse({
            'success': True,
            'data': {
                'actualizados': resultado['actualizados'],
                'creados':      resultado['creados'],
                'errores':      len(resultado['errores']),
            }
        })

    except Exception as e:
        logger.error(f"Error en api_guardar_seleccion: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_cerrar_fecha(request):
    """Marca una fecha como cerrada para un contrato (bloquea escrituras)."""
    try:
        if not request.user.can_manage_contract_users():
            return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

        contrato_id = request.POST.get('contrato_id')
        fecha_s = request.POST.get('fecha')
        motivo = request.POST.get('motivo', '')

        if not contrato_id or not fecha_s:
            return JsonResponse({'success': False, 'error': 'Parámetros incompletos'}, status=400)

        contrato = get_object_or_404(Contrato, id=contrato_id)
        fecha = datetime.strptime(fecha_s, '%Y-%m-%d').date()

        existing = FechaCerrada.objects.filter(contrato=contrato, fecha=fecha).first()
        if existing:
            return JsonResponse({'success': True, 'mensaje': 'Fecha ya cerrada', 'fecha': fecha.isoformat()})

        FechaCerrada.objects.create(contrato=contrato, fecha=fecha, cerrado_por=request.user, motivo=motivo)

        return JsonResponse({'success': True, 'mensaje': 'Fecha marcada como cerrada', 'fecha': fecha.isoformat()})

    except Exception as e:
        logger.error(f"Error cerrando fecha: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_reabrir_fecha(request):
    """Reabre una fecha cerrada (requiere motivo)."""
    try:
        if not request.user.can_manage_contract_users():
            return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

        contrato_id = request.POST.get('contrato_id')
        fecha_s = request.POST.get('fecha')
        motivo = request.POST.get('motivo', '')

        if not contrato_id or not fecha_s or not motivo or len(motivo.strip()) < 5:
            return JsonResponse({'success': False, 'error': 'Parámetros incompletos o motivo insuficiente'}, status=400)

        contrato = get_object_or_404(Contrato, id=contrato_id)
        fecha = datetime.strptime(fecha_s, '%Y-%m-%d').date()

        fc = FechaCerrada.objects.filter(contrato=contrato, fecha=fecha).first()
        if not fc:
            return JsonResponse({'success': False, 'error': 'Fecha no estaba cerrada'}, status=404)

        fc.delete()
        logger.info(f"Fecha reabierta: {contrato} {fecha} por {request.user} motivo: {motivo}")

        return JsonResponse({'success': True, 'mensaje': 'Fecha reabierta', 'fecha': fecha.isoformat()})

    except Exception as e:
        logger.error(f"Error reabriendo fecha: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# API PARA ACTUALIZAR GUARDIA DE UN TRABAJADOR
# =============================================================================
@login_required
@require_http_methods(["POST"])
def api_actualizar_guardia(request):
    """
    Actualiza la guardia asignada (A/B/C) de un trabajador.

    POST body (JSON):
        { "trabajador_id": int, "guardia": "A"|"B"|"C"|"" }
    """
    try:
        if not request.user.can_manage_contract_users():
            return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

        payload      = json.loads(request.body)
        trabajador_id = int(payload['trabajador_id'])
        guardia       = payload.get('guardia', '').strip().upper()

        if guardia and guardia not in ('A', 'B', 'C'):
            return JsonResponse({'success': False, 'error': 'Guardia inválida'}, status=400)

        trabajador = get_object_or_404(Trabajador, id=trabajador_id)

        if not request.user.has_contract_permission(trabajador.contrato):
            return JsonResponse({'success': False, 'error': 'Sin permiso sobre ese contrato'}, status=403)

        trabajador.guardia_asignada = guardia or None
        trabajador.save(update_fields=['guardia_asignada'])

        return JsonResponse({'success': True, 'guardia': trabajador.guardia_asignada or ''})

    except (KeyError, ValueError):
        return JsonResponse({'success': False, 'error': 'Datos inválidos'}, status=400)
    except Exception as e:
        logger.error(f"Error en api_actualizar_guardia: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# API PARA OBTENER MÁQUINAS DEL CONTRATO
# =============================================================================
@login_required
@require_http_methods(["GET"])
def api_obtener_maquinas(request):
    """
    Endpoint AJAX para obtener las máquinas activas de un contrato.
    
    GET params:
        - contrato_id: ID del contrato (opcional)
    
    Returns:
        JSON con lista de máquinas
    """
    try:
        contrato_id = request.GET.get('contrato_id')
        
        # Filtrar máquinas
        maquinas_query = Maquina.objects.filter(estado='ACTIVO')
        
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id)
            maquinas_query = maquinas_query.filter(contrato=contrato)
        
        # Serializar máquinas
        maquinas_data = [
            {
                'id': maq.id,
                'nombre': maq.nombre,
                'codigo': maq.codigo,
                'tipo': maq.tipo_maquina
            }
            for maq in maquinas_query.order_by('nombre')
        ]
        
        return JsonResponse({
            'success': True,
            'data': maquinas_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# =============================================================================
# VISTA DE ESTADÍSTICAS (DASHBOARD)
# =============================================================================
@login_required
def tareo_v2_estadisticas(request):
    """
    Vista de estadísticas y resumen del tareo mensual.
    
    Muestra:
    - Total de días trabajados vs proyectados
    - Ausentismos por tipo
    - Trabajadores por guardia
    - Gráficos de tendencia
    """
    user = request.user
    
    # Determinar contrato
    if user.has_access_to_all_contracts():
        contrato_id = request.GET.get('contrato')
        contrato = get_object_or_404(Contrato, id=contrato_id, estado='ACTIVO') if contrato_id else None
        contratos_disponibles = Contrato.objects.filter(estado='ACTIVO')
    else:
        contrato = user.contrato
        contratos_disponibles = None
    
    if not contrato:
        messages.warning(request, 'Seleccione un contrato')
        return redirect('dashboard')
    
    # Calcular estadísticas del período operativo actual (26 del mes anterior → 25 del mes en curso)
    hoy = date.today()
    _mes_ant_est  = hoy.month - 1 if hoy.month > 1 else 12
    _anio_ant_est = hoy.year      if hoy.month > 1 else hoy.year - 1
    primer_dia_operativo = date(_anio_ant_est, _mes_ant_est, 26)
    ultimo_dia_operativo = date(hoy.year, hoy.month, 25)

    # Query de asistencias del período operativo
    emp_field = _emp_field_name()
    asistencias_mes = AsistenciaDiaria.objects.filter(
        **{f"{emp_field}__contrato": contrato, 'fecha__gte': primer_dia_operativo, 'fecha__lte': min(hoy, ultimo_dia_operativo)}
    )

    # Estadísticas básicas
    total_registros = asistencias_mes.count()
    if NEW_TAREO:
        proyecciones = asistencias_mes.filter(tipo='PROY').count()
        correcciones = asistencias_mes.filter(tipo='REAL').count()
    else:
        proyecciones = asistencias_mes.filter(es_proyeccion=True).count()
        correcciones = asistencias_mes.filter(es_proyeccion=False).count()

    # Por estado
    stats_por_estado = {}
    for estado_code, estado_label in getattr(AsistenciaDiaria, 'ESTADO_CHOICES', []):
        count = asistencias_mes.filter(estado=estado_code).count()
        if count > 0:
            stats_por_estado[estado_label] = count
    
    context = {
        'contrato': contrato,
        'contratos_disponibles': contratos_disponibles,
        'total_registros': total_registros,
        'proyecciones': proyecciones,
        'correcciones': correcciones,
        'stats_por_estado': stats_por_estado,
        'mes_actual': primer_dia_operativo,
    }
    
    return render(request, 'drilling/tareo/tareo_v2_estadisticas.html', context)


# =============================================================================
# VISTAS DE CIERRE MENSUAL Y AUDITORÍA
# =============================================================================

@login_required
def tareo_cierre_mensual(request):
    """
    Vista para revisar y cerrar contablemente el mes.
    Muestra resumen completo antes del cierre.
    """
    # Determinar contrato
    if request.user.is_staff or request.user.is_superuser:
        contratos = Contrato.objects.all()
        contrato_id = request.GET.get('contrato')
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id)
        else:
            contrato = contratos.first()
    else:
        contrato = request.user.contrato
        contratos = Contrato.objects.filter(id=contrato.id)
    
    # Mes y año
    mes = int(request.GET.get('mes', date.today().month))
    anio = int(request.GET.get('anio', date.today().year))
    
    # Obtener resumen del mes
    resumen = CierreMensualService.obtener_resumen_mes(contrato, anio, mes)
    
    # Obtener estado del cierre
    try:
        cierre = CierreMensualTareo.objects.get(contrato=contrato, anio=anio, mes=mes)
    except CierreMensualTareo.DoesNotExist:
        cierre = None
    
    context = {
        'contrato': contrato,
        'contratos': contratos,
        'mes': mes,
        'anio': anio,
        'resumen': resumen,
        'cierre': cierre,
        'meses': range(1, 13),
        'anios': range(date.today().year - 1, date.today().year + 2),
    }
    
    return render(request, 'drilling/tareo/cierre_mensual.html', context)


@login_required
@require_http_methods(["POST"])
def api_cerrar_mes(request):
    """
    API para cerrar contablemente un mes.
    """
    from .utils.tareo_service import CierreMensualService
    
    try:
        contrato_id = request.POST.get('contrato_id')
        mes = int(request.POST.get('mes'))
        anio = int(request.POST.get('anio'))
        observaciones = request.POST.get('observaciones', '')
        
        contrato = get_object_or_404(Contrato, id=contrato_id)
        
        # Verificar permisos
        if not (request.user.is_staff or request.user.role in ['manager', 'admin']):
            return JsonResponse({'error': 'Permisos insuficientes'}, status=403)
        
        # Cerrar mes
        resultado = CierreMensualService.cerrar_mes(
            contrato=contrato,
            anio=anio,
            mes=mes,
            usuario=request.user,
            observaciones=observaciones
        )
        
        if resultado['success']:
            return JsonResponse({
                'success': True,
                'mensaje': resultado['mensaje'],
                'cierre': {
                    'id': resultado['cierre'].id,
                    'estado': resultado['cierre'].estado,
                    'fecha_cierre': resultado['cierre'].fecha_cierre.strftime('%Y-%m-%d %H:%M'),
                    'total_trabajadores': resultado['cierre'].total_trabajadores,
                    'total_dias_trabajo': resultado['cierre'].total_dias_trabajo,
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'error': resultado['error']
            }, status=400)
            
    except Exception as e:
        logger.error(f"Error cerrando mes: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_reabrir_mes(request):
    """
    API para reabrir un mes cerrado (caso excepcional).
    """
    from .utils.tareo_service import CierreMensualService
    
    try:
        contrato_id = request.POST.get('contrato_id')
        mes = int(request.POST.get('mes'))
        anio = int(request.POST.get('anio'))
        motivo = request.POST.get('motivo', '')
        
        contrato = get_object_or_404(Contrato, id=contrato_id)
        
        # Solo admin puede reabrir
        if not request.user.is_staff:
            return JsonResponse({'error': 'Solo administradores pueden reabrir meses'}, status=403)
        
        resultado = CierreMensualService.reabrir_mes(
            contrato=contrato,
            anio=anio,
            mes=mes,
            usuario=request.user,
            motivo=motivo
        )
        
        if resultado['success']:
            return JsonResponse({
                'success': True,
                'mensaje': resultado['mensaje']
            })
        else:
            return JsonResponse({
                'success': False,
                'error': resultado['error']
            }, status=400)
            
    except Exception as e:
        logger.error(f"Error reabriendo mes: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_importar_desde_v1(request):
    """
    API para importar registros de AsistenciaTrabajador (V1) a AsistenciaDiaria (V2).
    Respeta correcciones manuales (es_proyeccion=False).
    Sólo actualiza/crea proyecciones.
    """
    from calendar import monthrange
    from datetime import date
    from .utils.tareo_service import TareoService

    try:
        contrato_id = request.POST.get('contrato_id')
        anio = int(request.POST.get('anio'))
        mes = int(request.POST.get('mes'))

        contrato = get_object_or_404(Contrato, id=contrato_id)

        # El periodo operativo es del 26 del mes anterior al 25 del mes en curso
        if mes > 1:
            mes_ant = mes - 1
            anio_ant = anio
        else:
            mes_ant = 12
            anio_ant = anio - 1

        fecha_inicio = date(anio_ant, mes_ant, 26)
        fecha_fin = date(anio, mes, 25)

        resultado = TareoService.importar_desde_v1(
            contrato=contrato,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            usuario=request.user,
        )

        logger.info(
            f"Importación V1→V2: contrato={contrato_id} periodo={fecha_inicio}/{fecha_fin} "
            f"creados={resultado['importados']} actualizados={resultado['actualizados']} "
            f"omitidos_manual={resultado['omitidos_manual']} sin_mapeo={resultado['sin_mapeo']}"
        )

        return JsonResponse({'success': True, 'data': resultado})

    except Exception as e:
        logger.error(f"Error importando desde V1: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def tareo_historial_trabajador(request, trabajador_id):
    """
    Vista de historial completo de cambios de un trabajador.
    Útil para auditorías y resolución de disputas.
    """
    from .utils.tareo_service import AuditoriaAsistenciaService
    from .models import HistorialCambioAsistencia
    
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    
    # Verificar permisos
    if not (request.user.is_staff or request.user.contrato == trabajador.contrato):
        messages.error(request, 'No tiene permisos para ver este historial')
        return redirect('dashboard')
    
    # Filtros – por defecto usa el período operativo actual (26 del mes anterior → 25 del mes en curso)
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str    = request.GET.get('fecha_fin')

    _hoy = date.today()
    _mes_ant  = _hoy.month - 1 if _hoy.month > 1 else 12
    _anio_ant = _hoy.year      if _hoy.month > 1 else _hoy.year - 1

    if fecha_inicio_str:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
    else:
        fecha_inicio = date(_anio_ant, _mes_ant, 26)    # inicio operativo

    if fecha_fin_str:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    else:
        fecha_fin = date(_hoy.year, _hoy.month, 25)     # fin operativo

    # Obtener historial
    historial = AuditoriaAsistenciaService.obtener_historial_trabajador(
        trabajador=trabajador,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(historial, 50)  # 50 cambios por página
    page_number = request.GET.get('page')
    historial_paginado = paginator.get_page(page_number)
    
    context = {
        'trabajador': trabajador,
        'historial': historial_paginado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'total_cambios': historial.count(),
    }
    
    return render(request, 'drilling/tareo/historial_trabajador.html', context)


@login_required
def tareo_reporte_nomina(request):
    """
    Vista para generar reporte de nómina del mes cerrado.
    Solo muestra meses cerrados para garantizar integridad.
    """
    from .models import CierreMensualTareo, AsistenciaDiaria
    from calendar import monthrange
    
    # Determinar contrato
    if request.user.is_staff or request.user.is_superuser:
        contratos = Contrato.objects.all()
        contrato_id = request.GET.get('contrato')
        if contrato_id:
            contrato = get_object_or_404(Contrato, id=contrato_id)
        else:
            contrato = contratos.first()
    else:
        contrato = request.user.contrato
        contratos = Contrato.objects.filter(id=contrato.id)
    
    # Obtener meses cerrados
    cierres_disponibles = CierreMensualTareo.objects.filter(
        contrato=contrato,
        estado='CERRADO'
    ).order_by('-anio', '-mes')
    
    # Seleccionar cierre
    cierre_id = request.GET.get('cierre_id')
    if cierre_id:
        cierre = get_object_or_404(CierreMensualTareo, id=cierre_id)
    elif cierres_disponibles.exists():
        cierre = cierres_disponibles.first()
    else:
        cierre = None
    
    # Generar reporte si hay cierre seleccionado
    reporte_trabajadores = []
    if cierre:
        # Usar fechas del mes operativo
        primer_dia = cierre.get_fecha_inicio_periodo()
        ultimo_dia = cierre.get_fecha_fin_periodo()
        
        trabajadores = Trabajador.objects.filter(
            contrato=contrato,
            estado='ACTIVO'
        ).order_by('apepat', 'apemat', 'nombres')
        
        for trabajador in trabajadores:
            emp_field = _emp_field_name()
            filtros = {f"{emp_field}": trabajador, 'fecha__gte': primer_dia, 'fecha__lte': ultimo_dia}
            asistencias = AsistenciaDiaria.objects.filter(**filtros)
            if NEW_TAREO:
                asistencias = asistencias.filter(tipo='REAL')
            else:
                asistencias = asistencias.filter(es_proyeccion=False)
            
            dias_trabajo = asistencias.filter(estado='TRABAJO').count()
            dias_descanso = asistencias.filter(estado='DESCANSO').count()
            faltas = asistencias.filter(estado='FALTA').count()
            vacaciones = asistencias.filter(estado='VACACIONES').count()
            permisos = asistencias.filter(estado='PERMISO').count()
            dm = asistencias.filter(estado='DM').count()
            
            reporte_trabajadores.append({
                'trabajador': trabajador,
                'dias_trabajo': dias_trabajo,
                'dias_descanso': dias_descanso,
                'faltas': faltas,
                'vacaciones': vacaciones,
                'permisos': permisos,
                'descanso_medico': dm,
                'total_dias': asistencias.count(),
            })
    
    context = {
        'contrato': contrato,
        'contratos': contratos,
        'cierres_disponibles': cierres_disponibles,
        'cierre': cierre,
        'reporte_trabajadores': reporte_trabajadores,
    }
    
    return render(request, 'drilling/tareo/reporte_nomina.html', context)


@login_required
def api_exportar_nomina_excel(request, cierre_id):
    """
    Exporta el reporte de nómina a Excel.
    """
    from .models import CierreMensualTareo, AsistenciaDiaria
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from calendar import monthrange
    import io
    
    cierre = get_object_or_404(CierreMensualTareo, id=cierre_id)
    
    # Verificar permisos
    if not (request.user.is_staff or request.user.contrato == cierre.contrato):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Nómina {cierre.mes}-{cierre.anio}"
    
    # Encabezados
    headers = [
        'DNI', 'Apellidos', 'Nombres', 'Cargo', 'Régimen',
        'Días Trabajo', 'Días Descanso', 'Faltas', 'Vacaciones',
        'Permisos', 'DM', 'Total Días'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0470AC", end_color="0470AC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos - usar mes operativo
    primer_dia = cierre.get_fecha_inicio_periodo()
    ultimo_dia = cierre.get_fecha_fin_periodo()
    
    trabajadores = Trabajador.objects.filter(
        contrato=cierre.contrato,
        estado='ACTIVO'
    ).order_by('apepat', 'apemat', 'nombres')
    
    row_num = 2
    for trabajador in trabajadores:
        emp_field = _emp_field_name()
        filtros = {f"{emp_field}": trabajador, 'fecha__gte': primer_dia, 'fecha__lte': ultimo_dia}
        asistencias = AsistenciaDiaria.objects.filter(**filtros)
        if NEW_TAREO:
            asistencias = asistencias.filter(tipo='REAL')
        else:
            asistencias = asistencias.filter(es_proyeccion=False)
        
        ws.cell(row=row_num, column=1, value=trabajador.dni)
        ws.cell(row=row_num, column=2, value=trabajador.apellidos)
        ws.cell(row=row_num, column=3, value=trabajador.nombres)
        ws.cell(row=row_num, column=4, value=trabajador.cargo or '')
        ws.cell(row=row_num, column=5, value=trabajador.regimen_laboral or '')
        ws.cell(row=row_num, column=6, value=asistencias.filter(estado='TRABAJO').count())
        ws.cell(row=row_num, column=7, value=asistencias.filter(estado='DESCANSO').count())
        ws.cell(row=row_num, column=8, value=asistencias.filter(estado='FALTA').count())
        ws.cell(row=row_num, column=9, value=asistencias.filter(estado='VACACIONES').count())
        ws.cell(row=row_num, column=10, value=asistencias.filter(estado='PERMISO').count())
        ws.cell(row=row_num, column=11, value=asistencias.filter(estado='DM').count())
        ws.cell(row=row_num, column=12, value=asistencias.count())
        
        row_num += 1
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="nomina_{cierre.contrato.nombre_contrato}_{cierre.mes}_{cierre.anio}.xlsx"'
    
    return response
