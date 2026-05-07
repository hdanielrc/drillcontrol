"""
Vistas del módulo de Planilla / Bonos.
"""
import json
from datetime import date
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db import models
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy, reverse

from .models import Contrato, Trabajador
from .models import TurnoAvance, TurnoTrabajador  # metraje acumulado por trabajador
from .models_payroll import (
    TipoBono,
    ConceptoBono,
    ConfiguracionBonoContrato,
    ConceptoBonoContrato,
    PeriodoBono,
    BonoTrabajador,
    BonoTrabajadorDetalle,
    CriterioBono,
    CalificacionCriterio,
    EstructuraSalarial,
    HistorialEstructuraSalarial,
    HistorialBonoTrabajador,
    ESTADOS_DIA_TRABAJADO,
    AsignacionEstructuraSalarial,
    PresupuestoPlanilla,
)
from .forms_payroll import (
    TipoBonoForm,
    ConceptoBonoFormSet,
    CriterioBonoFormSet,
    ConfiguracionBonoContratoForm,
    ConceptoBonoContratoFormSet,
    EscalaBonoContratoFormSet,
    AbrirPeriodoForm,
    PuntajeDetalleForm,
    EstructuraSalarialForm,
    AsignacionEstructuraSalarialForm,
)
from .utils.payroll_engine import (
    abrir_periodo,
    calcular_periodo,
    aprobar_periodo,
    cerrar_periodo,
    resumen_periodo,
    contar_dias_trabajados,
    contar_dias_trabajados_v2,
    calcular_dias_base_regimen,
    filtrar_trabajadores_por_cargo,
    generar_calificaciones_criterios,
    generar_detalles_vacios,
    calcular_bono_multi_concepto,
)
from .models_tareo import TareoPeriod
from .mixins import AdminOrContractFilterMixin, rol_requerido


# ===========================================
# DEMO — template standalone cuadro evaluación
# ===========================================

@login_required
def cuadro_evaluacion_demo(request):
    return render(request, 'drilling/planilla/cuadro_evaluacion_demo.html')


# ===========================================
# HUB PRINCIPAL DE PLANILLA
# ===========================================

@login_required
def planilla_hub(request):
    """Dashboard principal del módulo de planilla — centrado en cuadros."""
    user = request.user
    contrato = user.contrato
    today = date.today()

    # Tipos de bono activos (cada uno es un "cuadro")
    tipos_bono = TipoBono.objects.filter(activo=True).prefetch_related('conceptos').order_by('codigo')

    # Últimos períodos
    periodos_qs = PeriodoBono.objects.all().order_by('-anio', '-mes')
    if not user.has_access_to_all_contracts() and contrato:
        periodos_qs = periodos_qs.filter(contrato=contrato)
    periodos = periodos_qs[:10]

    # Contratos disponibles
    if user.has_access_to_all_contracts():
        contratos = Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato')
    elif contrato:
        contratos = Contrato.objects.filter(pk=contrato.pk)
    else:
        contratos = Contrato.objects.none()

    context = {
        'tipos_bono': tipos_bono,
        'periodos': periodos,
        'contratos': contratos,
        'anio_actual': today.year,
        'mes_actual': today.month,
    }
    return render(request, 'drilling/planilla/hub.html', context)


# ===========================================
# TIPOS DE BONO — CRUD
# ===========================================

@login_required
def tipo_bono_list(request):
    tipos = TipoBono.objects.all().prefetch_related('conceptos').order_by('codigo')
    return render(request, 'drilling/planilla/tipo_bono_list.html', {'tipos': tipos})


@login_required
def tipo_bono_create(request):
    if request.method == 'POST':
        form = TipoBonoForm(request.POST)
        formset = ConceptoBonoFormSet(request.POST, prefix='conceptos')
        if form.is_valid() and formset.is_valid():
            tipo = form.save(commit=False)
            tipo.es_sistema = False
            tipo.save()
            formset.instance = tipo
            formset.save()
            messages.success(request, f'Tipo de bono "{tipo.codigo}" creado.')
            return redirect('planilla-tipo-bono-list')
    else:
        form = TipoBonoForm()
        formset = ConceptoBonoFormSet(prefix='conceptos')
    return render(request, 'drilling/planilla/tipo_bono_form.html', {
        'form': form, 'conceptos_formset': formset, 'titulo': 'Nuevo Tipo de Bono'
    })


@login_required
def tipo_bono_update(request, pk):
    tipo = get_object_or_404(TipoBono, pk=pk)
    if request.method == 'POST':
        form = TipoBonoForm(request.POST, instance=tipo)
        formset = ConceptoBonoFormSet(request.POST, instance=tipo, prefix='conceptos')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f'Tipo de bono "{tipo.codigo}" actualizado.')
            return redirect('planilla-tipo-bono-edit', pk=tipo.pk)
    else:
        form = TipoBonoForm(instance=tipo)
        formset = ConceptoBonoFormSet(instance=tipo, prefix='conceptos')
    return render(request, 'drilling/planilla/tipo_bono_form.html', {
        'form': form, 'conceptos_formset': formset, 'tipo': tipo, 'titulo': f'Editar {tipo.codigo}'
    })


@login_required
def criterios_concepto(request, concepto_pk):
    """Gestionar criterios (sub-conceptos / checkboxes) de un concepto."""
    concepto = get_object_or_404(
        ConceptoBono.objects.select_related('tipo_bono'),
        pk=concepto_pk
    )
    if request.method == 'POST':
        formset = CriterioBonoFormSet(request.POST, instance=concepto, prefix='criterios')
        if formset.is_valid():
            formset.save()
            messages.success(request, f'Criterios de "{concepto.nombre}" actualizados.')
            return redirect('planilla-criterios-concepto', concepto_pk=concepto.pk)
    else:
        formset = CriterioBonoFormSet(instance=concepto, prefix='criterios')
    return render(request, 'drilling/planilla/criterios_concepto.html', {
        'concepto': concepto,
        'formset': formset,
    })


# ===========================================
# CONFIGURACIÓN DE BONO POR CONTRATO
# ===========================================

@login_required
def config_bono_list(request):
    configs = ConfiguracionBonoContrato.objects.select_related(
        'contrato', 'tipo_bono'
    ).order_by('contrato__nombre_contrato', 'tipo_bono__codigo')
    if not request.user.has_access_to_all_contracts() and request.user.contrato:
        configs = configs.filter(contrato=request.user.contrato)
    return render(request, 'drilling/planilla/config_bono_list.html', {'configuraciones': configs})


@login_required
def config_bono_create(request):
    if request.method == 'POST':
        form = ConfiguracionBonoContratoForm(request.POST, user=request.user)
        if form.is_valid():
            config = form.save()
            messages.success(request, f'Configuración creada: {config}')
            return redirect('planilla-config-bono-edit', pk=config.pk)
    else:
        form = ConfiguracionBonoContratoForm(user=request.user)
    return render(request, 'drilling/planilla/config_bono_form.html', {
        'form': form, 'titulo': 'Nueva Configuración de Bono'
    })


@login_required
def config_bono_edit(request, pk):
    config = get_object_or_404(ConfiguracionBonoContrato, pk=pk)
    es_multi = config.tipo_bono.tipo_calculo == 'MULTI_CONCEPTO'
    es_escalon = config.tipo_bono.tipo_calculo == 'ESCALONADO'

    if request.method == 'POST':
        form = ConfiguracionBonoContratoForm(request.POST, instance=config, user=request.user)
        conceptos_formset = ConceptoBonoContratoFormSet(
            request.POST, instance=config, prefix='conceptos'
        ) if es_multi else None
        escalas_formset = EscalaBonoContratoFormSet(
            request.POST, instance=config, prefix='escalas'
        ) if es_escalon else None

        all_valid = form.is_valid()
        if conceptos_formset:
            all_valid = all_valid and conceptos_formset.is_valid()
        if escalas_formset:
            all_valid = all_valid and escalas_formset.is_valid()

        if all_valid:
            form.save()
            if conceptos_formset:
                conceptos_formset.save()
            if escalas_formset:
                escalas_formset.save()
            messages.success(request, 'Configuración actualizada.')
            return redirect('planilla-config-bono-list')
    else:
        form = ConfiguracionBonoContratoForm(instance=config, user=request.user)
        conceptos_formset = ConceptoBonoContratoFormSet(
            instance=config, prefix='conceptos'
        ) if es_multi else None
        escalas_formset = EscalaBonoContratoFormSet(
            instance=config, prefix='escalas'
        ) if es_escalon else None

    import json as _json
    return render(request, 'drilling/planilla/config_bono_form.html', {
        'form': form,
        'config': config,
        'conceptos_formset': conceptos_formset,
        'escalas_formset': escalas_formset,
        'es_multi': es_multi,
        'es_escalon': es_escalon,
        'titulo': f'Editar Configuración — {config.tipo_bono.codigo}',
        # JSON pre-serializado para el JS del template
        'cargos_aplicables_json': _json.dumps(config.cargos_aplicables or []),
        'montos_por_cargo_json': _json.dumps(config.montos_por_cargo or {}),
        'tipo_calculo_json': _json.dumps(config.tipo_calculo_por_trabajador or {}),
    })


# ===========================================
# PERÍODOS DE BONOS
# ===========================================

@login_required
def periodo_list(request):
    user = request.user
    periodos = PeriodoBono.objects.select_related('contrato', 'calculado_por', 'aprobado_por')
    if not user.has_access_to_all_contracts() and user.contrato:
        periodos = periodos.filter(contrato=user.contrato)
    periodos = periodos.order_by('-anio', '-mes')

    form = AbrirPeriodoForm(initial={'anio': date.today().year, 'mes': date.today().month})

    return render(request, 'drilling/planilla/periodo_list.html', {
        'periodos': periodos,
        'form_abrir': form,
    })


@login_required
@rol_requerido('GERENCIA', 'CONTROL_PROYECTOS', 'ADMINISTRADOR', 'RESIDENTE')
def periodo_abrir(request):
    if request.method != 'POST':
        return redirect('planilla-periodo-list')

    form = AbrirPeriodoForm(request.POST)
    if not form.is_valid():
        messages.error(request, 'Datos del período inválidos.')
        return redirect('planilla-periodo-list')

    user = request.user
    contrato = user.contrato
    if not contrato:
        messages.error(request, 'No tienes contrato asignado.')
        return redirect('planilla-periodo-list')

    anio = int(form.cleaned_data['anio'])
    mes = int(form.cleaned_data['mes'])

    try:
        periodo, registros, eliminados = abrir_periodo(contrato, anio, mes, user)
        msg = f'Período {mes:02d}/{anio} abierto. {registros} registros de bonos creados.'
        if eliminados:
            msg += f' {eliminados} registros fuera de configuración eliminados.'
        messages.success(request, msg)
    except ValueError as e:
        messages.warning(request, str(e))

    return redirect('planilla-periodo-list')


@login_required
def periodo_detalle(request, pk):
    periodo = get_object_or_404(PeriodoBono, pk=pk)

    # Resumen
    resumen = resumen_periodo(periodo)

    # Cargar configuraciones de bono vigentes para filtrar por cargo
    configs_by_tipo = {
        c.tipo_bono_id: c
        for c in ConfiguracionBonoContrato.objects.filter(contrato=periodo.contrato, activo=True)
    }

    # Bonos agrupados por tipo, filtrados por cargos_aplicables de la config
    bonos_por_tipo = {}
    bonos_qs = BonoTrabajador.objects.filter(
        periodo=periodo
    ).select_related('trabajador', 'tipo_bono').order_by('tipo_bono__codigo', 'trabajador__apepat')

    for bono in bonos_qs:
        # Excluir trabajadores cuyo cargo no está en los cargos configurados
        config = configs_by_tipo.get(bono.tipo_bono_id)
        if config and config.cargos_aplicables:
            worker_cargo = (bono.trabajador.cargo or '').strip()
            worker_cargo_hc = (bono.trabajador.cargo_headcount or '').strip()
            if not any(
                c == worker_cargo or c == worker_cargo_hc
                for c in config.cargos_aplicables
            ):
                continue

        codigo = bono.tipo_bono.codigo
        if codigo not in bonos_por_tipo:
            bonos_por_tipo[codigo] = {
                'tipo_bono': bono.tipo_bono,
                'bonos': [],
            }
        bonos_por_tipo[codigo]['bonos'].append(bono)

    for codigo, grupo in bonos_por_tipo.items():
        if grupo['tipo_bono'].tipo_calculo != 'MULTI_CONCEPTO':
            grupo['progreso'] = None
            continue
        bono_pks = [b.pk for b in grupo['bonos']]
        evaluated_set = set(
            CalificacionCriterio.objects.filter(
                bono_trabajador_id__in=bono_pks, cumple=False
            ).values_list('bono_trabajador_id', flat=True)
        ) | set(
            BonoTrabajadorDetalle.objects.filter(
                bono_id__in=bono_pks, puntaje__gt=0
            ).values_list('bono_id', flat=True)
        )
        grupo['progreso'] = {'evaluados': len(evaluated_set), 'total': len(bono_pks)}

    context = {
        'periodo': periodo,
        'resumen': resumen,
        'bonos_por_tipo': bonos_por_tipo,
    }
    return render(request, 'drilling/planilla/periodo_detalle.html', context)


@login_required
@rol_requerido('GERENCIA', 'CONTROL_PROYECTOS', 'ADMINISTRADOR', 'RESIDENTE')
def periodo_calcular(request, pk):
    periodo = get_object_or_404(PeriodoBono, pk=pk)
    try:
        calcular_periodo(periodo, request.user)
        messages.success(request, f'Período {periodo.mes:02d}/{periodo.anio} calculado exitosamente.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('planilla-periodo-detalle', pk=pk)


@login_required
@rol_requerido('GERENCIA', 'CONTROL_PROYECTOS', 'ADMINISTRADOR')
def periodo_aprobar(request, pk):
    periodo = get_object_or_404(PeriodoBono, pk=pk)
    try:
        aprobar_periodo(periodo, request.user)
        messages.success(request, f'Período {periodo.mes:02d}/{periodo.anio} aprobado.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('planilla-periodo-detalle', pk=pk)


# ===========================================
# EVALUACIÓN DE PUNTAJES (MULTI-CONCEPTO)
# ===========================================

@login_required
def evaluar_bono(request, bono_pk):
    """Vista para ingresar puntajes de un bono multi-concepto."""
    bono = get_object_or_404(
        BonoTrabajador.objects.select_related('trabajador', 'tipo_bono', 'periodo'),
        pk=bono_pk
    )

    if bono.periodo.estado == 'CERRADO':
        messages.error(request, 'El período está cerrado.')
        return redirect('planilla-periodo-detalle', pk=bono.periodo.pk)

    detalles = bono.detalles.select_related('concepto').order_by('concepto__orden')

    if request.method == 'POST':
        all_valid = True
        for detalle in detalles:
            field_name = f'puntaje_{detalle.pk}'
            try:
                puntaje = Decimal(request.POST.get(field_name, '0'))
                if 0 <= puntaje <= 100:
                    detalle.puntaje = puntaje
                    detalle.save(update_fields=['puntaje'])
                else:
                    all_valid = False
            except (ValueError, TypeError):
                all_valid = False

        if all_valid:
            messages.success(request, f'Puntajes guardados para {bono.trabajador.get_full_name}.')
        else:
            messages.warning(request, 'Algunos puntajes no son válidos (deben ser 0-100).')
        return redirect('planilla-evaluar-bono', bono_pk=bono_pk)

    return render(request, 'drilling/planilla/evaluar_bono.html', {
        'bono': bono,
        'detalles': detalles,
    })


@login_required
def evaluar_bono_masivo(request, periodo_pk, tipo_bono_pk):
    """DEPRECATED — redirige a cuadro_evaluacion con contexto del período."""
    periodo = get_object_or_404(PeriodoBono, pk=periodo_pk)
    return redirect(
        f"{reverse('planilla-cuadro', args=[tipo_bono_pk])}"
        f"?anio={periodo.anio}&mes={periodo.mes}&from_periodo={periodo_pk}"
    )


# ===========================================
# API ENDPOINTS
# ===========================================

@login_required
def api_conceptos_tipo_bono(request, tipo_bono_pk):
    """Retorna los conceptos de un tipo de bono (para carga dinámica)."""
    conceptos = ConceptoBono.objects.filter(tipo_bono_id=tipo_bono_pk).order_by('orden')
    data = [{'id': c.pk, 'codigo': c.codigo, 'nombre': c.nombre, 'peso': float(c.peso_default)} for c in conceptos]
    return JsonResponse(data, safe=False)


@login_required
def api_exportar_bonos_excel(request, periodo_pk):
    """Exporta los bonos de un período a Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    periodo = get_object_or_404(PeriodoBono, pk=periodo_pk)
    bonos = BonoTrabajador.objects.filter(
        periodo=periodo
    ).select_related('trabajador', 'tipo_bono').order_by('trabajador__apepat', 'tipo_bono__codigo')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Bonos {periodo.mes:02d}-{periodo.anio}'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0470AC', end_color='0470AC', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers
    headers = ['DNI', 'Nombres', 'Apellido Pat.', 'Cargo', 'Código Bono', 'Nombre Bono',
               'Categoría', 'Días Trab.', 'Días Base', 'Factor', 'Calculado', 'Ajuste', 'Final']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for row, bono in enumerate(bonos, 2):
        data = [
            bono.trabajador.dni,
            bono.trabajador.nombres,
            bono.trabajador.apepat,
            bono.trabajador.cargo,
            bono.tipo_bono.codigo,
            bono.tipo_bono.nombre,
            bono.tipo_bono.categoria,
            bono.dias_trabajados,
            bono.dias_base,
            float(bono.factor_cumplimiento),
            float(bono.monto_calculado),
            float(bono.monto_ajuste),
            float(bono.monto_final),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border

    # Auto-width
    for col in ws.columns:
        max_width = max(len(str(cell.value or '')) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_width, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=bonos_{periodo.contrato.nombre_contrato}_{periodo.mes:02d}_{periodo.anio}.xlsx'
    wb.save(response)
    return response


# ===========================================
# CUADRO DE EVALUACIÓN (formato Excel)
# ===========================================

@login_required
def cuadro_evaluacion(request, tipo_bono_pk):
    """
    Vista principal del cuadro de evaluación estilo Excel.
    Muestra todos los trabajadores agrupados por contrato, con secciones
    y criterios como columnas de checkboxes.
    El % de cada sección se toma de ConceptoGlobalPeriodo cuando el código del
    ConceptoBono coincide con un ConceptoGlobal del sistema; si no existe se usa
    el conteo de criterios manual (checkboxes).
    Query params: ?anio=2026&mes=4
    """
    import calendar
    from collections import OrderedDict
    import unicodedata
    from .models_payroll import ConceptoGlobalPeriodo
    from drilling.utils.payroll_engine import resolver_estructura_salarial as _resolver_est
    from drilling.utils.payroll_engine import resolver_estructuras_por_maquinas as _resolver_ests

    # Prefijos ordenados de mayor a menor longitud para evitar falsos matches.
    # Cualquier ConceptoBono.codigo que empiece con uno de estos prefijos
    # (tras quitar tildes y mayúsculas) se mapea al código oficial de ConceptoGlobal.
    _PREFIJOS_GLOBAL = [
        ('RESULTADO_OPERATIVO', 'RESULTADO_OPERATIVO'),
        ('RESULTADO_OP',        'RESULTADO_OPERATIVO'),
        ('RESULTADO',           'RESULTADO_OPERATIVO'),
        ('PRODUCCION',          'PRODUCCION'),
        ('SEGURIDAD',           'SEGURIDAD'),
        ('VALORIZACION',        'VALORIZACION'),
        ('COSTO_METRO',         'CXM'),
        ('COSTO',               'CXM'),
        ('RROO',                'RESULTADO_OPERATIVO'),
        ('PROD',                'PRODUCCION'),
        ('SEGU',                'SEGURIDAD'),
        ('SEG',                 'SEGURIDAD'),
        ('VALO',                'VALORIZACION'),
        ('VAL',                 'VALORIZACION'),
        ('CXM',                 'CXM'),
        ('RO',                  'RESULTADO_OPERATIVO'),
    ]

    # Multiplicadores de bono para trabajadores de tipo METRAJE.
    # Si el indicador global llega al 100 %, se suma ese porcentaje al bono base por metro.
    _METRAJE_MULT = {
        'SEGURIDAD': Decimal('0.04'),   # SEG al 100 % → +4 %
        'CXM':       Decimal('0.04'),   # CXM al 100 % → +4 %
    }

    def _strip_accents(s):
        """Quita tildes y devuelve en mayúsculas."""
        return ''.join(
            c for c in unicodedata.normalize('NFD', s.upper())
            if unicodedata.category(c) != 'Mn'
        )

    def _codigo_a_global(codigo_seccion):
        """
        Convierte el código de un ConceptoBono al código oficial de ConceptoGlobal.
        Primero busca coincidencia exacta en el dict de prefijos, luego startswith
        (de mayor a menor longitud) para tolerar sufijos extras como '_01', etc.
        """
        norm = _strip_accents(codigo_seccion)
        for prefijo, global_cod in _PREFIJOS_GLOBAL:
            if norm == prefijo or norm.startswith(prefijo):
                return global_cod
        return None

    tipo_bono = get_object_or_404(TipoBono, pk=tipo_bono_pk, activo=True)
    anio = int(request.GET.get('anio') or date.today().year)
    mes = int(request.GET.get('mes') or date.today().month)
    from_periodo_pk = request.GET.get('from_periodo')
    _, ultimo_dia = calendar.monthrange(anio, mes)
    fecha_inicio = date(anio, mes, 1)
    fecha_fin = date(anio, mes, ultimo_dia)

    user = request.user

    # Secciones y criterios del tipo de bono
    secciones = ConceptoBono.objects.filter(
        tipo_bono=tipo_bono
    ).prefetch_related('criterios').order_by('orden')

    # Obtener contratos con configuración activa para este tipo de bono
    configs = ConfiguracionBonoContrato.objects.filter(
        tipo_bono=tipo_bono, activo=True,
    ).select_related('contrato')
    if not user.has_access_to_all_contracts() and user.contrato:
        configs = configs.filter(contrato=user.contrato)

    datos_por_contrato = OrderedDict()

    for config in configs:
        contrato = config.contrato
        # Cargar conceptos globales del período para este contrato
        # Indexados por código normalizado (sin acentos, mayúsculas)
        cgp_qs = ConceptoGlobalPeriodo.objects.filter(
            contrato=contrato, anio=anio, mes=mes
        ).select_related('concepto')
        conceptos_globales_periodo = {
            _strip_accents(cgp.concepto.codigo): cgp
            for cgp in cgp_qs
        }

        # Sincronizar V1 (AsistenciaTrabajador) → V2 (AsistenciaDiaria) para
        # este contrato y período operativo, igual que hace la vista del tareo.
        # Garantiza que contar_dias_trabajados refleje los mismos datos que el
        # Resumen de Planilla.
        try:
            from .utils.tareo_service import TareoService as _TareoService
            _mes_ant = mes - 1 if mes > 1 else 12
            _anio_ant = anio if mes > 1 else anio - 1
            _sync_inicio = date(_anio_ant, _mes_ant, 23)
            _TareoService.importar_desde_v1(
                contrato=contrato,
                fecha_inicio=_sync_inicio,
                fecha_fin=fecha_fin,
                usuario=request.user,
            )
        except Exception:
            pass  # El sync es best-effort; no debe romper la vista

        # Obtener o crear período
        periodo, _ = PeriodoBono.objects.get_or_create(
            contrato=contrato, anio=anio, mes=mes,
            defaults={'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'estado': 'ABIERTO'}
        )

        # Obtener trabajadores del contrato que aplican al cuadro
        trabajadores = Trabajador.objects.filter(
            contrato=contrato, estado='ACTIVO'
        ).order_by('apepat', 'apemat', 'nombres')
        trabajadores = filtrar_trabajadores_por_cargo(trabajadores, tipo_bono, config=config)

        filas_cumplimiento = []
        filas_metraje = []

        # ── Pre-carga para prorrateo de metraje ──────────────────────────────
        from .tareo_compat import AsistenciaDiaria, NEW_TAREO as _NEW_TAREO
        from .utils.periodo_operativo import get_rango_mes_operativo, cantidad_dias_mes_operativo
        from django.db.models import Sum as _SumM, Count as _CountM

        # Todo usa el mismo período operativo 26-25 para consistencia
        _op_inicio, _op_fin = get_rango_mes_operativo(anio, mes)
        _dias_mes_op = cantidad_dias_mes_operativo(anio, mes)

        # metros_acumulados por maquina_id en el período operativo
        _metros_maquina = {}
        _rows = (
            TurnoAvance.objects.filter(
                turno__contrato=contrato,
                turno__fecha__gte=_op_inicio,
                turno__fecha__lte=_op_fin,
                turno__estado__in=['COMPLETADO', 'APROBADO'],
            )
            .values('turno__maquina_id')
            .annotate(total=_SumM('metros_perforados'))
        )
        for _r in _rows:
            if _r['turno__maquina_id']:
                _metros_maquina[_r['turno__maquina_id']] = Decimal(str(_r['total'] or 0))

        # nombre de máquinas para el desglose en template
        from .models import Maquina as _Maquina, ProgramacionMes as _ProgramacionMes
        _maquina_nombres = dict(_Maquina.objects.filter(contrato=contrato).values_list('id', 'nombre'))

        # meta programada por máquina: usa ProgramacionMes (misma fuente que /gerencia/programacion/)
        _meta_maquina = dict(
            _ProgramacionMes.objects.filter(maquina__contrato=contrato, año=anio, mes=mes)
            .values_list('maquina_id', 'meta_metros')
        )

        # días trabajados por trabajador+máquina en el período operativo 26-25
        # misma lógica que la tabla Excel del tareo (views_tareo.py línea 1795-1807):
        # TareoEntry usa tipo='REAL'; AsistenciaDiaria legacy usa es_proyeccion=False
        # resultado: {trabajador_id: [(maquina_id, dias), ...]}
        _dias_trab_maquinas = {}
        _real_filter = {'tipo': 'REAL'} if _NEW_TAREO else {'es_proyeccion': False}
        _ad_rows = (
            AsistenciaDiaria.objects.filter(
                trabajador__contrato=contrato,
                fecha__gte=_op_inicio,
                fecha__lte=_op_fin,
                maquina_snapshot__isnull=False,
                **_real_filter,
            )
            .values('trabajador_id', 'maquina_snapshot_id')
            .annotate(dias=_CountM('id'))
        )
        for _ad in _ad_rows:
            _tid = _ad['trabajador_id']
            _dias_trab_maquinas.setdefault(_tid, []).append(
                (_ad['maquina_snapshot_id'], _ad['dias'])
            )

        for trab in trabajadores:
            cargo_trab = trab.cargo or trab.cargo_headcount or ''
            from decimal import Decimal as _D

            # Mapa maquina_id → EstructuraSalarial para prorrateo por máquina
            _structs_trab = _resolver_ests(trab, config.contrato, cargo_trab, fecha_inicio, fecha_fin)
            _est_por_maquina = {maq_id: est for est, _, maq_id in _structs_trab if maq_id}

            # ── Determinar tipo de cálculo del trabajador ─────────────────────
            _default_calc = getattr(config, 'tipo_calculo_default', 'cumplimiento') or 'cumplimiento'
            tipo_calc = (config.tipo_calculo_por_trabajador or {}).get(trab.dni, _default_calc)

            def _bono_base_para(cfg, cargo, tipo):
                """Retorna el monto base según el tipo de cálculo del trabajador."""
                if tipo == 'metraje':
                    est = _resolver_est(trab, cfg.contrato, cargo, fecha_inicio, fecha_fin)
                    if est and est.bono_por_metraje:
                        # Almacenar solo la TARIFA por metro (no el producto con metraje_base)
                        return est.bono_por_metraje
                # cumplimiento (default): per-cargo o global
                if cfg.montos_por_cargo and cargo in cfg.montos_por_cargo:
                    return _D(str(cfg.montos_por_cargo[cargo]))
                return cfg.monto_base_mensual

            monto_inicial = _bono_base_para(config, cargo_trab, tipo_calc)

            # Obtener o crear BonoTrabajador
            bono_trab, bt_created = BonoTrabajador.objects.get_or_create(
                periodo=periodo, trabajador=trab, tipo_bono=tipo_bono,
                defaults={
                    'configuracion': config,
                    'bono_base': monto_inicial,
                    'dias_trabajados': 0, 'dias_base': 0,
                    'factor_cumplimiento': Decimal('1'),
                    'monto_calculado': Decimal('0'),
                    'monto_ajuste': Decimal('0'),
                    'monto_final': Decimal('0'),
                }
            )

            # Siempre sincronizar bono_base desde la configuración si está vacío o desactualizado
            cargo_trab = trab.cargo or trab.cargo_headcount or ''
            monto_esperado = _bono_base_para(config, cargo_trab, tipo_calc)

            fields_to_update = []
            if not bono_trab.bono_base or bono_trab.bono_base != monto_esperado:
                bono_trab.bono_base = monto_esperado
                fields_to_update.append('bono_base')

            # Recalcular días trabajados y días base
            if tipo_bono.usa_periodo_operativo_tareo:
                # Bonos BA-: días_trabajados = 30 - count(estados NO trabajados en período 26-25).
                # Estados trabajados: TD/TN/DL/DA/MDL. Todo lo demás (V-A, permisos, DM…) resta 1 día.
                # El período puede tener 31 días; el máximo siempre es 30 (fijo).
                _estados_ba = (
                    'TD', 'TRABAJO_DIA', 'TN', 'TRABAJO_NOCHE',
                    'DL', 'DIA_LIBRE', 'DESCANSO', 'DA', 'DIA_APOYO', 'MDL',
                )
                _no_trabajados = AsistenciaDiaria.objects.filter(
                    trabajador=trab,
                    fecha__gte=_op_inicio,
                    fecha__lte=_op_fin,
                    **_real_filter,
                ).exclude(
                    estado__in=_estados_ba,
                ).count()
                dias_trab = max(Decimal('0'), Decimal('30') - Decimal(str(_no_trabajados)))
                dias_base = 30
            else:
                dias_trab = contar_dias_trabajados(trab, fecha_inicio, fecha_fin)
                dias_base = calcular_dias_base_regimen(trab, fecha_inicio, fecha_fin) or 30
            bono_trab.dias_trabajados = dias_trab
            bono_trab.dias_base = dias_base
            fields_to_update += ['dias_trabajados', 'dias_base']

            if fields_to_update:
                bono_trab.save(update_fields=fields_to_update)

            # Generar detalles y calificaciones si faltan (solo cumplimiento)
            if bt_created and tipo_calc != 'metraje':
                generar_detalles_vacios(bono_trab, config)
                generar_calificaciones_criterios(bono_trab)

            # ── TRABAJADORES DE METRAJE o AMBOS: cálculo directo ─────────────
            if tipo_calc in ('metraje', 'ambos'):
                seg_cgp = conceptos_globales_periodo.get('SEGURIDAD')
                seg_puntaje = float(seg_cgp.porcentaje_bono) if seg_cgp else 0.0
                seg_activo = seg_puntaje >= 100

                # CXM no aplica para BA-OPERADORES — solo Seguridad suma bonificación
                mult_seg = Decimal('0.04') if seg_activo else Decimal('0')
                mult_total = Decimal('1') + mult_seg
                mult_total_pct = int(mult_total * 100)  # 100 o 104

                est_m = _resolver_est(trab, config.contrato, cargo_trab, fecha_inicio, fecha_fin)
                metraje_base_val = est_m.metraje_base if est_m and est_m.metraje_base else Decimal('0')

                # Metraje acumulado del trabajador: suma de metros perforados
                # en los turnos donde aparece como trabajador en el período.
                from django.db.models import Sum as _Sum
                _metros = TurnoAvance.objects.filter(
                    turno__contrato=contrato,
                    turno__fecha__gte=fecha_inicio,
                    turno__fecha__lte=fecha_fin,
                    turno__trabajadores_turno__trabajador=trab,
                ).aggregate(total=_Sum('metros_perforados'))['total']
                metraje_acum_val = (
                    Decimal(str(_metros)).quantize(Decimal('0.001'))
                    if _metros else Decimal('0')
                )
                # Para "ambos", bono_base tiene el monto de cumplimiento; leer tarifa desde EstructuraSalarial
                if tipo_calc == 'ambos':
                    bono_por_metro_val = (est_m.bono_por_metraje if est_m and est_m.bono_por_metraje else Decimal('0'))
                else:
                    bono_por_metro_val = bono_trab.bono_base  # tarifa por metro (unit rate)

                # Persiste el valor calculado para referencia histórica
                if bono_trab.metraje_acumulado != metraje_acum_val:
                    bono_trab.metraje_acumulado = metraje_acum_val
                    bono_trab.save(update_fields=['metraje_acumulado'])

                # monto_ajustado referencial (tarifa dominante × mult); se usa como fallback en el loop
                monto_ajustado = (bono_por_metro_val * mult_total).quantize(Decimal('0.001'))
                # bono_calculado se asigna después del loop de prorrateo por máquina

                # Prorrateo por máquina: una entrada por cada máquina donde trabajó
                # Bonos BA-: denominador fijo 30; resto: días reales del período operativo
                _d_op = Decimal('30') if tipo_bono.usa_periodo_operativo_tareo else Decimal(str(_dias_mes_op))
                _dias_op_limite = 30 if tipo_bono.usa_periodo_operativo_tareo else _dias_mes_op
                _maq_entradas = _dias_trab_maquinas.get(trab.pk, [])
                _desglose_maquinas = []
                _base_prorrateo_total = Decimal('0')
                _acum_prorrateo_total = Decimal('0')
                _total_prorrateo_total = Decimal('0')
                _total_dias_en_maq = 0
                _dias_meta_cumplida = 0  # días en máquinas que superaron la meta

                for _maq_id, _dias_raw in _maq_entradas:
                    # BA-: días = total trabajados del tareo (TD/TN/DL/DA/MDL), no días con máquina asignada
                    if tipo_bono.usa_periodo_operativo_tareo:
                        _dias_en_maq = min(int(bono_trab.dias_trabajados), 30)
                    else:
                        _dias_en_maq = min(_dias_raw, _dias_op_limite)
                    _d_en_maq = Decimal(str(_dias_en_maq))
                    _metros_maq = _metros_maquina.get(_maq_id, Decimal('0'))
                    _meta_maq = _meta_maquina.get(_maq_id)
                    # Regla: ≥ 90% de la meta en esa máquina → aporta al Bono Meta
                    _pct_maq = (float(_metros_maq) / float(_meta_maq) * 100) if _meta_maq else 0.0
                    _meta_cumplida = _meta_maq is not None and _pct_maq >= 90
                    # Estructura específica de esta máquina; fallback a est_m (dominante)
                    _est_maq = _est_por_maquina.get(_maq_id) or est_m
                    _metraje_base_maq = (
                        _est_maq.metraje_base
                        if _est_maq and _est_maq.metraje_base
                        else metraje_base_val
                    )
                    _bono_metro_maq = (
                        _est_maq.bono_por_metraje
                        if _est_maq and _est_maq.bono_por_metraje
                        else bono_por_metro_val
                    )
                    _monto_ajustado_maq = (_bono_metro_maq * mult_total).quantize(Decimal('0.001'))
                    _base_p = (
                        (_metraje_base_maq / _d_op * _d_en_maq).quantize(Decimal('0.01'))
                        if _d_op > 0 else Decimal('0')
                    )
                    _acum_p = (
                        (_metros_maq / _d_op * _d_en_maq).quantize(Decimal('0.01'))
                        if _d_op > 0 else Decimal('0')
                    )
                    _total_p = ((_acum_p - _base_p) * _monto_ajustado_maq).quantize(Decimal('0.01'))
                    _desglose_maquinas.append({
                        'maquina_id': _maq_id,
                        'maquina_nombre': _maquina_nombres.get(_maq_id, f'Máq. {_maq_id}'),
                        'dias': _dias_en_maq,
                        'metros_acum': float(_metros_maq),
                        'meta_metros': float(_meta_maq) if _meta_maq is not None else None,
                        'pct_cumplimiento': round(_pct_maq, 1),
                        'meta_cumplida': _meta_cumplida,
                        'bono_por_metro': float(_bono_metro_maq),
                        'metraje_base_maq': float(_metraje_base_maq),
                        'base_prorrateo': float(_base_p),
                        'acum_prorrateo': float(_acum_p),
                        'total': float(_total_p),
                    })
                    _base_prorrateo_total += _base_p
                    _acum_prorrateo_total += _acum_p
                    _total_prorrateo_total += _total_p
                    _total_dias_en_maq += _dias_en_maq
                    if _meta_cumplida:
                        _dias_meta_cumplida += _dias_en_maq

                # BA-: normalizar _total_dias_en_maq al valor real (evita inflación por multi-máquina)
                if tipo_bono.usa_periodo_operativo_tareo and _maq_entradas:
                    _total_dias_en_maq = min(int(bono_trab.dias_trabajados), 30)

                # Administradores/Residentes/Logísticos (BA-SUPERVISIÓN): no aparecen en turnos.
                # BASE PRORRATEADA = (Σ metros_maquinas / N_maquinas) / 30 × dias_trabajados
                # MONTO = base_prorrateada × tarifa_por_metro
                # dias_mes_operativo es siempre 30 fijo; dias_trabajados máx 30 desde TareoV2.
                if not _maq_entradas:
                    _desglose_maquinas = []
                    _dias_fallback = min(bono_trab.dias_trabajados, _dias_op_limite)
                    _d_fb = Decimal(str(_dias_fallback))
                    if _d_op > 0 and _dias_fallback > 0:
                        _est_sup = _resolver_est(trab, config.contrato, cargo_trab, fecha_inicio, fecha_fin)
                        _estructuras_sup = [_est_sup] if _est_sup else []
                        # Determinar conjunto de máquinas según EstructuraSalarial
                        _maq_ids_sup = [e.maquina_id for e in _estructuras_sup if e.maquina_id]
                        if _maq_ids_sup:
                            # Estructuras con máquinas específicas
                            _metros_sup = {mid: _metros_maquina.get(mid, Decimal('0')) for mid in _maq_ids_sup}
                        else:
                            # Sin máquina asignada → todas las máquinas del contrato
                            _metros_sup = dict(_metros_maquina)

                        _n_maq_sup = len(_metros_sup)
                        _suma_metros_sup = sum(_metros_sup.values(), Decimal('0'))
                        # Promedio de metraje acumulado entre las máquinas
                        _metros_prom_sup = (
                            (_suma_metros_sup / Decimal(str(_n_maq_sup))).quantize(Decimal('0.001'))
                            if _n_maq_sup > 0 else Decimal('0')
                        )
                        # Base prorrateada = (metraje_base / 30) × dias_trabajados, máx metraje_base
                        # dias_mes_operativo = 30 fijo; dias_trabajados máx 30 (TareoV2 26-25)
                        _acum_prorrateo_total = min(
                            metraje_base_val,
                            (metraje_base_val / Decimal('30') * _d_fb),
                        ).quantize(Decimal('0.001'))
                        _base_prorrateo_total = _acum_prorrateo_total
                        _total_prorrateo_total = (
                            _acum_prorrateo_total * monto_ajustado
                        ).quantize(Decimal('0.01'))
                        _total_dias_en_maq = _dias_fallback

                # bono_calculado = suma prorateada por máquina (ya incluye tarifas específicas por equipo)
                bono_calculado = _total_prorrateo_total

                # ── Bono Meta para metraje puro: base mensual del cargo × factor ≥90% ──
                # Base: montos_por_cargo[cargo] o monto_base_mensual (misma fuente que cumplimiento).
                # Factor: días en máquinas ≥90% / total días en máquinas.
                # Requiere MetaMaquina configurada; si no hay datos → 0 (sin penalización implícita).
                _monto_meta_base = (
                    Decimal(str(config.montos_por_cargo[cargo_trab]))
                    if (config.montos_por_cargo and cargo_trab in config.montos_por_cargo)
                    else config.monto_base_mensual
                )
                if _total_dias_en_maq > 0 and _dias_meta_cumplida > 0:
                    _bono_meta_calc_m = (
                        _monto_meta_base
                        * Decimal(str(_dias_meta_cumplida))
                        / Decimal(str(_total_dias_en_maq))
                    ).quantize(Decimal('0.01'))
                else:
                    _bono_meta_calc_m = Decimal('0')

                # Hay MetaMaquina configurada para al menos una máquina del desglose?
                _tiene_meta_config = any(
                    m['meta_metros'] is not None for m in _desglose_maquinas
                ) if _desglose_maquinas else False

                # Auto-persistir monto_final solo para metraje puro (para "ambos" lo gestiona el bloque de cumplimiento)
                _total_con_meta = _total_prorrateo_total + _bono_meta_calc_m
                if tipo_calc != 'ambos' and bono_trab.monto_final != _total_con_meta:
                    bono_trab.monto_calculado = _total_con_meta
                    bono_trab.monto_final = (_total_con_meta + bono_trab.monto_ajuste).quantize(Decimal('0.01'))
                    bono_trab.registrar_historial(
                        fuente='CALCULO',
                        usuario=request.user,
                        observacion=f'Cálculo automático metraje: S/{float(bono_calculado):.2f} + Bono Meta: S/{float(_bono_meta_calc_m):.2f}',
                    )
                    bono_trab.save(update_fields=['monto_calculado', 'monto_final'])

                filas_metraje.append({
                    'bono_pk': bono_trab.pk,
                    'trabajador_pk': trab.pk,
                    'nombre_completo': f"{trab.apepat} {trab.apemat} {trab.nombres}".strip(),
                    'cargo': cargo_trab,
                    'dias_trabajados': bono_trab.dias_trabajados,
                    'dias_operativos': bono_trab.dias_base,
                    'bono_por_metraje': float(bono_por_metro_val),
                    'metraje_base': float(metraje_base_val),
                    'seg_puntaje': seg_puntaje,
                    'seg_activo': seg_activo,
                    'mult_total_pct': mult_total_pct,
                    'monto_ajustado': float(monto_ajustado),
                    'bono_calculado': float(bono_calculado),
                    'total': float(_total_prorrateo_total),
                    'dias_mes_operativo': int(_d_op),
                    'dias_en_maquina': _total_dias_en_maq,
                    'dias_meta_cumplida': _dias_meta_cumplida,
                    'metraje_base_prorrateo': float(_base_prorrateo_total),
                    'metraje_acum_prorrateo': float(_acum_prorrateo_total),
                    'desglose_maquinas': _desglose_maquinas,
                    'tipo_calculo': tipo_calc,
                    'bono_meta_total': float(_bono_meta_calc_m),
                    'bono_meta_base': float(_monto_meta_base),
                    'tiene_meta_config': _tiene_meta_config,
                    'gran_total': float(_total_con_meta),
                })
                if tipo_calc != 'ambos':
                    continue  # no procesar secciones para trabajadores de metraje puro

            # ── Factor de máquinas para BONO META (cumplimiento) ─────────────
            # Regla: ≥90% en una máquina → esa máquina aporta al Bono Meta.
            # factor = días en máquinas ≥90% / total días trabajados en máquinas.
            # Si el trabajador no tiene entradas de máquina, factor = 1.0 (sin penalización).
            _maq_entradas_cumpl = _dias_trab_maquinas.get(trab.pk, [])
            _dias_ok_cumpl = 0
            _total_dias_maq_cumpl = 0
            _desglose_factor_cumpl = []
            for _maq_id_c, _dias_raw_c in _maq_entradas_cumpl:
                _dias_c = min(_dias_raw_c, _dias_mes_op)
                _metros_c = _metros_maquina.get(_maq_id_c, Decimal('0'))
                _meta_c = _meta_maquina.get(_maq_id_c)
                _pct_c = (float(_metros_c) / float(_meta_c) * 100) if _meta_c else 0.0
                _aporta_c = _meta_c is not None and _pct_c >= 90
                _total_dias_maq_cumpl += _dias_c
                if _aporta_c:
                    _dias_ok_cumpl += _dias_c
                _desglose_factor_cumpl.append({
                    'maquina_nombre': _maquina_nombres.get(_maq_id_c, f'Máq. {_maq_id_c}'),
                    'dias': _dias_c,
                    'pct_cumplimiento': round(_pct_c, 1),
                    'aporta': _aporta_c,
                })
            if _total_dias_maq_cumpl > 0:
                _factor_maq_cumpl = _dias_ok_cumpl / _total_dias_maq_cumpl
            else:
                _factor_maq_cumpl = 1.0  # sin datos de máquina → sin penalización

            # Construir datos de secciones con criterios (solo cumplimiento)
            secciones_data = []
            total_monto_trab = Decimal('0')
            for seccion in secciones:
                criterios_seccion = seccion.criterios.filter(activo=True).order_by('orden')
                criterios_data = []
                cumplidos = 0
                total_crit = criterios_seccion.count()

                for criterio in criterios_seccion:
                    calif, _ = CalificacionCriterio.objects.get_or_create(
                        bono_trabajador=bono_trab, criterio=criterio,
                        defaults={'cumple': True}
                    )
                    criterios_data.append({
                        'criterio_pk': criterio.pk,
                        'nombre': criterio.nombre,
                        'cumple': calif.cumple,
                    })
                    if calif.cumple:
                        cumplidos += 1

                # Puntaje: según modalidad de la sección
                global_codigo = _codigo_a_global(seccion.codigo)
                cgp = conceptos_globales_periodo.get(global_codigo) if global_codigo else None

                if cgp is not None:
                    # Concepto global del contrato: mayor prioridad en todos los bonos.
                    # Esto hace que PRODUCCIÓN en BA-ADMINISTRACIÓN use el mismo %
                    # que en logísticos y residentes.
                    puntaje = float(cgp.porcentaje_bono)
                    fuente_puntaje = 'global'
                    # Persistir el puntaje global en BonoTrabajadorDetalle para que
                    # calcular_periodo/calcular_bono_multi_concepto usen el valor correcto.
                    _monto_max = (
                        bono_trab.bono_base * seccion.peso_default / Decimal('100')
                    ).quantize(Decimal('0.01'))
                    BonoTrabajadorDetalle.objects.update_or_create(
                        bono=bono_trab,
                        concepto=seccion,
                        defaults={
                            'puntaje': Decimal(str(round(puntaje))),
                            'monto_max_concepto': _monto_max,
                        }
                    )
                elif seccion.tabla_calificacion:
                    # Tabla fija: leer desde BonoTrabajadorDetalle (persistido)
                    detalle_tc, _ = BonoTrabajadorDetalle.objects.get_or_create(
                        bono=bono_trab,
                        concepto=seccion,
                        defaults={
                            'puntaje': Decimal('100'),
                            'monto_max_concepto': Decimal(str(
                                float(bono_trab.bono_base) * float(seccion.peso_default) / 100
                            )),
                            'monto_calculado': Decimal('0'),
                        }
                    )
                    puntaje = int(detalle_tc.puntaje)
                    fuente_puntaje = 'tabla_fija'
                else:
                    # Fallback: conteo manual de criterios (checkboxes)
                    puntaje = round(cumplidos * 100 / total_crit) if total_crit > 0 else 100
                    fuente_puntaje = 'manual'

                peso = float(seccion.peso_default)
                bono_base = float(bono_trab.bono_base)

                # Multiplicador metraje para esta sección (0 si no aplica)
                metraje_mult = _METRAJE_MULT.get(global_codigo, Decimal('0'))

                # Fórmula cumplimiento: bono_base × peso% × puntaje%
                # Para la sección PRODUCCION (Bono Meta): aplicar factor de máquinas ≥90%.
                monto_seccion_base = round(bono_base * (peso / 100) * (puntaje / 100), 2)
                if global_codigo == 'PRODUCCION' and _total_dias_maq_cumpl > 0:
                    monto_seccion = round(monto_seccion_base * _factor_maq_cumpl, 2)
                    _factor_maq_aplicado = _factor_maq_cumpl
                else:
                    monto_seccion = monto_seccion_base
                    _factor_maq_aplicado = None

                total_monto_trab += Decimal(str(monto_seccion))

                secciones_data.append({
                    'concepto_pk': seccion.pk,
                    'nombre': seccion.nombre,
                    'peso': peso,
                    'criterios': criterios_data,
                    'puntaje': puntaje,
                    'fuente_puntaje': fuente_puntaje,
                    'monto': monto_seccion,
                    'monto_bruto': monto_seccion_base,
                    'factor_maq': round(_factor_maq_aplicado * 100, 1) if _factor_maq_aplicado is not None else None,
                    'desglose_factor_maq': _desglose_factor_cumpl if _factor_maq_aplicado is not None else [],
                    'metraje_mult_pct': float(metraje_mult * 100),  # 0, 4 ó 8
                    'tabla_calificacion': seccion.tabla_calificacion,
                })

            # Fórmula cumplimiento: suma de secciones (sin factor días — se
            # aplicará más adelante cuando se integre el cálculo por días).
            total_monto_trab = total_monto_trab.quantize(Decimal('0.01'))

            # Para trabajadores 'ambos', calcular bono_meta prorrateado:
            # Regla: factor = días en máquinas ≥90% / total días trabajados en máquinas.
            if tipo_calc == 'ambos' and filas_metraje:
                _fila_m = filas_metraje[-1]
                _dias_ok = Decimal(str(_fila_m['dias_meta_cumplida']))
                _total_dias_maq_dec = Decimal(str(_fila_m['dias_en_maquina']))
                if _total_dias_maq_dec > 0 and _dias_ok > 0:
                    _bono_meta_calc = (total_monto_trab * _dias_ok / _total_dias_maq_dec).quantize(Decimal('0.01'))
                else:
                    _bono_meta_calc = Decimal('0')
                _fila_m['bono_meta_total'] = float(_bono_meta_calc)
                _fila_m['bono_meta_base'] = float(total_monto_trab)
                _fila_m['gran_total'] = round(_fila_m['total'] + float(_bono_meta_calc), 2)

            # Auto-persistir monto_final en cada carga (no esperar a que el usuario pulse Guardar)
            if bono_trab.monto_final != total_monto_trab:
                bono_trab.monto_calculado = total_monto_trab
                bono_trab.monto_final = (total_monto_trab + bono_trab.monto_ajuste).quantize(Decimal('0.01'))
                bono_trab.save(update_fields=['monto_calculado', 'monto_final'])
                bono_trab.registrar_historial(
                    fuente='CALCULO',
                    usuario=request.user,
                    observacion=f'Cálculo automático cumplimiento: {bono_trab.dias_trabajados}/{bono_trab.dias_base} días',
                )

            filas_cumplimiento.append({
                'bono_pk': bono_trab.pk,
                'trabajador_pk': trab.pk,
                'nombre_completo': f"{trab.apepat} {trab.apemat} {trab.nombres}".strip(),
                'cargo': trab.cargo or trab.cargo_headcount or '',
                'dias_trabajados': bono_trab.dias_trabajados,
                'dias_operativos': bono_trab.dias_base,
                'bono_base': float(bono_trab.bono_base),
                'tipo_calculo': tipo_calc,
                'secciones': secciones_data,
                'total': float(total_monto_trab),
            })

        if filas_cumplimiento or filas_metraje:
            datos_por_contrato[contrato.nombre_contrato] = {
                'contrato_pk': contrato.pk,
                'periodo_pk': periodo.pk,
                'filas': filas_cumplimiento,
                'filas_metraje': filas_metraje,
            }

    # Preparar estructura de secciones para el header
    secciones_header = []
    for seccion in secciones:
        criterios = seccion.criterios.filter(activo=True).order_by('orden')
        # Para secciones con tabla_calificacion no hay columnas de criterios:
        # solo 2 columnas (dropdown-% y monto)
        num_crit_cols = 0 if seccion.tabla_calificacion else criterios.count()
        secciones_header.append({
            'nombre': seccion.nombre,
            'peso': float(seccion.peso_default),
            'criterios': list(criterios.values_list('nombre', flat=True)),
            'colspan': num_crit_cols + 2,
            'tabla_calificacion': seccion.tabla_calificacion,
        })

    MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    todos_tipos = list(
        TipoBono.objects.filter(activo=True, tipo_calculo='MULTI_CONCEPTO')
        .order_by('codigo').values('pk', 'codigo', 'nombre')
    )
    tipo_anterior = tipo_siguiente = None
    for i, t in enumerate(todos_tipos):
        if t['pk'] == tipo_bono.pk:
            if i > 0:
                tipo_anterior = todos_tipos[i - 1]
            if i < len(todos_tipos) - 1:
                tipo_siguiente = todos_tipos[i + 1]
            break

    context = {
        'tipo_bono': tipo_bono,
        'anio': anio,
        'mes': mes,
        'mes_nombre': MESES[mes],
        'secciones_header': secciones_header,
        'datos_por_contrato': datos_por_contrato,
        'rango_anios': range(2025, date.today().year + 2),
        'from_periodo_pk': from_periodo_pk,
        'tipo_anterior': tipo_anterior,
        'tipo_siguiente': tipo_siguiente,
    }
    return render(request, 'drilling/planilla/cuadro_evaluacion.html', context)


@login_required
def cuadro_guardar(request, tipo_bono_pk):
    """
    Endpoint AJAX para guardar calificaciones de criterios y bono_base.
    POST body JSON: {
        "bonos": [
            {"bono_pk": 1, "bono_base": 1300, "criterios": {"55": true, "56": false, ...}},
            ...
        ]
    }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    bonos_data = data.get('bonos', [])
    actualizados = 0

    try:
        for bd in bonos_data:
            bono_pk = bd.get('bono_pk')
            if not bono_pk:
                continue

            try:
                bono_trab = BonoTrabajador.objects.select_related(
                    'periodo', 'trabajador', 'tipo_bono'
                ).get(pk=bono_pk)
            except BonoTrabajador.DoesNotExist:
                continue

            # Actualizar bono_base si cambió
            nuevo_base = bd.get('bono_base')
            if nuevo_base is not None:
                bono_trab.bono_base = Decimal(str(nuevo_base))
                bono_trab.save(update_fields=['bono_base'])

            # Actualizar días si los envían
            dias_trab = bd.get('dias_trabajados')
            dias_op = bd.get('dias_operativos')
            if dias_trab is not None:
                bono_trab.dias_trabajados = int(dias_trab)
            if dias_op is not None:
                bono_trab.dias_base = int(dias_op)
            if dias_trab is not None or dias_op is not None:
                bono_trab.save(update_fields=['dias_trabajados', 'dias_base'])

            # Actualizar metraje_acumulado (trabajadores de tipo metraje)
            metraje_acumulado = bd.get('metraje_acumulado')
            if metraje_acumulado is not None:
                bono_trab.metraje_acumulado = Decimal(str(metraje_acumulado))
                bono_trab.save(update_fields=['metraje_acumulado'])

            # Actualizar monto_final (total calculado enviado desde el DOM)
            monto_final = bd.get('monto_final')
            if monto_final is not None:
                monto_final_dec = Decimal(str(monto_final))
                bono_trab.monto_calculado = monto_final_dec
                bono_trab.monto_final = monto_final_dec + bono_trab.monto_ajuste
                bono_trab.save(update_fields=['monto_calculado', 'monto_final'])
                bono_trab.registrar_historial(
                    fuente='GUARDAR',
                    usuario=request.user,
                    observacion='Guardado manual desde cuadro de calificación',
                )

            # Actualizar criterios
            criterios_dict = bd.get('criterios', {})
            for crit_pk_str, cumple in criterios_dict.items():
                CalificacionCriterio.objects.filter(
                    bono_trabajador=bono_trab,
                    criterio_id=int(crit_pk_str),
                ).update(cumple=bool(cumple))

            # Actualizar puntajes de tabla fija (INCIDENCIAS / PRODUCCION_META)
            puntajes_tabla = bd.get('puntajes_tabla', {})
            for concepto_pk_str, puntaje_val in puntajes_tabla.items():
                BonoTrabajadorDetalle.objects.update_or_create(
                    bono=bono_trab,
                    concepto_id=int(concepto_pk_str),
                    defaults={'puntaje': Decimal(str(puntaje_val))},
                )

            actualizados += 1

    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('Error en cuadro_guardar')
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

    return JsonResponse({'ok': True, 'actualizados': actualizados})


@login_required
def cuadro_calcular(request, tipo_bono_pk):
    """
    Recalcula todos los bonos de un tipo para un mes/año dado.
    POST params: anio, mes
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)

    import calendar
    from decimal import ROUND_HALF_UP

    tipo_bono = get_object_or_404(TipoBono, pk=tipo_bono_pk)
    anio = int(request.POST.get('anio', date.today().year))
    mes = int(request.POST.get('mes', date.today().month))

    periodos = PeriodoBono.objects.filter(anio=anio, mes=mes)
    user = request.user
    if not user.has_access_to_all_contracts() and user.contrato:
        periodos = periodos.filter(contrato=user.contrato)

    total_calculados = 0
    for periodo in periodos:
        bonos = BonoTrabajador.objects.filter(
            periodo=periodo, tipo_bono=tipo_bono,
        ).select_related('trabajador', 'configuracion')

        for bono in bonos:
            config = bono.configuracion
            if not config:
                continue

            dias_trab = bono.dias_trabajados
            dias_base = bono.dias_base or 30

            if tipo_bono.tipo_calculo == 'MULTI_CONCEPTO':
                from .utils.payroll_engine import _round2, ZERO
                monto, factor = calcular_bono_multi_concepto(bono, config, dias_trab, dias_base)
                bono.factor_cumplimiento = factor
                bono.monto_calculado = monto
                bono.monto_final = _round2(monto + bono.monto_ajuste)
                bono.save()
                bono.registrar_historial(
                    fuente='RECALC',
                    usuario=user,
                    observacion=f'Recalcular manual: factor={float(factor):.4f}',
                )
                total_calculados += 1

    messages.success(request, f'Se recalcularon {total_calculados} bonos para {tipo_bono.nombre}.')
    from_periodo = request.POST.get('from_periodo', '')
    redirect_url = f"{reverse('planilla-cuadro', args=[tipo_bono_pk])}?anio={anio}&mes={mes}"
    if from_periodo:
        redirect_url += f"&from_periodo={from_periodo}"
    return redirect(redirect_url)


# ===========================================
# CONCEPTOS GLOBALES DE CONTRATO
# ===========================================

@login_required
def conceptos_globales(request):
    """
    Vista principal de conceptos globales: muestra todos los indicadores
    por contrato/período con sus valores y % de bono calculado.
    Usa mes operativo: día 26 del mes anterior al día 25 del mes actual.
    Query params: ?anio=2026&mes=4&contrato=ID
    """
    from .models_payroll import ConceptoGlobal, ConceptoGlobalPeriodo
    from .utils.conceptos_globales_engine import (
        inicializar_conceptos_periodo,
        get_rango_mes_operativo,
        _auto_cargar_datos,
        calcular_concepto_global,
    )
    import logging
    logger = logging.getLogger('drilling')

    user = request.user
    today = date.today()
    anio = int(request.GET.get('anio') or today.year)
    mes = int(request.GET.get('mes') or today.month)

    # Rango del mes operativo
    fecha_inicio_op, fecha_fin_op = get_rango_mes_operativo(anio, mes)

    # Contratos disponibles según permisos
    if user.has_access_to_all_contracts():
        contratos = Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato')
    elif user.contrato:
        contratos = Contrato.objects.filter(pk=user.contrato.pk)
    else:
        contratos = Contrato.objects.none()

    contrato_id = request.GET.get('contrato')
    if contrato_id:
        contratos = contratos.filter(pk=contrato_id)

    conceptos = ConceptoGlobal.objects.filter(activo=True).order_by('orden')

    datos_por_contrato = {}
    debug_autocarga = []
    for contrato in contratos:
        # Inicializar si no existen
        inicializar_conceptos_periodo(contrato, anio, mes)

        periodos = ConceptoGlobalPeriodo.objects.filter(
            contrato=contrato, anio=anio, mes=mes
        ).select_related('concepto').order_by('concepto__orden')

        # Siempre recargar datos automáticos (metros, meta, máquinas)
        # para que reflejen el estado actual de los reportes
        for cgp in periodos:
            if cgp.concepto.tipo in ('PRODUCCION', 'CXM'):
                try:
                    _auto_cargar_datos(cgp, anio, mes)
                    calcular_concepto_global(cgp)
                    cgp.save()
                    debug_autocarga.append(
                        f"OK {cgp.concepto.codigo}/{contrato.nombre_contrato}: "
                        f"metros={cgp.metros_acumulados}, meta={cgp.meta_programada}, "
                        f"maquinas={cgp.cantidad_maquinas}"
                    )
                except Exception as e:
                    debug_autocarga.append(
                        f"ERROR {cgp.concepto.codigo}/{contrato.nombre_contrato}: {e}"
                    )
                    logger.error(
                        f"[ConceptosGlobales] Error auto-cargando {cgp.concepto.codigo} "
                        f"contrato={contrato.nombre_contrato}: {e}",
                        exc_info=True,
                    )

        datos_por_contrato[contrato] = list(periodos)

    MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    context = {
        'anio': anio,
        'mes': mes,
        'mes_nombre': MESES[mes],
        'fecha_inicio_op': fecha_inicio_op,
        'fecha_fin_op': fecha_fin_op,
        'conceptos': conceptos,
        'datos_por_contrato': datos_por_contrato,
        'contratos_disponibles': Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato') if user.has_access_to_all_contracts() else contratos,
        'contrato_seleccionado': contrato_id,
        'rango_anios': range(2025, today.year + 2),
        'debug_autocarga': debug_autocarga,
    }
    return render(request, 'drilling/planilla/conceptos_globales.html', context)


@login_required
def conceptos_globales_guardar(request):
    """
    Endpoint AJAX para guardar valores de conceptos globales.
    POST JSON: {
        "registros": [
            {
                "id": 1,
                "metros_acumulados": 1500.0,
                "meta_programada": 1200.0,
                "cantidad_maquinas": 2,
                ...
            },
            ...
        ]
    }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)

    from .models_payroll import ConceptoGlobalPeriodo
    from .utils.conceptos_globales_engine import calcular_concepto_global

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    registros = data.get('registros', [])
    actualizados = 0
    resultados = []

    CAMPOS_NUMERICOS = [
        'metros_acumulados', 'meta_programada', 'cantidad_maquinas',
        'accidentes_incapacitantes', 'eficiencia_cobro',
        'total_abastecido', 'meta_cxm_programada', 'rentabilidad',
    ]
    CAMPOS_ENTEROS = ['cantidad_maquinas', 'accidentes_incapacitantes']

    for reg in registros:
        pk = reg.get('id')
        if not pk:
            continue
        try:
            cgp = ConceptoGlobalPeriodo.objects.select_related('concepto').get(pk=pk)
        except ConceptoGlobalPeriodo.DoesNotExist:
            continue

        # Actualizar campos de entrada
        campos_actualizados = []
        for campo in CAMPOS_NUMERICOS:
            if campo in reg:
                valor = reg[campo]
                if campo in CAMPOS_ENTEROS:
                    setattr(cgp, campo, int(valor))
                else:
                    setattr(cgp, campo, Decimal(str(valor)))
                campos_actualizados.append(campo)

        if reg.get('observaciones') is not None:
            cgp.observaciones = reg['observaciones']
            campos_actualizados.append('observaciones')

        # Recalcular
        valor, porcentaje = calcular_concepto_global(cgp)
        campos_actualizados.extend(['valor_calculado', 'porcentaje_bono', 'updated_at'])
        cgp.save(update_fields=campos_actualizados)
        actualizados += 1

        resultados.append({
            'id': cgp.pk,
            'codigo': cgp.concepto.codigo,
            'valor_calculado': float(valor),
            'porcentaje_bono': float(porcentaje),
        })

    return JsonResponse({'ok': True, 'actualizados': actualizados, 'resultados': resultados})


@login_required
def conceptos_globales_calcular(request):
    """
    Recalcula todos los conceptos globales de un contrato para un período.
    POST params: contrato_id, anio, mes
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)

    from .utils.conceptos_globales_engine import calcular_todos_conceptos_contrato

    contrato_id = request.POST.get('contrato_id')
    anio = int(request.POST.get('anio', date.today().year))
    mes = int(request.POST.get('mes', date.today().month))

    contrato = get_object_or_404(Contrato, pk=contrato_id)
    resultados = calcular_todos_conceptos_contrato(contrato, anio, mes)

    messages.success(request, f'Se calcularon {len(resultados)} conceptos globales para {contrato.nombre_contrato}.')
    return redirect(f"{reverse('planilla-conceptos-globales')}?anio={anio}&mes={mes}&contrato={contrato_id}")


@login_required
def api_conceptos_globales_contrato(request, contrato_id, anio, mes):
    """
    API que retorna los conceptos globales calculados de un contrato/período.
    Útil para que otros módulos consulten los % de bono.
    """
    from .models_payroll import ConceptoGlobalPeriodo

    periodos = ConceptoGlobalPeriodo.objects.filter(
        contrato_id=contrato_id, anio=anio, mes=mes
    ).select_related('concepto').order_by('concepto__orden')

    data = {}
    for cgp in periodos:
        data[cgp.concepto.codigo] = {
            'id': cgp.pk,
            'nombre': cgp.concepto.nombre,
            'tipo': cgp.concepto.tipo,
            'valor_calculado': float(cgp.valor_calculado),
            'porcentaje_bono': float(cgp.porcentaje_bono),
            'inputs': {
                'metros_acumulados': float(cgp.metros_acumulados),
                'meta_programada': float(cgp.meta_programada),
                'cantidad_maquinas': cgp.cantidad_maquinas,
                'accidentes_incapacitantes': cgp.accidentes_incapacitantes,
                'eficiencia_cobro': float(cgp.eficiencia_cobro),
                'total_abastecido': float(cgp.total_abastecido),
                'meta_cxm_programada': float(cgp.meta_cxm_programada),
                'rentabilidad': float(cgp.rentabilidad),
            },
        }

    return JsonResponse(data)


@login_required
def api_diagnostico_conceptos(request):
    """
    Diagnóstico: muestra qué datos encuentra el sistema para cargar metros y metas.
    GET params: contrato_id, anio, mes
    """
    from django.db.models import Sum, Count
    from .models import TurnoAvance, Turno, Maquina, ProgramacionMes, MetaTurno
    from .utils.periodo_operativo import get_rango_mes_operativo

    contrato_id = request.GET.get('contrato_id')
    anio = int(request.GET.get('anio', date.today().year))
    mes = int(request.GET.get('mes', date.today().month))

    # Acepta ID numérico o nombre parcial del contrato
    if contrato_id and contrato_id.isdigit():
        contrato = get_object_or_404(Contrato, pk=int(contrato_id))
    elif contrato_id:
        contrato = Contrato.objects.filter(
            nombre_contrato__icontains=contrato_id
        ).first()
        if not contrato:
            return JsonResponse({'error': f'No se encontró contrato con nombre "{contrato_id}"'}, status=404)
    else:
        # Listar contratos disponibles
        contratos = list(Contrato.objects.filter(estado='ACTIVO').values('id', 'nombre_contrato'))
        return JsonResponse({'error': 'Falta contrato_id', 'contratos_disponibles': contratos})
    fecha_inicio, fecha_fin = get_rango_mes_operativo(anio, mes)

    # 1. Máquinas operativas
    maquinas = list(Maquina.objects.filter(
        contrato=contrato, estado='OPERATIVO'
    ).values('id', 'nombre', 'estado'))

    maquinas_todas = list(Maquina.objects.filter(
        contrato=contrato
    ).values('id', 'nombre', 'estado'))

    # 2. Turnos en el período
    turnos_por_estado = list(Turno.objects.filter(
        contrato=contrato,
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin,
    ).values('estado').annotate(cnt=Count('id')))

    # 3. TurnoAvance en el período (todos los estados)
    avance_todos = TurnoAvance.objects.filter(
        turno__contrato=contrato,
        turno__fecha__gte=fecha_inicio,
        turno__fecha__lte=fecha_fin,
    ).aggregate(
        total=Sum('metros_perforados'),
        count=Count('id'),
    )

    # 4. TurnoAvance solo COMPLETADO/APROBADO
    avance_filtrado = TurnoAvance.objects.filter(
        turno__contrato=contrato,
        turno__fecha__gte=fecha_inicio,
        turno__fecha__lte=fecha_fin,
        turno__estado__in=['COMPLETADO', 'APROBADO'],
    ).aggregate(
        total=Sum('metros_perforados'),
        count=Count('id'),
    )

    # 5. Muestra de TurnoAvance
    samples = list(TurnoAvance.objects.filter(
        turno__contrato=contrato,
        turno__fecha__gte=fecha_inicio,
        turno__fecha__lte=fecha_fin,
    ).select_related('turno', 'turno__maquina').values(
        'turno__fecha', 'turno__estado', 'turno__maquina__nombre',
        'metros_perforados',
    ).order_by('-turno__fecha')[:10])

    # 6. ProgramacionMes
    maquinas_ids = [m['id'] for m in maquinas]
    progs = list(ProgramacionMes.objects.filter(
        maquina_id__in=maquinas_ids, año=anio, mes=mes,
    ).values('maquina__nombre', 'meta_metros', 'dia_inicio'))

    # 7. Turnos sin TurnoAvance (posible causa: reportes sin avance creado)
    turnos_sin_avance = Turno.objects.filter(
        contrato=contrato,
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin,
    ).exclude(
        avance__isnull=False,
    ).count()

    # 8. Test directo de las funciones de auto-carga
    from .utils.conceptos_globales_engine import (
        cargar_metros_acumulados,
        cargar_meta_programada,
        cargar_cantidad_maquinas,
    )
    test_autocarga = {}
    try:
        metros = cargar_metros_acumulados(contrato, fecha_inicio, fecha_fin)
        test_autocarga['metros_acumulados'] = float(metros)
    except Exception as e:
        test_autocarga['metros_acumulados_error'] = str(e)

    try:
        meta = cargar_meta_programada(contrato, anio, mes)
        test_autocarga['meta_programada'] = float(meta)
    except Exception as e:
        test_autocarga['meta_programada_error'] = str(e)

    try:
        maq = cargar_cantidad_maquinas(contrato)
        test_autocarga['cantidad_maquinas'] = maq
    except Exception as e:
        test_autocarga['cantidad_maquinas_error'] = str(e)

    # 9. Estado actual de ConceptoGlobalPeriodo
    from .models_payroll import ConceptoGlobalPeriodo
    cgps = list(ConceptoGlobalPeriodo.objects.filter(
        contrato=contrato, anio=anio, mes=mes,
    ).select_related('concepto').values(
        'id', 'concepto__codigo', 'concepto__tipo',
        'metros_acumulados', 'meta_programada', 'cantidad_maquinas',
        'valor_calculado', 'porcentaje_bono',
    ))

    return JsonResponse({
        'contrato': contrato.nombre_contrato,
        'periodo': f'{fecha_inicio} a {fecha_fin}',
        'anio': anio,
        'mes': mes,
        'maquinas_operativas': maquinas,
        'maquinas_todas': maquinas_todas,
        'turnos_por_estado': turnos_por_estado,
        'turno_avance_todos_estados': {
            'total_metros': float(avance_todos['total'] or 0),
            'count': avance_todos['count'],
        },
        'turno_avance_completado_aprobado': {
            'total_metros': float(avance_filtrado['total'] or 0),
            'count': avance_filtrado['count'],
        },
        'muestra_avances': samples,
        'programacion_mes': progs,
        'turnos_sin_avance': turnos_sin_avance,
        'test_autocarga': test_autocarga,
        'conceptos_globales_periodo_actual': cgps,
    })


# ===========================================
# ESTRUCTURA SALARIAL — CRUD
# ===========================================


def _sync_sueldo_trabajadores(estructura):
    """Actualiza Trabajador.sueldo para todos los activos que coincidan
    con el contrato + centro_costo + cargo de la estructura."""
    cc = estructura.contrato_servicio.codigo_centro_costo
    actualizados = Trabajador.objects.filter(
        contrato=estructura.contrato,
        centro_costo=cc,
        cargo=estructura.cargo_contratado,
        estado='ACTIVO',
    ).update(sueldo=estructura.sueldo_basico)
    return actualizados

@login_required
def estructura_salarial_list(request):
    """Lista de todas las estructuras salariales."""
    from django.db.models import Exists, OuterRef
    user = request.user
    contrato_filter = request.GET.get('contrato', '')

    estructuras = EstructuraSalarial.objects.select_related(
        'contrato', 'contrato_servicio', 'maquina', 'creado_por'
    ).annotate(
        tiene_variantes_maquina=Exists(
            EstructuraSalarial.objects.filter(
                contrato=OuterRef('contrato'),
                contrato_servicio=OuterRef('contrato_servicio'),
                cargo_contratado=OuterRef('cargo_contratado'),
                maquina__isnull=False,
            )
        )
    ).order_by('contrato__nombre_contrato', 'contrato_servicio__tipo_servicio', 'cargo_contratado', 'maquina__nombre')

    if not user.has_access_to_all_contracts() and user.contrato:
        estructuras = estructuras.filter(contrato=user.contrato)
    elif contrato_filter:
        estructuras = estructuras.filter(contrato_id=contrato_filter)

    if user.has_access_to_all_contracts():
        contratos = Contrato.objects.exclude(estado='FINALIZADO').order_by('nombre_contrato')
    elif user.contrato:
        contratos = Contrato.objects.filter(pk=user.contrato.pk)
    else:
        contratos = Contrato.objects.none()

    return render(request, 'drilling/planilla/estructura_salarial_list.html', {
        'estructuras': estructuras,
        'contratos': contratos,
        'contrato_filter': contrato_filter,
    })


@login_required
def estructura_salarial_create(request):
    """Crear/actualizar estructuras salariales en bloque para un contrato+CTR."""
    from .models import ContratoServicio

    if request.method == 'POST':
        from .models import Maquina
        contrato_id = request.POST.get('contrato')
        ctr_id = request.POST.get('ctr_id')
        maquina_id = request.POST.get('maquina') or None

        contrato = get_object_or_404(Contrato, pk=contrato_id)
        ctr = get_object_or_404(ContratoServicio, pk=ctr_id)
        maquina = Maquina.objects.filter(pk=maquina_id, contrato=contrato).first() if maquina_id else None

        # Recopilar datos de cada fila de cargo del POST
        i = 0
        creados = 0
        actualizados = 0
        while f'cargo_{i}' in request.POST:
            if not request.POST.get(f'incluir_{i}'):
                i += 1
                continue

            cargo = request.POST[f'cargo_{i}'].strip()
            if not cargo:
                i += 1
                continue

            sueldo = request.POST.get(f'sueldo_basico_{i}', '0') or '0'
            bono = request.POST.get(f'bono_por_metraje_{i}', '0') or '0'
            metraje = request.POST.get(f'metraje_base_{i}', '0') or '0'
            bonif = request.POST.get(f'bonificacion_area_{i}', '0') or '0'

            existing = EstructuraSalarial.objects.filter(
                contrato=contrato,
                contrato_servicio=ctr,
                cargo_contratado=cargo,
                maquina=maquina,
            ).first()

            if existing:
                existing.guardar_historial(request.user, 'Actualización masiva')
                existing.sueldo_basico = Decimal(sueldo)
                existing.bono_por_metraje = Decimal(bono)
                existing.metraje_base = Decimal(metraje)
                existing.bonificacion_area = Decimal(bonif)
                existing.version += 1
                existing.save()
                _sync_sueldo_trabajadores(existing)
                actualizados += 1
            else:
                est = EstructuraSalarial.objects.create(
                    contrato=contrato,
                    contrato_servicio=ctr,
                    cargo_contratado=cargo,
                    maquina=maquina,
                    sueldo_basico=Decimal(sueldo),
                    bono_por_metraje=Decimal(bono),
                    metraje_base=Decimal(metraje),
                    bonificacion_area=Decimal(bonif),
                    creado_por=request.user,
                    version=1,
                )
                est.guardar_historial(request.user, 'Creación inicial')
                _sync_sueldo_trabajadores(est)
                creados += 1
            i += 1

        partes = []
        if creados:
            partes.append(f'{creados} creadas')
        if actualizados:
            partes.append(f'{actualizados} actualizadas')
        if partes:
            messages.success(request, f'Estructuras salariales: {", ".join(partes)}.')
        else:
            messages.warning(request, 'No se seleccionó ningún cargo.')
        return redirect('planilla-estructura-salarial-list')

    # GET: mostrar selector de contrato + CTR
    user = request.user
    if user.has_access_to_all_contracts():
        contratos = Contrato.objects.exclude(estado='FINALIZADO').order_by('nombre_contrato')
    elif user.contrato:
        contratos = Contrato.objects.filter(pk=user.contrato.pk)
    else:
        contratos = Contrato.objects.none()

    from .models import Maquina
    return render(request, 'drilling/planilla/estructura_salarial_form.html', {
        'contratos': contratos,
        'titulo': 'Nueva Estructura Salarial',
    })


@login_required
def estructura_salarial_edit(request, pk):
    """Editar estructura salarial existente (genera historial de versión)."""
    estructura = get_object_or_404(EstructuraSalarial, pk=pk)

    if request.method == 'POST':
        form = EstructuraSalarialForm(request.POST, instance=estructura)
        if form.is_valid():
            motivo = form.cleaned_data.get('motivo_cambio', '')
            if not motivo:
                form.add_error('motivo_cambio', 'Debe indicar el motivo del cambio al editar.')
                cargos = list(
                    Trabajador.objects.filter(estado='ACTIVO')
                    .values_list('cargo', flat=True).distinct().order_by('cargo')
                )
                return render(request, 'drilling/planilla/estructura_salarial_form.html', {
                    'form': form,
                    'titulo': f'Editar Estructura Salarial — {estructura.cargo_contratado}',
                    'estructura': estructura,
                    'cargos': cargos,
                })

            # Guardar snapshot de la versión anterior
            estructura.guardar_historial(
                usuario=request.user,
                motivo=motivo,
            )
            # Incrementar versión y guardar
            estructura = form.save(commit=False)
            estructura.version += 1
            estructura.save()
            _sync_sueldo_trabajadores(estructura)
            messages.success(request, f'Estructura salarial actualizada a v{estructura.version}.')
            return redirect('planilla-estructura-salarial-list')
    else:
        form = EstructuraSalarialForm(instance=estructura)

    cargos = list(
        Trabajador.objects.filter(estado='ACTIVO')
        .values_list('cargo', flat=True).distinct().order_by('cargo')
    )

    return render(request, 'drilling/planilla/estructura_salarial_form.html', {
        'form': form,
        'titulo': f'Editar Estructura Salarial — {estructura.cargo_contratado}',
        'estructura': estructura,
        'cargos': cargos,
    })


@login_required
def estructura_salarial_historial(request, pk):
    """Ver historial de versiones de una estructura salarial."""
    estructura = get_object_or_404(
        EstructuraSalarial.objects.select_related('contrato', 'contrato_servicio'),
        pk=pk
    )
    historial = estructura.historial.select_related('modificado_por').order_by('-version')

    return render(request, 'drilling/planilla/estructura_salarial_historial.html', {
        'estructura': estructura,
        'historial': historial,
    })


@login_required
def api_ctr_por_contrato(request, contrato_id):
    """API: devuelve los CTR (ContratoServicio) de un contrato, para filtrar dinámicamente.
    Incluye tanto los ContratoServicio existentes como el CC principal del Contrato si no está ya registrado.
    """
    from .models import Contrato, ContratoServicio

    ctrs = list(ContratoServicio.objects.filter(
        contrato_id=contrato_id, activo=True
    ).values('id', 'tipo_servicio', 'codigo_centro_costo', 'descripcion'))

    # Safety net: si el CC principal del contrato no está en ContratoServicio, agregarlo
    try:
        contrato = Contrato.objects.get(pk=contrato_id)
        if contrato.codigo_centro_costo:
            cc_existentes = {c['codigo_centro_costo'] for c in ctrs}
            if contrato.codigo_centro_costo not in cc_existentes:
                ctrs.insert(0, {
                    'id': None,
                    'tipo_servicio': 'DDH',
                    'codigo_centro_costo': contrato.codigo_centro_costo,
                    'descripcion': f'CTR Principal ({contrato.nombre_contrato})',
                })
    except Contrato.DoesNotExist:
        pass

    return JsonResponse(ctrs, safe=False)


@login_required
def api_maquinas_por_contrato(request, contrato_id):
    """API: devuelve las máquinas activas de un contrato para el selector de estructura salarial."""
    from .models import Maquina
    maquinas = list(Maquina.objects.filter(
        contrato_id=contrato_id,
    ).exclude(estado='FUERA_SERVICIO').values('id', 'nombre', 'tipo', 'estado').order_by('nombre'))
    return JsonResponse(maquinas, safe=False)


@login_required
def api_cargos_por_ctr(request, contrato_id, ctr_id):
    """API: devuelve los cargos distintos de trabajadores activos para un contrato+CTR,
    junto con la estructura salarial existente si la hay."""
    from django.db.models import Count
    from .models import ContratoServicio, Maquina

    ctr = get_object_or_404(ContratoServicio, pk=ctr_id)
    codigo_cc = ctr.codigo_centro_costo
    maquina_id = request.GET.get('maquina_id') or None
    maquina = Maquina.objects.filter(pk=maquina_id).first() if maquina_id else None

    cargos = (
        Trabajador.objects.filter(
            contrato_id=contrato_id,
            centro_costo=codigo_cc,
            estado='ACTIVO',
        )
        .exclude(cargo='')
        .values('cargo')
        .annotate(cantidad=Count('id'))
        .order_by('cargo')
    )

    result = []
    for c in cargos:
        cargo_nombre = c['cargo']
        existing = EstructuraSalarial.objects.filter(
            contrato_id=contrato_id,
            contrato_servicio=ctr,
            cargo_contratado=cargo_nombre,
            maquina=maquina,
        ).first()

        # Estructura general (maquina=None) — usada como fallback cuando se selecciona máquina
        general = None
        if maquina:
            general = EstructuraSalarial.objects.filter(
                contrato_id=contrato_id,
                contrato_servicio=ctr,
                cargo_contratado=cargo_nombre,
                maquina__isnull=True,
            ).first()

        # Cuántas máquinas específicas ya tiene este cargo
        maquinas_count = EstructuraSalarial.objects.filter(
            contrato_id=contrato_id,
            contrato_servicio=ctr,
            cargo_contratado=cargo_nombre,
            maquina__isnull=False,
        ).count()

        item = {
            'cargo': cargo_nombre,
            'cantidad': c['cantidad'],
            'tiene_estructura': existing is not None,
            'tiene_general': general is not None,
            'maquinas_count': maquinas_count,
        }
        if existing:
            item['estructura'] = {
                'id': existing.pk,
                'sueldo_basico': str(existing.sueldo_basico),
                'bono_por_metraje': str(existing.bono_por_metraje),
                'metraje_base': str(existing.metraje_base),
                'bonificacion_area': str(existing.bonificacion_area),
                'version': existing.version,
            }
        if general:
            item['estructura_general'] = {
                'sueldo_basico': str(general.sueldo_basico),
                'bono_por_metraje': str(general.bono_por_metraje),
                'metraje_base': str(general.metraje_base),
                'bonificacion_area': str(general.bonificacion_area),
            }
        result.append(item)

    return JsonResponse({'ctr_id': ctr.pk, 'codigo_cc': codigo_cc, 'cargos': result})


# ===========================================
# APIs PARA CONFIGURACIÓN DE BONOS
# ===========================================

@login_required
def api_cargos_activos_contrato(request, contrato_id):
    """
    API: devuelve los cargos distintos de trabajadores ACTIVOS en un contrato.
    Usado para poblar el selector de cargos en el formulario de configuración de bono.
    """
    cargos = (
        Trabajador.objects.filter(contrato_id=contrato_id, estado='ACTIVO')
        .exclude(cargo__isnull=True)
        .exclude(cargo='')
        .values_list('cargo', flat=True)
        .distinct()
        .order_by('cargo')
    )
    return JsonResponse({'cargos': list(cargos)})


@login_required
def api_bonificacion_area_cargo(request):
    """
    API: dado un contrato_id y una lista de cargos, devuelve el bonificacion_area
    desde EstructuraSalarial para cada cargo.
    GET params: contrato_id, cargos (lista separada por comas)

    Retorna: {
        "cargos": {
            "OPERADOR": {"bonificacion_area": "1200.00", "encontrado": true},
            "ASISTENTE": {"bonificacion_area": "0.00", "encontrado": false},
        },
        "sugerido": "1200.00"   <- primer valor encontrado, o promedio si difieren
    }
    """
    contrato_id = request.GET.get('contrato_id')
    cargos_param = request.GET.get('cargos', '')

    if not contrato_id:
        return JsonResponse({'error': 'contrato_id requerido'}, status=400)

    cargos_list = [c.strip() for c in cargos_param.split(',') if c.strip()]
    if not cargos_list:
        return JsonResponse({'cargos': {}, 'sugerido': None})

    resultado = {}
    valores = []

    for cargo in cargos_list:
        # Buscar en estructura salarial; si hay varias entradas (distintos CTR)
        # tomamos el primer valor no nulo de bonificacion_area
        estructura = (
            EstructuraSalarial.objects.filter(
                contrato_id=contrato_id,
                cargo_contratado=cargo,
                activo=True,
            )
            .order_by('-bonificacion_area')  # primero los de mayor valor
            .first()
        )
        if estructura:
            val = float(estructura.bonificacion_area or 0)
            resultado[cargo] = {
                'bonificacion_area': str(estructura.bonificacion_area or '0.00'),
                'bono_por_metraje':  str(estructura.bono_por_metraje or '0.00'),
                'metraje_base':      str(estructura.metraje_base or '0.00'),
                'sueldo_basico':     str(estructura.sueldo_basico or '0.00'),
                'encontrado': True,
            }
            if val:
                valores.append(val)
        else:
            resultado[cargo] = {
                'bonificacion_area': '0.00',
                'bono_por_metraje':  '0.00',
                'metraje_base':      '0.00',
                'sueldo_basico':     '0.00',
                'encontrado': False,
            }

    # Valor sugerido: si todos son iguales → ese valor; si difieren → el mayor
    sugerido = None
    if valores:
        sugerido = f"{max(valores):.2f}"

    return JsonResponse({'cargos': resultado, 'sugerido': sugerido})


@login_required
def api_trabajadores_por_cargo(request):
    """
    API: dado un contrato_id y una lista de cargos (opcionales), devuelve los
    trabajadores ACTIVOS del contrato agrupados por cargo, con nombre y DNI.
    GET params:
        contrato_id  — requerido
        cargos       — comma-separated; vacío = todos los cargos
    Response: { cargos: { "CARGO": [{ dni, nombre }, ...] } }
    """
    contrato_id = request.GET.get('contrato_id')
    cargos_param = request.GET.get('cargos', '')

    if not contrato_id:
        return JsonResponse({'error': 'contrato_id requerido'}, status=400)

    cargos_list = [c.strip() for c in cargos_param.split(',') if c.strip()]

    from django.db.models import Q
    qs = Trabajador.objects.filter(
        contrato_id=contrato_id, estado='ACTIVO'
    ).order_by('cargo', 'apepat', 'apemat', 'nombres')

    if cargos_list:
        qs = qs.filter(
            Q(cargo__in=cargos_list) | Q(cargo_headcount__in=cargos_list)
        )

    resultado = {}
    for trab in qs:
        cargo = trab.cargo_headcount or trab.cargo or ''
        if not cargo:
            continue
        if cargo not in resultado:
            resultado[cargo] = []
        resultado[cargo].append({
            'dni': trab.dni,
            'nombre': f"{trab.apepat} {trab.apemat} {trab.nombres}".strip(),
        })

    return JsonResponse({'cargos': resultado})


# ===========================================
# ETAPA 0 — VALIDACIÓN PRE-CÁLCULO
# ===========================================

@login_required
def periodo_validar(request, pk):
    """
    Muestra el diagnóstico de alertas para un período antes de calcular.
    Permite al usuario identificar y corregir datos faltantes.
    """
    from .utils.payroll_engine import validar_periodo
    periodo = get_object_or_404(PeriodoBono, pk=pk)
    validacion = validar_periodo(periodo)
    context = {
        'periodo': periodo,
        'validacion': validacion,
    }
    return render(request, 'drilling/planilla/periodo_validacion.html', context)


# ===========================================
# ETAPA 3 — RESUMEN CONSOLIDADO POR TRABAJADOR
# ===========================================

@login_required
def periodo_resumen_trabajador(request, pk, trabajador_pk):
    """
    Vista consolidada de bonos de un trabajador en un período.
    Muestra: sueldo básico referencial + bonos remunerativos + extraordinarios.
    """
    from .utils.payroll_engine import resumen_por_trabajador
    periodo = get_object_or_404(PeriodoBono, pk=pk)
    trabajador = get_object_or_404(Trabajador, pk=trabajador_pk)
    datos = resumen_por_trabajador(periodo, trabajador)
    context = {
        'periodo': periodo,
        'trabajador': trabajador,
        'datos': datos,
    }
    return render(request, 'drilling/planilla/resumen_trabajador.html', context)


# ===========================================
# ETAPA 4 — CONCILIACIÓN TAREO-PAYROLL
# ===========================================

@login_required
def periodo_conciliacion(request, pk):
    """
    Compara los días calculados en payroll contra los registros de asistencia.
    Detecta desajustes entre cierre de tareo y período de bonos.
    """
    from .utils.payroll_engine import conciliar_periodo_tareo
    periodo = get_object_or_404(PeriodoBono, pk=pk)
    conciliacion = conciliar_periodo_tareo(periodo)
    context = {
        'periodo': periodo,
        'conciliacion': conciliacion,
    }
    return render(request, 'drilling/planilla/conciliacion.html', context)


# ===========================================
# ETAPA 5 — BOLETA DE PAGO HTML IMPRIMIBLE
# ===========================================

@login_required
def periodo_boleta(request, pk, trabajador_pk):
    """
    Genera la boleta de pago imprimible de un trabajador para el período.
    Muestra: sueldo básico + bonos desglosados por tipo + total.
    """
    from .utils.payroll_engine import resumen_por_trabajador
    periodo = get_object_or_404(PeriodoBono, pk=pk)
    trabajador = get_object_or_404(Trabajador, pk=trabajador_pk)
    datos = resumen_por_trabajador(periodo, trabajador)
    context = {
        'periodo': periodo,
        'trabajador': trabajador,
        'datos': datos,
        'es_boleta': True,
    }
    return render(request, 'drilling/planilla/boleta_pago.html', context)


# ===========================================
# ETAPA 6 — PRESUPUESTO VS REAL
# ===========================================

@login_required
def presupuesto_planilla(request):
    """
    Vista de comparación presupuesto vs real de planilla por contrato y mes.
    Muestra: monto presupuestado ingresado vs monto real calculado en PeriodoBono.
    """
    from django.db.models import Sum
    from .models_payroll import PresupuestoPlanilla

    user = request.user
    contrato = user.contrato
    today = date.today()

    anio = int(request.GET.get('anio', today.year))
    mes = int(request.GET.get('mes', today.month))

    # Contratos visibles
    if user.has_access_to_all_contracts():
        from .models import Contrato
        contratos = Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato')
    elif contrato:
        contratos = Contrato.objects.filter(pk=contrato.pk)
    else:
        contratos = Contrato.objects.none()

    filas = []
    for c in contratos:
        presupuestos = PresupuestoPlanilla.objects.filter(
            contrato=c, anio=anio, mes=mes
        )
        total_presupuestado = presupuestos.aggregate(total=Sum('monto_presupuestado'))['total'] or 0

        periodo_real = PeriodoBono.objects.filter(
            contrato=c, anio=anio, mes=mes
        ).first()

        total_real = 0
        if periodo_real:
            total_real = BonoTrabajador.objects.filter(
                periodo=periodo_real
            ).aggregate(total=Sum('monto_final'))['total'] or 0

        desviacion_abs = float(total_real) - float(total_presupuestado)
        desviacion_pct = (desviacion_abs / float(total_presupuestado) * 100) if total_presupuestado else None

        filas.append({
            'contrato': c,
            'total_presupuestado': total_presupuestado,
            'total_real': total_real,
            'desviacion_abs': desviacion_abs,
            'desviacion_pct': desviacion_pct,
            'periodo': periodo_real,
            'presupuestos_detalle': list(presupuestos),
        })

    # Formulario para ingresar presupuesto
    if request.method == 'POST':
        contrato_id = request.POST.get('contrato_id')
        concepto = request.POST.get('concepto', 'BONOS')
        monto = request.POST.get('monto', '0')
        anio_post = int(request.POST.get('anio', anio))
        mes_post = int(request.POST.get('mes', mes))
        try:
            from .models import Contrato as ContratoModel
            contrato_obj = ContratoModel.objects.get(pk=contrato_id)
            PresupuestoPlanilla.objects.update_or_create(
                contrato=contrato_obj,
                anio=anio_post,
                mes=mes_post,
                concepto=concepto,
                defaults={
                    'monto_presupuestado': monto,
                    'registrado_por': user,
                }
            )
            messages.success(request, 'Presupuesto guardado.')
        except Exception as e:
            messages.error(request, f'Error al guardar presupuesto: {e}')
        return redirect(f"{request.path}?anio={anio_post}&mes={mes_post}")

    context = {
        'filas': filas,
        'anio': anio,
        'mes': mes,
        'contratos': contratos,
        'meses': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'anios': range(today.year - 2, today.year + 2),
    }
    return render(request, 'drilling/planilla/presupuesto.html', context)


# =============================================================================
# ASIGNACIÓN DE ESTRUCTURA SALARIAL POR TRABAJADOR
# =============================================================================

@login_required
def asignacion_estructura_list(request):
    """
    Lista de trabajadores del contrato con su asignación de estructura salarial.
    Muestra: asignados (con estructura específica), sin asignación explícita.
    """
    from django.db.models import Prefetch

    user = request.user
    contrato_filter = request.GET.get('contrato', '')
    cargo_filter = request.GET.get('cargo', '')

    # Determinar contrato activo
    if not user.has_access_to_all_contracts() and user.contrato:
        contrato_activo = user.contrato
    elif contrato_filter:
        contrato_activo = get_object_or_404(Contrato, pk=contrato_filter)
    else:
        contrato_activo = None

    trabajadores_qs = Trabajador.objects.none()
    if contrato_activo:
        trabajadores_qs = Trabajador.objects.filter(
            contrato=contrato_activo,
        ).select_related('maquina_asignada').prefetch_related(
            Prefetch(
                'asignaciones_estructura_salarial',
                queryset=AsignacionEstructuraSalarial.objects.filter(
                    activo=True
                ).select_related('estructura_salarial__maquina', 'estructura_salarial__contrato_servicio'),
                to_attr='asignacion_activa_list',
            )
        ).order_by('cargo', 'apepat', 'nombres')

        if cargo_filter:
            trabajadores_qs = trabajadores_qs.filter(cargo=cargo_filter)

    # Lista de cargos únicos para el filtro
    cargos_disponibles = []
    if contrato_activo:
        cargos_disponibles = (
            Trabajador.objects.filter(contrato=contrato_activo)
            .values_list('cargo', flat=True)
            .distinct()
            .order_by('cargo')
        )

    if user.has_access_to_all_contracts():
        contratos = Contrato.objects.filter(estado='ACTIVO').order_by('nombre_contrato')
    elif user.contrato:
        contratos = Contrato.objects.filter(pk=user.contrato.pk)
    else:
        contratos = Contrato.objects.none()

    # Enriquecer cada trabajador con su asignación activa (si existe)
    trabajadores_con_asignacion = []
    for trab in trabajadores_qs:
        asignacion = trab.asignacion_activa_list[0] if trab.asignacion_activa_list else None
        trabajadores_con_asignacion.append({
            'trabajador': trab,
            'asignacion': asignacion,
        })

    return render(request, 'drilling/planilla/asignacion_estructura_list.html', {
        'trabajadores': trabajadores_con_asignacion,
        'contrato_activo': contrato_activo,
        'contratos': contratos,
        'contrato_filter': str(contrato_activo.pk) if contrato_activo else '',
        'cargos_disponibles': cargos_disponibles,
        'cargo_filter': cargo_filter,
    })


@login_required
def asignacion_estructura_asignar(request, trabajador_pk):
    """
    Crear o reemplazar la asignación de estructura salarial para un trabajador.
    Si ya existe una asignación activa, la desactiva y crea una nueva.
    """
    from django.utils import timezone

    trabajador = get_object_or_404(Trabajador, pk=trabajador_pk)
    contrato = trabajador.contrato
    cargo = trabajador.cargo or ''

    # Asignación activa previa (si existe)
    asignacion_previa = AsignacionEstructuraSalarial.objects.filter(
        trabajador=trabajador,
        activo=True,
    ).select_related('estructura_salarial__maquina').first()

    # Estructuras disponibles para este cargo y contrato
    estructuras_disponibles = EstructuraSalarial.objects.filter(
        contrato=contrato,
        cargo_contratado=cargo,
        activo=True,
    ).select_related('maquina', 'contrato_servicio').order_by('maquina__nombre')

    # Sugerir la estructura que coincide con la máquina asignada al trabajador
    estructura_sugerida = None
    if not asignacion_previa and trabajador.maquina_asignada_id:
        estructura_sugerida = estructuras_disponibles.filter(
            maquina=trabajador.maquina_asignada
        ).first()
        if not estructura_sugerida:
            estructura_sugerida = estructuras_disponibles.filter(maquina__isnull=True).first()

    if request.method == 'POST':
        form = AsignacionEstructuraSalarialForm(
            request.POST,
            contrato=contrato,
            cargo=cargo,
            maquina_trabajador=trabajador.maquina_asignada,
        )
        if form.is_valid():
            if getattr(form, 'add_warning', False):
                messages.warning(
                    request,
                    f'Asignación guardada, pero la estructura seleccionada corresponde a una máquina distinta '
                    f'a la operativa de {trabajador.apepat} {trabajador.nombres}. Revisa si es correcto.'
                )
            # Desactivar asignación previa
            if asignacion_previa:
                asignacion_previa.activo = False
                asignacion_previa.fecha_fin = form.cleaned_data['fecha_inicio']
                asignacion_previa.save(update_fields=['activo', 'fecha_fin'])

            nueva = form.save(commit=False)
            nueva.contrato = contrato
            nueva.trabajador = trabajador
            nueva.activo = True
            nueva.creado_por = request.user
            nueva.save()

            messages.success(
                request,
                f'Estructura salarial asignada correctamente a {trabajador.apepat} {trabajador.nombres}.'
            )
            return redirect(reverse('planilla-hub'))
    else:
        form = AsignacionEstructuraSalarialForm(
            initial={
                'trabajador': trabajador.pk,
                'fecha_inicio': date.today(),
                'estructura_salarial': (
                    asignacion_previa.estructura_salarial_id if asignacion_previa
                    else (estructura_sugerida.pk if estructura_sugerida else None)
                ),
            },
            contrato=contrato,
            cargo=cargo,
            maquina_trabajador=trabajador.maquina_asignada,
        )

    return render(request, 'drilling/planilla/asignacion_estructura_form.html', {
        'form': form,
        'trabajador': trabajador,
        'contrato': contrato,
        'asignacion_previa': asignacion_previa,
        'estructuras_disponibles': estructuras_disponibles,
        'estructura_sugerida': estructura_sugerida,
    })


@login_required
def asignacion_estructura_desactivar(request, pk):
    """Desactiva una asignación activa (quita la asignación explícita del trabajador)."""
    asignacion = get_object_or_404(AsignacionEstructuraSalarial, pk=pk, activo=True)
    contrato_pk = asignacion.contrato_id
    if request.method == 'POST':
        asignacion.activo = False
        asignacion.save(update_fields=['activo'])
        messages.success(request, 'Asignación desactivada.')
    return redirect(reverse('planilla-hub'))


@login_required
def api_estructuras_por_cargo(request):
    """
    API: devuelve las estructuras salariales disponibles para un cargo dado en un contrato.
    Parámetros: ?contrato_id=X&cargo=AYUDANTE+DDH-I
    """
    contrato_id = request.GET.get('contrato_id')
    cargo = request.GET.get('cargo', '').strip()

    if not contrato_id or not cargo:
        return JsonResponse({'estructuras': []})

    estructuras = EstructuraSalarial.objects.filter(
        contrato_id=contrato_id,
        cargo_contratado=cargo,
        activo=True,
    ).select_related('maquina', 'contrato_servicio').order_by('maquina__nombre')

    data = []
    for e in estructuras:
        data.append({
            'id': e.pk,
            'label': f"{e.cargo_contratado} — {'Máq: ' + e.maquina.nombre if e.maquina else 'General'} | CTR: {e.contrato_servicio.codigo_centro_costo}",
            'maquina': e.maquina.nombre if e.maquina else None,
            'sueldo_basico': str(e.sueldo_basico),
            'bono_por_metraje': str(e.bono_por_metraje),
            'metraje_base': str(e.metraje_base),
            'bonificacion_area': str(e.bonificacion_area),
        })

    return JsonResponse({'estructuras': data})
